"""
EXTRATOR SICRO — VERSÃO SÊNIOR COMENTADA V16 - INTERFACE COM CANTEIRO PRESERVADO

Objetivo:
    Ler relatórios sintéticos e analíticos do SICRO, localizar composições,
    extrair blocos técnicos e gerar um Excel consolidado com as abas atuais.

Fluxo preservado:
    E -> F -> D -> E -> F

Princípio desta revisão:
    Melhorar legibilidade, rastreabilidade, comentários, constantes e robustez
    sem alterar o fluxo operacional do sistema.
"""

import re
import unicodedata
from pathlib import Path
from datetime import datetime
import logging
from copy import copy

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ============================================================
# CONFIGURAÇÕES GERAIS DO SISTEMA
# ============================================================

# Quando True, o programa copia as composições analíticas para a aba COMPOSICOES.
# Quando False, o processo fica mais rápido e gera apenas as abas consolidadas.
GERAR_COMPOSICOES = True

# Hora mensal usada no cálculo de efetivo.
# Mantido como constante para evitar números soltos no código.
HORAS_MES_PADRAO = 182.49
FATOR_PICO_EQUIPE = 1.33

# Nomes oficiais dos blocos SICRO usados na leitura do relatório analítico.
BLOCO_EQUIPAMENTOS = "A - EQUIPAMENTOS"
BLOCO_MAO_OBRA = "B - MAO DE OBRA"
BLOCO_MATERIAL = "C - MATERIAL"
BLOCO_ATIVIDADES_AUXILIARES = "D - ATIVIDADES AUXILIARES"
BLOCO_TEMPO_FIXO = "E - TEMPO FIXO"
BLOCO_MOMENTO_TRANSPORTE = "F - MOMENTO DE TRANSPORTE"

# Logs simples para facilitar auditoria e debug da execução.
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger(__name__)

# Mapeia o código normalizado usado internamente para o código original digitado pelo usuário.
# Isso permite manter a lógica de busca sem alterar a saída da aba NAO_ENCONTRADOS.
CODIGOS_ORIGINAIS_POR_NORMALIZADO = {}


def normalizar_texto(texto):
    """
    Normaliza textos para comparação interna do sistema.

    O que faz:
        - remove acentos;
        - converte para maiúsculo;
        - remove espaços extras;
        - evita erro com NaN/None.

    Utilização:
        Comparação de títulos de blocos SICRO,
        cabeçalhos e descrições.
    """
    if texto is None:
        return ""
    try:
        if pd.isna(texto):
            return ""
    except Exception:
        pass
    texto = str(texto).strip().upper()
    texto = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode("ASCII")
    return texto


def normalizar_codigo(valor):
    """
    Normaliza códigos SICRO.

    Exemplo:
        "0407819.0" -> "0407819"

    Regras:
        - remove caracteres não numéricos;
        - remove .0 do Excel;
        - completa zeros à esquerda até 7 dígitos.
    """
    if valor is None:
        return ""
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    codigo = re.sub(r"\D", "", texto)
    if not codigo:
        return ""
    return codigo.zfill(7)


def limpar_valor(valor):
    """
    Remove NaN e valores inválidos vindos do pandas/openpyxl.

    Retorna:
        None quando o valor estiver vazio.
    """
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass
    return valor


def ler_codigos_digitados():
    """
    Lê os códigos digitados manualmente pelo usuário.

    Entrada esperada:
        0407819,1107892,0705322

    Retorna:
        lista de códigos SICRO normalizados sem duplicidade.
    """
    entrada = input("Digite os códigos separados por vírgula: ").strip()
    if not entrada:
        raise ValueError("Nenhum código foi informado.")

    codigos = []
    CODIGOS_ORIGINAIS_POR_NORMALIZADO.clear()

    for item in entrada.split(","):
        codigo_original = str(item).strip()
        codigo_normalizado = normalizar_codigo(codigo_original)
        if not codigo_normalizado:
            continue

        # Mantém o primeiro valor exatamente como o usuário digitou para uso na saída.
        if codigo_normalizado not in CODIGOS_ORIGINAIS_POR_NORMALIZADO:
            CODIGOS_ORIGINAIS_POR_NORMALIZADO[codigo_normalizado] = codigo_original

        codigos.append(codigo_normalizado)

    if not codigos:
        raise ValueError("Nenhum código válido foi informado.")

    vistos = set()
    codigos_unicos = []
    for codigo in codigos:
        if codigo not in vistos:
            vistos.add(codigo)
            codigos_unicos.append(codigo)
    return codigos_unicos


def identificar_coluna_codigo(df):
    """
    Identifica automaticamente a coluna de código do arquivo sintético.

    Estratégia:
        1. procura nomes clássicos:
           CODIGO, COD, CODIGO SICRO etc.
        2. se não encontrar, analisa o padrão dos valores.

    Retorna:
        nome da coluna identificada.
    """
    colunas_originais = list(df.columns)
    colunas_normalizadas = {col: normalizar_texto(col) for col in colunas_originais}
    nomes_prioritarios = {
        "CODIGO", "COD", "CODIGO DO SERVICO", "CODIGO SERVICO", "CODIGO SICRO", "CODIGO_SICRO"
    }
    for col_original, col_normalizada in colunas_normalizadas.items():
        if col_normalizada in nomes_prioritarios:
            return col_original
    for col_original, col_normalizada in colunas_normalizadas.items():
        if "COD" in col_normalizada:
            return col_original
    for col in colunas_originais:
        serie = df[col].dropna().astype(str).head(30)
        if len(serie) == 0:
            continue
        qtd_codigos = sum(bool(re.fullmatch(r"\D*\d{6,8}\D*", v.strip())) for v in serie)
        if qtd_codigos >= max(1, len(serie) // 2):
            return col
    raise ValueError(f"Não foi possível identificar a coluna de código. Colunas encontradas: {colunas_originais}")


def carregar_itens_sintetico(arquivo_sintetico, codigos_desejados):
    """
    Carrega o relatório sintético e filtra apenas os códigos desejados.

    Retorna:
        DataFrame filtrado mantendo a ordem digitada pelo usuário.
    """
    df = pd.read_excel(arquivo_sintetico)
    if df.empty:
        raise ValueError("O arquivo sintético está vazio.")
    coluna_codigo = identificar_coluna_codigo(df)
    df = df.copy()
    df["CODIGO_NORMALIZADO"] = df[coluna_codigo].apply(normalizar_codigo)
    df_filtrado = df[df["CODIGO_NORMALIZADO"].isin(codigos_desejados)].copy()
    if df_filtrado.empty:
        return pd.DataFrame()
    colunas_finais = [c for c in df_filtrado.columns if c != "CODIGO_NORMALIZADO"]
    df_filtrado = df_filtrado[colunas_finais].copy()
    ordem = {codigo: i for i, codigo in enumerate(codigos_desejados)}
    df_filtrado["_ordem"] = df_filtrado[coluna_codigo].apply(normalizar_codigo).map(ordem)
    df_filtrado = df_filtrado.sort_values("_ordem").drop(columns="_ordem")
    return df_filtrado


def carregar_analitico_em_memoria(arquivo_analitico):
    """
    Carrega o relatório analítico inteiro em memória usando pandas.

    Retorna:
        dados: matriz em memória, onde cada linha do Excel vira uma lista.
        max_col: quantidade máxima de colunas detectadas.

    Benefício:
        Evita acesso lento célula a célula via openpyxl durante as buscas.
    """
    df = pd.read_excel(arquivo_analitico, sheet_name=0, header=None, dtype=object)
    df = df.where(pd.notna(df), None)
    dados = df.values.tolist()
    return dados, df.shape[1]


def valor_celula(dados, row, col=1):
    """
    Lê uma célula da matriz em memória simulando a indexação do Excel.

    row e col começam em 1, igual ao Excel.
    Retorna None quando a posição não existe.
    """
    if row < 1 or col < 1:
        return None
    r = row - 1
    c = col - 1
    if r >= len(dados) or c >= len(dados[r]):
        return None
    return limpar_valor(dados[r][c])


def valores_linha(dados, row, max_col=None):
    if row < 1 or row > len(dados):
        return []
    linha = dados[row - 1]
    if max_col is None:
        max_col = len(linha)
    return [limpar_valor(v) for v in linha[:max_col]]


def obter_texto_dados(dados, row, col=1):
    valor = valor_celula(dados, row, col)
    if valor is None:
        return ""
    return str(valor).strip()


def texto_linha(dados, row, max_col=None):
    return normalizar_texto(" ".join(str(v or "") for v in valores_linha(dados, row, max_col)))


def eh_linha_inicio_composicao(dados, row_idx):
    codigo = normalizar_codigo(valor_celula(dados, row_idx, 1))
    if not re.fullmatch(r"\d{7}", codigo):
        return False
    linha_anterior_1 = normalizar_texto(obter_texto_dados(dados, row_idx - 1, 1))
    linha_anterior_2 = normalizar_texto(obter_texto_dados(dados, row_idx - 2, 1))
    if "SISTEMA DE CUSTOS REFERENCIAIS DE OBRAS - SICRO" in linha_anterior_2:
        return True
    if "CUSTO UNITARIO DE REFERENCIA" in linha_anterior_1:
        return True
    return False


def linha_inicio_real_composicao(dados, row_idx):
    if row_idx >= 3:
        linha_menos_2 = normalizar_texto(obter_texto_dados(dados, row_idx - 2, 1))
        if "SISTEMA DE CUSTOS REFERENCIAIS DE OBRAS - SICRO" in linha_menos_2:
            return row_idx - 2
    if row_idx >= 2:
        linha_menos_1 = normalizar_texto(obter_texto_dados(dados, row_idx - 1, 1))
        if "CUSTO UNITARIO DE REFERENCIA" in linha_menos_1:
            return row_idx - 1
    return row_idx


def localizar_blocos_composicoes(dados):
    """
    Localiza todas as composições existentes no relatório analítico.

    Resultado:
        dicionário:
            codigo -> (linha_inicio, linha_fim)

    Isso permite acessar rapidamente qualquer composição
    sem precisar varrer o Excel inteiro novamente.
    """
    blocos = {}
    max_row = len(dados)
    inicios = []
    for r in range(1, max_row + 1):
        if eh_linha_inicio_composicao(dados, r):
            codigo = normalizar_codigo(valor_celula(dados, r, 1))
            linha_inicio_real = linha_inicio_real_composicao(dados, r)
            inicios.append((codigo, linha_inicio_real, r))
    for i, (codigo, linha_inicio_real, linha_codigo) in enumerate(inicios):
        if i < len(inicios) - 1:
            linha_fim = inicios[i + 1][1] - 1
        else:
            linha_fim = max_row
        blocos[codigo] = (linha_inicio_real, linha_fim)
    return blocos


def eh_cabecalho_secao(texto):
    texto = normalizar_texto(texto)
    return bool(re.match(r"^[A-Z]\s*-\s*", texto))


def localizar_dados_composicao(dados, linha_inicio, linha_fim, codigo_composicao):
    for r in range(linha_inicio, linha_fim + 1):
        codigo_linha = normalizar_codigo(valor_celula(dados, r, 1))
        if codigo_linha == codigo_composicao:
            return r, valor_celula(dados, r, 2)
    return None, ""


def localizar_producao_equipe_composicao(dados, linha_inicio, linha_fim, max_col):
    """
    Localiza, dentro do bloco da composição, a célula com o texto
    "Produção da equipe" e retorna o primeiro valor encontrado à direita.

    No relatório analítico SICRO observado:
    - coluna G: Produção da equipe
    - coluna H: valor da produção
    - coluna I: unidade
    """
    for r in range(linha_inicio, linha_fim + 1):
        linha = valores_linha(dados, r, max_col)

        for idx, valor in enumerate(linha):
            if "PRODUCAO DA EQUIPE" in normalizar_texto(valor):
                for valor_direita in linha[idx + 1:]:
                    valor_limpo = limpar_valor(valor_direita)
                    if valor_limpo is not None and str(valor_limpo).strip() != "":
                        return valor_limpo

    return None


def mapear_producao_equipe_por_codigo(dados, max_col, blocos):
    producao_por_codigo = {}

    for codigo_composicao, (linha_inicio, linha_fim) in blocos.items():
        producao = localizar_producao_equipe_composicao(
            dados=dados,
            linha_inicio=linha_inicio,
            linha_fim=linha_fim,
            max_col=max_col
        )
        producao_por_codigo[codigo_composicao] = producao

    return producao_por_codigo


def adicionar_producao_equipe_ao_sintetico(df_sintetico, producao_por_codigo):
    if df_sintetico.empty:
        return df_sintetico

    df = df_sintetico.copy()
    coluna_codigo = identificar_coluna_codigo(df)

    df["Produção Equipe"] = df[coluna_codigo].apply(
        lambda valor: producao_por_codigo.get(normalizar_codigo(valor))
    )

    return df


def extrair_itens_equipamentos(dados, max_col, blocos, codigos_desejados):
    """
    Extrai itens do bloco:
        A - EQUIPAMENTOS

    Retorna:
        DataFrame contendo:
            - código da composição;
            - código do equipamento;
            - descrição do equipamento.
    """
    registros = []
    vistos = set()
    for codigo_composicao in codigos_desejados:
        if codigo_composicao not in blocos:
            continue
        linha_inicio, linha_fim = blocos[codigo_composicao]
        _, descricao_composicao = localizar_dados_composicao(dados, linha_inicio, linha_fim, codigo_composicao)
        dentro_equipamentos = False
        for r in range(linha_inicio, linha_fim + 1):
            texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
            texto = texto_linha(dados, r, max_col)
            if texto_coluna_a == "A - EQUIPAMENTOS" or "A - EQUIPAMENTOS" in texto:
                dentro_equipamentos = True
                continue
            if not dentro_equipamentos:
                continue
            if "CUSTO UNITARIO TOTAL DE EQUIPAMENTOS" in texto or eh_cabecalho_secao(texto_coluna_a):
                break
            linha_valores = valores_linha(dados, r, max_col)
            if not any(v is not None and str(v).strip() != "" for v in linha_valores):
                continue
            codigo_item = valor_celula(dados, r, 1)
            descricao_item = valor_celula(dados, r, 2)
            codigo_norm = normalizar_codigo(codigo_item)
            descricao_txt = str(descricao_item or "").strip()
            if not codigo_norm and not descricao_txt:
                continue
            chave = (codigo_composicao, codigo_norm, descricao_txt.upper())
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append({
                "CODIGO_COMPOSICAO": codigo_composicao,
                "DESCRICAO_COMPOSICAO": descricao_composicao,
                "CODIGO_ITEM": codigo_item,
                "DESCRICAO_ITEM_EQUIPAMENTO": descricao_item,
            })
    return pd.DataFrame(registros)



def converter_numero(valor):
    """
    Converte valores vindos do Excel para float de forma robusta.

    Aceita:
        - None / NaN;
        - números int/float;
        - textos com vírgula decimal, como "4,22034";
        - textos com ponto decimal, como "4.22034";
        - textos com milhar e decimal, como "1.234,56".

    Retorna:
        float: número convertido ou 0.0 quando não for possível converter.

    Uso no sistema:
        Esta função alimenta cálculos de produção, mão de obra,
        coeficientes de atividades auxiliares e quantidades dos blocos.
    """
    if valor is None:
        return 0.0

    try:
        if pd.isna(valor):
            return 0.0
    except Exception:
        pass

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return 0.0

    texto = texto.replace("\xa0", "").replace(" ", "")

    # Formato brasileiro: 1.234,56
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    # Decimal com vírgula: 4,22034
    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return 0.0

def localizar_quantidade_atividade_auxiliar_na_linha(linha, idx_codigo):
    """
    Tenta localizar a quantidade/coeficiente da atividade auxiliar na própria linha do bloco D.

    No relatório SICRO, o bloco D costuma vir como:
    código | descrição | unidade | quantidade | custo unitário | custo total

    Como o layout pode variar um pouco, a regra abaixo pega o primeiro número útil
    depois da descrição/unidade, evitando usar o próprio código como quantidade.
    """
    for valor in linha[idx_codigo + 2: idx_codigo + 8]:
        numero = converter_numero(valor)
        if numero != 0:
            return numero
    return 0.0


def extrair_itens_mao_obra(dados, max_col, blocos, codigos_desejados):
    """
    Extrai itens do bloco:
        B - MÃO DE OBRA

    Retorna:
        Quantidade horária da equipe necessária
        para produzir 1 hora da patrulha SICRO.
    """
    registros = []
    for codigo_composicao in codigos_desejados:
        if codigo_composicao not in blocos:
            continue
        linha_inicio, linha_fim = blocos[codigo_composicao]
        _, descricao_composicao = localizar_dados_composicao(dados, linha_inicio, linha_fim, codigo_composicao)
        dentro_mao_obra = False
        for r in range(linha_inicio, linha_fim + 1):
            texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
            texto = texto_linha(dados, r, max_col)

            if texto_coluna_a == "B - MAO DE OBRA" or "B - MAO DE OBRA" in texto:
                dentro_mao_obra = True
                continue

            if not dentro_mao_obra:
                continue

            if "CUSTO UNITARIO TOTAL DE MAO DE OBRA" in texto or eh_cabecalho_secao(texto_coluna_a):
                break

            linha_valores = valores_linha(dados, r, max_col)
            if not any(v is not None and str(v).strip() != "" for v in linha_valores):
                continue

            codigo_item = valor_celula(dados, r, 1)
            descricao_item = valor_celula(dados, r, 2)
            codigo_norm = normalizar_codigo(codigo_item)
            descricao_txt = str(descricao_item or "").strip()

            if not codigo_norm and not descricao_txt:
                continue

            registros.append({
                "CODIGO_COMPOSICAO": codigo_composicao,
                "DESCRICAO_COMPOSICAO": descricao_composicao,
                "CODIGO_ITEM": codigo_item,
                "DESCRICAO_ITEM_MAO_OBRA": descricao_item,
                "QUANTIDADE": converter_numero(valor_celula(dados, r, 3)),
            })

    return pd.DataFrame(registros)
def extrair_itens_tempo_fixo(dados, max_col, blocos, codigos_desejados):
    """
    Extrai itens do bloco:
        E - TEMPO FIXO

    Esses itens são utilizados para:
        - carga;
        - descarga;
        - movimentação;
        - logística SICRO.
    """
    registros = []
    vistos = set()
    for codigo_composicao in codigos_desejados:
        if codigo_composicao not in blocos:
            continue
        linha_inicio, linha_fim = blocos[codigo_composicao]
        _, descricao_composicao = localizar_dados_composicao(dados, linha_inicio, linha_fim, codigo_composicao)
        dentro_tempo_fixo = False
        for r in range(linha_inicio, linha_fim + 1):
            texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
            texto = texto_linha(dados, r, max_col)
            if texto_coluna_a == "E - TEMPO FIXO" or "E - TEMPO FIXO" in texto:
                dentro_tempo_fixo = True
                continue
            if not dentro_tempo_fixo:
                continue
            if "CUSTO UNITARIO TOTAL DE TEMPO FIXO" in texto or eh_cabecalho_secao(texto_coluna_a):
                break
            linha_valores = valores_linha(dados, r, max_col)
            if not any(v is not None and str(v).strip() != "" for v in linha_valores):
                continue
            codigo_item = valor_celula(dados, r, 1)
            descricao_item = valor_celula(dados, r, 2)
            codigo_transporte = valor_celula(dados, r, 3)
            chave = (
                codigo_composicao,
                normalizar_codigo(codigo_item),
                str(descricao_item or "").strip(),
                normalizar_codigo(codigo_transporte),
                str(valor_celula(dados, r, 4) or "").strip(),
                str(valor_celula(dados, r, 5) or "").strip(),
            )
            if chave in vistos:
                continue
            vistos.add(chave)
            registros.append({
                "CODIGO_COMPOSICAO": codigo_composicao,
                "DESCRICAO_COMPOSICAO": descricao_composicao,
                "CODIGO_ITEM": codigo_item,
                "DESCRICAO_ITEM_TEMPO_FIXO": descricao_item,
                "CODIGO_TRANSPORTE": codigo_transporte,
                "QUANTIDADE": valor_celula(dados, r, 4),
                "UNIDADE": valor_celula(dados, r, 5),
                "CUSTO_UNITARIO": valor_celula(dados, r, 7),
                "CUSTO_UNITARIO_TOTAL": valor_celula(dados, r, 9),
            })
    return pd.DataFrame(registros)


def extrair_itens_momento_transporte(dados, max_col, blocos, codigos_desejados):
    """
    Extrai itens do bloco:
        F - MOMENTO DE TRANSPORTE

    Utilizado para:
        - DMT;
        - transporte LN/RP/P;
        - custos logísticos.
    """
    registros = []
    vistos = set()
    for codigo_composicao in codigos_desejados:
        if codigo_composicao not in blocos:
            continue
        linha_inicio, linha_fim = blocos[codigo_composicao]
        _, descricao_composicao = localizar_dados_composicao(dados, linha_inicio, linha_fim, codigo_composicao)
        dentro_momento_transporte = False
        for r in range(linha_inicio, linha_fim + 1):
            texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
            texto = texto_linha(dados, r, max_col)
            if texto_coluna_a == "F - MOMENTO DE TRANSPORTE" or "F - MOMENTO DE TRANSPORTE" in texto:
                dentro_momento_transporte = True
                continue
            if not dentro_momento_transporte:
                continue
            if (
                "CUSTO UNITARIO TOTAL DE MOMENTO DE TRANSPORTE" in texto
                or "CUSTO UNITARIO TOTAL DO MOMENTO DE TRANSPORTE" in texto
                or eh_cabecalho_secao(texto_coluna_a)
            ):
                break
            linha_valores = valores_linha(dados, r, max_col)
            if not any(v is not None and str(v).strip() != "" for v in linha_valores):
                continue
            chave = (codigo_composicao, tuple(str(v or "").strip() for v in linha_valores))
            if chave in vistos:
                continue
            vistos.add(chave)
            registro = {"CODIGO_COMPOSICAO": codigo_composicao, "DESCRICAO_COMPOSICAO": descricao_composicao}
            for idx, valor in enumerate(linha_valores, start=1):
                registro[f"COLUNA_{idx}"] = valor
            registros.append(registro)
    return pd.DataFrame(registros)

def extrair_codigos_de_bloco(dados, max_col, blocos, codigo_composicao, nome_bloco, termos_fim=None):
    if termos_fim is None:
        termos_fim = []
    if codigo_composicao not in blocos:
        return []
    linha_inicio, linha_fim = blocos[codigo_composicao]
    nome_bloco_norm = normalizar_texto(nome_bloco)
    termos_fim_norm = [normalizar_texto(t) for t in termos_fim]
    codigos_encontrados = []
    vistos = set()
    dentro_bloco = False
    for r in range(linha_inicio, linha_fim + 1):
        linha = valores_linha(dados, r, max_col)
        texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
        texto = normalizar_texto(" ".join(str(v or "") for v in linha))
        if nome_bloco_norm in texto:
            dentro_bloco = True
            continue
        if not dentro_bloco:
            continue
        if any(termo in texto for termo in termos_fim_norm):
            break
        if eh_cabecalho_secao(texto_coluna_a):
            break
        for valor in linha:
            codigo = normalizar_codigo(valor)
            if not re.fullmatch(r"\d{7}", codigo):
                continue
            if codigo == codigo_composicao:
                continue
            if codigo not in blocos:
                continue
            if codigo not in vistos:
                vistos.add(codigo)
                codigos_encontrados.append(codigo)
    return codigos_encontrados


def extrair_codigos_atividades_auxiliares(dados, max_col, blocos, codigo_composicao):
    return extrair_codigos_de_bloco(
        dados=dados,
        max_col=max_col,
        blocos=blocos,
        codigo_composicao=codigo_composicao,
        nome_bloco="D - ATIVIDADES AUXILIARES",
        termos_fim=["CUSTO UNITARIO TOTAL DE ATIVIDADES AUXILIARES"]
    )



def extrair_registros_atividades_auxiliares(dados, max_col, blocos, codigo_composicao):
    """
    Extrai relações do bloco:
        D - ATIVIDADES AUXILIARES

    Resultado:
        composição principal -> composição auxiliar

    Também extrai:
        - coeficiente da auxiliar;
        - descrição;
        - vínculo analítico.
    """
    """
    Lê o bloco D - ATIVIDADES AUXILIARES da composição informada e retorna
    todos os códigos auxiliares encontrados, preservando o código original
    exibido no relatório analítico.
    """
    registros = []
    vistos = set()

    if codigo_composicao not in blocos:
        return registros

    linha_inicio, linha_fim = blocos[codigo_composicao]
    _, descricao_composicao = localizar_dados_composicao(
        dados, linha_inicio, linha_fim, codigo_composicao
    )

    dentro_atividades_auxiliares = False

    for r in range(linha_inicio, linha_fim + 1):
        linha = valores_linha(dados, r, max_col)
        texto_coluna_a = normalizar_texto(valor_celula(dados, r, 1))
        texto = normalizar_texto(" ".join(str(v or "") for v in linha))

        if "D - ATIVIDADES AUXILIARES" in texto:
            dentro_atividades_auxiliares = True
            continue

        if not dentro_atividades_auxiliares:
            continue

        if "CUSTO UNITARIO TOTAL DE ATIVIDADES AUXILIARES" in texto:
            break

        if eh_cabecalho_secao(texto_coluna_a):
            break

        if not any(v is not None and str(v).strip() != "" for v in linha):
            continue

        for idx, valor in enumerate(linha):
            codigo_auxiliar_norm = normalizar_codigo(valor)

            if not re.fullmatch(r"\d{7}", codigo_auxiliar_norm):
                continue

            if codigo_auxiliar_norm == codigo_composicao:
                continue

            # Mantém apenas códigos que realmente possuem composição no analítico.
            if codigo_auxiliar_norm not in blocos:
                continue

            codigo_auxiliar_original = limpar_codigo_original(valor)

            descricao_auxiliar = ""
            if idx + 1 < len(linha):
                descricao_auxiliar = linha[idx + 1]
            if descricao_auxiliar is None or str(descricao_auxiliar).strip() == "":
                descricao_auxiliar = valor_celula(dados, r, 2)

            chave = (codigo_composicao, codigo_auxiliar_norm)
            if chave in vistos:
                continue
            vistos.add(chave)

            quantidade_auxiliar = localizar_quantidade_atividade_auxiliar_na_linha(linha, idx)

            registros.append({
                "CODIGO_COMPOSICAO": codigo_composicao,
                "DESCRICAO_COMPOSICAO": descricao_composicao,
                "CODIGO_ATIVIDADE_AUXILIAR": codigo_auxiliar_original,
                "CODIGO_ATIVIDADE_AUXILIAR_NORMALIZADO": codigo_auxiliar_norm,
                "DESCRICAO_ATIVIDADE_AUXILIAR": descricao_auxiliar,
                "QUANTIDADE_ATIVIDADE_AUXILIAR": quantidade_auxiliar,
                "POSSUI_ANALITICO": "SIM" if codigo_auxiliar_norm in blocos else "NAO",
            })

    return registros


def extrair_codigos_tempo_fixo(dados, max_col, blocos, codigo_composicao):
    return extrair_codigos_de_bloco(
        dados=dados,
        max_col=max_col,
        blocos=blocos,
        codigo_composicao=codigo_composicao,
        nome_bloco="E - TEMPO FIXO",
        termos_fim=["CUSTO UNITARIO TOTAL DE TEMPO FIXO"]
    )


def extrair_codigos_momento_transporte(dados, max_col, blocos, codigo_composicao):
    return extrair_codigos_de_bloco(
        dados=dados,
        max_col=max_col,
        blocos=blocos,
        codigo_composicao=codigo_composicao,
        nome_bloco="F - MOMENTO DE TRANSPORTE",
        termos_fim=["CUSTO UNITARIO TOTAL DE MOMENTO DE TRANSPORTE", "CUSTO UNITARIO TOTAL DO MOMENTO DE TRANSPORTE"]
    )


def aplicar_formatacao_basica(ws):
    """
    Aplica formatação visual na aba COMPOSICOES.

    Formata:
        - títulos;
        - cabeçalhos;
        - seções SICRO;
        - bordas;
        - alinhamentos.
    """
    fill_titulo = PatternFill("solid", fgColor="D9EAF7")
    fill_secao = PatternFill("solid", fgColor="EAF4EA")
    fill_header = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        if not row:
            continue
        texto_a = normalizar_texto(row[0].value)
        tem_valor = any(cell.value is not None and str(cell.value).strip() != "" for cell in row)
        if not tem_valor:
            continue
        for cell in row:
            cell.alignment = Alignment(vertical="center")
        if "SISTEMA DE CUSTOS REFERENCIAIS DE OBRAS - SICRO" in texto_a or "CUSTO UNITARIO DE REFERENCIA" in texto_a:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = fill_titulo
                cell.border = border
            continue
        if eh_cabecalho_secao(texto_a):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = fill_secao
                cell.border = border
            continue
        texto_atual = normalizar_texto(" ".join(str(cell.value or "") for cell in row))
        if "CODIGO" in texto_atual and "DESCR" in texto_atual:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = fill_header
                cell.border = border
            continue
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border


def classificar_codigo_nao_encontrado(codigo):
    """Classifica códigos não encontrados para facilitar auditoria do orçamento."""
    codigo_txt = limpar_codigo_original(codigo)
    codigo_norm = normalizar_codigo(codigo_txt)

    if re.search(r"[A-Za-z]", codigo_txt):
        return "CODIGO_MANUAL_OU_PROJETO", "Código não localizado no relatório analítico SICRO. Pode representar item manual, MOB, ADM, BET ou código interno do orçamento."

    if re.fullmatch(r"\d{7}", codigo_norm):
        return "SICRO_NAO_LOCALIZADO", "Código com padrão SICRO, mas não encontrado nos relatórios informados. Verificar referência, estado, mês/ano e se o código existe no analítico."

    return "NAO_CLASSIFICADO", "Código não reconhecido automaticamente. Verificar digitação e origem do item."


def criar_aba_nao_encontrados(wb_destino, codigos_nao_encontrados):
    """Cria a aba NAO_ENCONTRADOS com classificação para facilitar auditoria."""
    ws = wb_destino.create_sheet("NAO_ENCONTRADOS")
    headers = ["CODIGO_NAO_ENCONTRADO", "TIPO_PROVAVEL", "OBSERVACAO"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = header
        ws.cell(1, col_idx).font = Font(bold=True)
        ws.cell(1, col_idx).fill = PatternFill("solid", fgColor="F2F2F2")

    for i, codigo in enumerate(codigos_nao_encontrados, start=2):
        codigo_original = CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(codigo, codigo)
        tipo, observacao = classificar_codigo_nao_encontrado(codigo_original)
        ws.cell(i, 1).value = codigo_original
        ws.cell(i, 2).value = tipo
        ws.cell(i, 3).value = observacao

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 90


def escrever_dataframe_em_aba(wb, nome_aba, df):
    ws = wb.create_sheet(nome_aba)
    if df.empty:
        ws["A1"] = "Nenhum dado encontrado."
        return ws
    for c_idx, coluna in enumerate(df.columns, start=1):
        ws.cell(row=1, column=c_idx, value=coluna)
    for r_idx, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c_idx, valor in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=limpar_valor(valor))
    fill_header = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
    for idx, coluna in enumerate(df.columns, start=1):
        valores = [len(str(coluna))]
        valores.extend(len(str(v)) for v in df[coluna].fillna("").astype(str).head(500))
        largura = min(max(valores) + 2, 60)
        ws.column_dimensions[get_column_letter(idx)].width = largura
    return ws


def copiar_bloco_dados(dados, max_col, ws_destino, linha_inicio, linha_fim, linhas_em_branco=2):
    for r in range(linha_inicio, linha_fim + 1):
        ws_destino.append(valores_linha(dados, r, max_col))
    for _ in range(linhas_em_branco):
        ws_destino.append([])


def ajustar_larguras_composicoes(ws, max_col):
    for c in range(1, max_col + 1):
        letra = get_column_letter(c)
        if c == 1:
            ws.column_dimensions[letra].width = 16
        elif c == 2:
            ws.column_dimensions[letra].width = 55
        else:
            ws.column_dimensions[letra].width = 14


def classificar_equipamento_para_mobilizacao(descricao):
    """Classificação preliminar de equipamentos para apoiar a futura aba MOB/DESMOB."""
    texto = normalizar_texto(descricao)

    termos_grande_porte = [
        "CAMINHAO", "CAMINHÃO", "TRATOR", "ESCAVADEIRA", "RETROESCAVADEIRA",
        "MOTONIVELADORA", "ROLO", "PA CARREGADEIRA", "PÁ CARREGADEIRA",
        "CARREGADEIRA", "VIBROACABADORA", "USINA", "GUINDASTE", "MUNCK",
        "COMPRESSOR", "PERFURATRIZ", "BETONEIRA", "DISTRIBUIDOR", "ESPARGIDOR",
        "ESPARIGIDOR", "ACABADORA", "FRESADORA", "RECICLADORA", "PAVIMENTADORA",
        "TANQUE", "BASCULANTE", "CARROCERIA", "CAVALO MECANICO", "CAVALO MECÂNICO",
        "PRANCHA", "SEMI-REBOQUE", "SEMI REBOQUE", "GRADE DE DISCOS", "PULVIMISTURADOR",
        "MISTURADOR", "BRITADOR", "CENTRAL", "DRAGA", "BATE-ESTACA", "BATE ESTACA"
    ]

    termos_pequeno_porte = [
        "FERRAMENTA", "VIBRADOR", "MARTELO", "SERRA", "BOMBA", "PLACA VIBRATORIA",
        "PLACA VIBRATÓRIA", "COMPACTADOR MANUAL", "COMPACTADOR DE PERCUSSAO",
        "COMPACTADOR DE PERCUSSÃO", "NIVEL", "NÍVEL", "TEODOLITO", "GERADOR PORTATIL",
        "GERADOR PORTÁTIL", "MOTOSSERRA", "CORTADORA", "FURADEIRA"
    ]

    if any(t in texto for t in termos_grande_porte):
        return "GRANDE_PORTE", "SIM", "Classificação automática preliminar. Confirmar antes da mobilização."
    if any(t in texto for t in termos_pequeno_porte):
        return "PEQUENO_PORTE", "AVALIAR", "Classificação automática preliminar. Pode ser mobilizado junto com equipe ou ferramentas."
    return "A_CLASSIFICAR", "AVALIAR", "Equipamento sem regra automática. Classificar manualmente."


def criar_dataframe_equipamentos(registros_equipamentos):
    """
    Consolida os equipamentos encontrados em uma visão limpa para uso diário.

    Mantém na aba visível EQUIPAMENTOS apenas as informações essenciais:
        - código;
        - descrição;
        - quantidade de ocorrências;
        - quantidade de composições;
        - indicação preliminar de mobilização.

    A rastreabilidade completa, incluindo composições onde aparece,
    classificação preliminar e observações, permanece na aba técnica oculta
    EQUIPAMENTOS_POR_COMPOSICAO.
    """
    consolidados = {}

    for item in registros_equipamentos:
        codigo = item.get("CODIGO_ITEM")
        descricao = item.get("DESCRICAO_ITEM_EQUIPAMENTO")
        codigo_original = limpar_codigo_original(codigo)
        codigo_norm = normalizar_codigo(codigo)
        descricao_txt = str(descricao or "").strip()

        if not codigo_original and not descricao_txt:
            continue

        chave = (codigo_norm if codigo_norm else normalizar_texto(codigo_original), normalizar_texto(descricao_txt))
        if chave not in consolidados:
            _, necessita_mob, _ = classificar_equipamento_para_mobilizacao(descricao_txt)
            consolidados[chave] = {
                "CODIGO": codigo_original,
                "DESCRICAO": descricao_txt,
                "OCORRENCIAS": 0,
                "COMPOSICOES": set(),
                "NECESSITA_MOBILIZACAO": necessita_mob,
            }

        consolidados[chave]["OCORRENCIAS"] += 1
        codigo_comp = normalizar_codigo(item.get("CODIGO_COMPOSICAO"))
        if codigo_comp:
            consolidados[chave]["COMPOSICOES"].add(codigo_comp)

    registros_saida = []
    for item in consolidados.values():
        # A aba EQUIPAMENTOS fica propositalmente simples e gerencial.
        # A rastreabilidade, classificação e indicação de mobilização permanecem
        # na aba técnica EQUIPAMENTOS_POR_COMPOSICAO.
        registros_saida.append({
            "CODIGO": item["CODIGO"],
            "DESCRICAO": item["DESCRICAO"],
            "OCORRENCIAS": item["OCORRENCIAS"],
        })

    df = pd.DataFrame(registros_saida)
    if not df.empty and "OCORRENCIAS" in df.columns:
        df = df.sort_values(["OCORRENCIAS", "DESCRICAO"], ascending=[False, True]).reset_index(drop=True)
    return df


def criar_dataframe_equipamentos_por_composicao(registros_equipamentos):
    """
    Cria uma aba técnica de rastreabilidade dos equipamentos.

    Objetivo:
        Mostrar de qual composição cada equipamento foi extraído.

    Essa aba deve ficar oculta na limpeza visual final, mas é importante para
    auditoria, conferência de mobilização/desmobilização e rastreabilidade.
    """
    registros_saida = []
    vistos = set()

    for item in registros_equipamentos:
        codigo_comp = normalizar_codigo(item.get("CODIGO_COMPOSICAO"))
        descricao_comp = str(item.get("DESCRICAO_COMPOSICAO") or "").strip()
        codigo_equip = limpar_codigo_original(item.get("CODIGO_ITEM"))
        descricao_equip = str(item.get("DESCRICAO_ITEM_EQUIPAMENTO") or "").strip()

        if not codigo_comp and not codigo_equip and not descricao_equip:
            continue

        chave = (
            codigo_comp,
            normalizar_texto(descricao_comp),
            normalizar_texto(codigo_equip),
            normalizar_texto(descricao_equip),
        )
        if chave in vistos:
            continue
        vistos.add(chave)

        classificacao, necessita_mob, observacao = classificar_equipamento_para_mobilizacao(descricao_equip)

        registros_saida.append({
            "CODIGO_COMPOSICAO": codigo_comp,
            "DESCRICAO_COMPOSICAO": descricao_comp,
            "CODIGO_EQUIPAMENTO": codigo_equip,
            "DESCRICAO_EQUIPAMENTO": descricao_equip,
            "CLASSIFICACAO_PRELIMINAR": classificacao,
            "NECESSITA_MOBILIZACAO": necessita_mob,
            "OBSERVACAO": observacao,
        })

    return pd.DataFrame(registros_saida)


def localizar_coluna_por_termos(df, termos):
    """Localiza uma coluna de um DataFrame por termos no cabeçalho normalizado."""
    termos_norm = [normalizar_texto(t) for t in termos]
    for col in df.columns:
        col_norm = normalizar_texto(col)
        if any(t in col_norm for t in termos_norm):
            return col
    return None


def escrever_aba_entradas_orcamento(wb, df_sintetico, codigos_desejados):
    """
    Cria a aba QUANTIDADES, centralizando quantidade e meses por serviço.

    Essa aba passa a ser a entrada principal para os cálculos de mão de obra,
    evitando que o usuário precise procurar a linha de quantidade dentro da aba PESSOAS.
    """
    ws = wb.create_sheet("QUANTIDADES")

    headers = ["CODIGO", "DESCRICAO", "UNIDADE", "QUANTIDADE", "MESES", "OBSERVACAO"]
    fill_header = PatternFill("solid", fgColor="F2F2F2")
    fill_input = PatternFill("solid", fgColor="FFF2CC")

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx, header)
        ws.cell(1, col_idx).font = Font(bold=True)
        ws.cell(1, col_idx).fill = fill_header

    mapa_sintetico = {}
    if df_sintetico is not None and not df_sintetico.empty:
        try:
            col_codigo = identificar_coluna_codigo(df_sintetico)
        except Exception:
            col_codigo = None

        col_descricao = localizar_coluna_por_termos(df_sintetico, ["DESCRICAO", "DESCRIÇÃO", "SERVICO", "SERVIÇO"])
        col_unidade = localizar_coluna_por_termos(df_sintetico, ["UNIDADE", "UNID"])
        col_quantidade = localizar_coluna_por_termos(df_sintetico, ["QUANTIDADE", "QUANT", "QTD"])

        if col_codigo:
            for _, row in df_sintetico.iterrows():
                codigo_norm = normalizar_codigo(row.get(col_codigo))
                if not codigo_norm:
                    continue
                mapa_sintetico[codigo_norm] = {
                    "DESCRICAO": row.get(col_descricao) if col_descricao else "",
                    "UNIDADE": row.get(col_unidade) if col_unidade else "",
                    "QUANTIDADE": row.get(col_quantidade) if col_quantidade else "",
                }

    for row_idx, codigo in enumerate(codigos_desejados, start=2):
        codigo_norm = normalizar_codigo(codigo)
        codigo_original = CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(codigo_norm, codigo_norm)
        info = mapa_sintetico.get(codigo_norm, {})

        ws.cell(row_idx, 1, codigo_original)
        ws.cell(row_idx, 2, limpar_valor(info.get("DESCRICAO", "")))
        ws.cell(row_idx, 3, limpar_valor(info.get("UNIDADE", "")))
        ws.cell(row_idx, 4, limpar_valor(info.get("QUANTIDADE", "")))
        ws.cell(row_idx, 5, None)
        ws.cell(row_idx, 6, "Preencher/validar QUANTIDADE e MESES para alimentar PESSOAS, MAO_OBRA_CALCULO e HISTOGRAMA_MO.")

        ws.cell(row_idx, 4).fill = fill_input
        ws.cell(row_idx, 5).fill = fill_input

    larguras = {"A": 18, "B": 55, "C": 12, "D": 16, "E": 12, "F": 90}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = "A2"
    return ws



def escrever_aba_pessoas(wb, registros_mao_obra, registros_atividades_auxiliares=None, codigos_tarefas_principais=None, producao_por_codigo=None):
    """
    Cria a aba PESSOAS consolidando mão de obra das composições principais
    e das atividades auxiliares.

    Regra correta aplicada:
    - Para cada composição principal, soma a mão de obra direta do bloco B.
    - Para cada atividade auxiliar do bloco D, multiplica:
        coeficiente da atividade auxiliar no bloco D x (quantidade de mão de obra do bloco B da auxiliar / produção da equipe da auxiliar).
    - Se a auxiliar possuir outra auxiliar, o cálculo é feito de forma recursiva.
    - A tabela final mostra, por composição principal, o coeficiente consolidado de cada função.
    - A linha Quantidade fica para preenchimento manual da quantidade executada da composição principal.
    - As horas totais são calculadas por: coeficiente consolidado x quantidade executada.
    - A linha Meses da obra permite dimensionar o efetivo médio mensal.
    - O resumo final mantém dois indicadores:
        COLABORADORES_TOTAL = horas totais / 182,49;
        COLABORADORES_MES = horas totais / (meses x 182,49).
    """
    ws = wb.create_sheet("PESSOAS")

    if not registros_mao_obra:
        ws["A1"] = "Nenhum dado encontrado."
        return ws, 0

    registros_atividades_auxiliares = registros_atividades_auxiliares or []
    codigos_tarefas_principais = codigos_tarefas_principais or []
    producao_por_codigo = producao_por_codigo or {}

    def obter_producao_equipe(codigo_comp):
        # Retorna a produção da equipe da composição.
        # Se não encontrar, usa 1,0 para manter a mão de obra original e não zerar o cálculo.
        producao = converter_numero(producao_por_codigo.get(normalizar_codigo(codigo_comp)))
        return producao if producao != 0 else 1.0
    codigos_principais = []
    vistos_principais = set()
    for c in codigos_tarefas_principais:
        cn = normalizar_codigo(c)
        if cn and cn not in vistos_principais:
            codigos_principais.append(cn)
            vistos_principais.add(cn)

    # ============================================================
    # 1) Mão de obra direta por composição
    #    coef_direto[CODIGO_COMPOSICAO][PESSOA] = soma(quantidade do bloco B / Produção da equipe)
    # ============================================================
    descricoes_composicoes = {}
    pessoas = {}
    coef_direto = {}

    for item in registros_mao_obra:
        codigo_comp = normalizar_codigo(item.get("CODIGO_COMPOSICAO"))
        descricao_comp = str(item.get("DESCRICAO_COMPOSICAO") or "").strip()
        codigo_pessoa = limpar_codigo_original(item.get("CODIGO_ITEM"))
        descricao_pessoa = str(item.get("DESCRICAO_ITEM_MAO_OBRA") or "").strip()
        quantidade = converter_numero(item.get("QUANTIDADE"))
        producao_equipe = obter_producao_equipe(codigo_comp)
        quantidade_ajustada = quantidade / producao_equipe if producao_equipe != 0 else 0

        if not codigo_comp or not descricao_pessoa:
            continue

        if descricao_comp and codigo_comp not in descricoes_composicoes:
            descricoes_composicoes[codigo_comp] = descricao_comp

        chave_pessoa = (normalizar_texto(codigo_pessoa), normalizar_texto(descricao_pessoa))
        if chave_pessoa not in pessoas:
            pessoas[chave_pessoa] = {"CODIGO": codigo_pessoa, "DESCRICAO": descricao_pessoa}

        coef_direto.setdefault(codigo_comp, {})
        coef_direto[codigo_comp][chave_pessoa] = coef_direto[codigo_comp].get(chave_pessoa, 0.0) + quantidade_ajustada

    # ============================================================
    # 2) Relações do bloco D: composição pai -> auxiliar x coeficiente
    # ============================================================
    relacoes_por_pai = {}

    for aux in registros_atividades_auxiliares:
        codigo_pai = normalizar_codigo(aux.get("CODIGO_COMPOSICAO"))
        codigo_aux = normalizar_codigo(aux.get("CODIGO_ATIVIDADE_AUXILIAR_NORMALIZADO") or aux.get("CODIGO_ATIVIDADE_AUXILIAR"))
        qtd_aux = converter_numero(aux.get("QUANTIDADE_ATIVIDADE_AUXILIAR"))

        if not codigo_pai or not codigo_aux:
            continue

        descricao_pai = str(aux.get("DESCRICAO_COMPOSICAO") or "").strip()
        descricao_aux = str(aux.get("DESCRICAO_ATIVIDADE_AUXILIAR") or "").strip()

        if descricao_pai and codigo_pai not in descricoes_composicoes:
            descricoes_composicoes[codigo_pai] = descricao_pai
        if descricao_aux and codigo_aux not in descricoes_composicoes:
            descricoes_composicoes[codigo_aux] = descricao_aux

        # Se o coeficiente vier vazio ou zero, mantém zero para não inventar quantitativo.
        relacoes_por_pai.setdefault(codigo_pai, []).append((codigo_aux, qtd_aux))

    # ============================================================
    # 3) Coeficiente consolidado por composição
    #    consolidado = mão de obra direta ajustada por produção + soma(auxiliar ajustada por produção * coeficiente do bloco D)
    # ============================================================
    cache_consolidado = {}

    def somar_dict(destino, origem, fator=1.0):
        for pessoa, valor in origem.items():
            destino[pessoa] = destino.get(pessoa, 0.0) + (converter_numero(valor) * fator)

    def calcular_coeficiente_consolidado(codigo_comp, pilha=None):
        codigo_comp = normalizar_codigo(codigo_comp)
        if not codigo_comp:
            return {}

        if codigo_comp in cache_consolidado:
            return cache_consolidado[codigo_comp]

        if pilha is None:
            pilha = set()

        # Evita loop infinito se existir referência circular no analítico.
        if codigo_comp in pilha:
            return {}

        pilha.add(codigo_comp)

        resultado = {}

        # Mão de obra direta do próprio bloco B da composição.
        somar_dict(resultado, coef_direto.get(codigo_comp, {}), fator=1.0)

        # Mão de obra indireta via atividades auxiliares do bloco D.
        for codigo_aux, coef_aux in relacoes_por_pai.get(codigo_comp, []):
            if coef_aux == 0:
                continue
            coef_auxiliar_consolidado = calcular_coeficiente_consolidado(codigo_aux, pilha)
            somar_dict(resultado, coef_auxiliar_consolidado, fator=coef_aux)

        pilha.remove(codigo_comp)
        cache_consolidado[codigo_comp] = resultado
        return resultado

    def calcular_coeficiente_complementar(codigo_comp, pilha=None):
        """
        Retorna somente a parcela complementar da composição principal.
        Regra: coeficiente do bloco D x mão de obra consolidada da composição auxiliar.
        A mão de obra consolidada da auxiliar já considera bloco B dividido pela produção da equipe e possíveis D dentro dela.
        """
        codigo_comp = normalizar_codigo(codigo_comp)
        if not codigo_comp:
            return {}

        if pilha is None:
            pilha = set()

        # Evita loop infinito se existir referência circular no analítico.
        if codigo_comp in pilha:
            return {}

        pilha.add(codigo_comp)
        resultado = {}

        for codigo_aux, coef_aux in relacoes_por_pai.get(codigo_comp, []):
            if coef_aux == 0:
                continue
            coef_auxiliar_consolidado = calcular_coeficiente_consolidado(codigo_aux, pilha)
            somar_dict(resultado, coef_auxiliar_consolidado, fator=coef_aux)

        pilha.remove(codigo_comp)
        return resultado

    # Mantém apenas as composições principais como colunas da tabela final.
    tarefas = []
    for codigo in codigos_principais:
        coef_total = calcular_coeficiente_consolidado(codigo)
        if not coef_total and codigo not in descricoes_composicoes:
            continue
        coef_principal = coef_direto.get(codigo, {})
        coef_complementar = calcular_coeficiente_complementar(codigo)
        tarefas.append({
            "CODIGO_COMPOSICAO": codigo,
            "DESCRICAO_COMPOSICAO": descricoes_composicoes.get(codigo, ""),
            "COEFICIENTES": coef_total,
            "COEF_PRINCIPAL": coef_principal,
            "COEF_COMPLEMENTAR": coef_complementar,
        })

    # Segurança: se não recebeu lista de principais, gera colunas para as composições que existem na mão de obra direta.
    if not tarefas:
        for codigo in coef_direto.keys():
            coef_total = calcular_coeficiente_consolidado(codigo)
            coef_principal = coef_direto.get(codigo, {})
            coef_complementar = calcular_coeficiente_complementar(codigo)
            tarefas.append({
                "CODIGO_COMPOSICAO": codigo,
                "DESCRICAO_COMPOSICAO": descricoes_composicoes.get(codigo, ""),
                "COEFICIENTES": coef_total,
                "COEF_PRINCIPAL": coef_principal,
                "COEF_COMPLEMENTAR": coef_complementar,
            })

    # Recalcula a lista de pessoas considerando também as que aparecem apenas dentro de auxiliares consolidadas.
    for tarefa in tarefas:
        for chave_pessoa in tarefa["COEFICIENTES"].keys():
            if chave_pessoa not in pessoas:
                pessoas[chave_pessoa] = {"CODIGO": chave_pessoa[0], "DESCRICAO": chave_pessoa[1]}

    lista_pessoas_chaves = list(pessoas.keys())

    if not tarefas or not lista_pessoas_chaves:
        ws["A1"] = "Nenhum dado encontrado."
        return ws, 0

    fill_titulo = PatternFill("solid", fgColor="D9EAF7")
    fill_header = PatternFill("solid", fgColor="F2F2F2")
    fill_quantidade = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def estilizar_linha(row_idx, fill=None, bold=False):
        for cell in ws[row_idx]:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if fill:
                    cell.fill = fill
                if bold:
                    cell.font = Font(bold=True)

    # ============================================================
    # TABELA 1 - COEFICIENTE CONSOLIDADO POR COMPOSIÇÃO PRINCIPAL
    # ============================================================
    ws.cell(1, 1, "LEVANTAMENTO DE COLABORADORES")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(1, 1).fill = fill_titulo

    ws.cell(2, 1, "CODIGO_COMPOSICAO")
    ws.cell(3, 1, "COLABORADOR / ATIVIDADE")

    for idx, tarefa in enumerate(tarefas, start=2):
        ws.cell(2, idx, CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(tarefa["CODIGO_COMPOSICAO"], tarefa["CODIGO_COMPOSICAO"]))
        ws.cell(3, idx, tarefa["DESCRICAO_COMPOSICAO"])

    estilizar_linha(2, fill_header, True)
    estilizar_linha(3, fill_header, True)

    linha_inicio_coef = 4
    for row_idx, chave_pessoa in enumerate(lista_pessoas_chaves, start=linha_inicio_coef):
        pessoa = pessoas[chave_pessoa]
        ws.cell(row_idx, 1, pessoa["DESCRICAO"])
        for col_idx, tarefa in enumerate(tarefas, start=2):
            valor = tarefa["COEFICIENTES"].get(chave_pessoa, 0.0)
            if valor:
                ws.cell(row_idx, col_idx, valor)
        estilizar_linha(row_idx)

    linha_total = linha_inicio_coef + len(lista_pessoas_chaves)
    ws.cell(linha_total, 1, "TOTAL")
    for col_idx in range(2, len(tarefas) + 2):
        col_letter = get_column_letter(col_idx)
        ws.cell(linha_total, col_idx, f"=SUM({col_letter}{linha_inicio_coef}:{col_letter}{linha_total-1})")
    estilizar_linha(linha_total, fill_header, True)

    # Quantidade e meses são centralizados na aba QUANTIDADES.
    # A aba PESSOAS mantém as linhas abaixo apenas como espelho para facilitar a conferência.
    linha_quantidade = linha_total + 3
    ws.cell(linha_quantidade, 1, "Quantidade")
    for col_idx, tarefa in enumerate(tarefas, start=2):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            linha_quantidade,
            col_idx,
            f'=IFERROR(INDEX(QUANTIDADES!$D:$D,MATCH({col_letter}$2,QUANTIDADES!$A:$A,0)),"")'
        )
    estilizar_linha(linha_quantidade, fill_quantidade, True)

    # Linha de duração por composição/frente de serviço.
    # Ela puxa primeiro a duração específica da aba QUANTIDADES.
    linha_meses = linha_quantidade + 1
    ws.cell(linha_meses, 1, "Meses da obra")
    for col_idx, tarefa in enumerate(tarefas, start=2):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=linha_meses,
            column=col_idx,
            value=f'=IFERROR(INDEX(QUANTIDADES!$E:$E,MATCH({col_letter}$2,QUANTIDADES!$A:$A,0)),"")'
        )
    estilizar_linha(linha_meses, fill_quantidade, True)

    # ============================================================
    # TABELA 2 - HORAS TOTAIS POR COLABORADOR E ATIVIDADE
    # ============================================================
    linha_inicio_horas = linha_meses + 3
    ws.cell(linha_inicio_horas, 1, "HORAS TOTAIS POR COLABORADOR E ATIVIDADE")
    ws.cell(linha_inicio_horas, 1).font = Font(bold=True, size=12)
    ws.cell(linha_inicio_horas, 1).fill = fill_titulo

    linha_header_horas_codigo = linha_inicio_horas + 1
    linha_header_horas_desc = linha_inicio_horas + 2

    ws.cell(linha_header_horas_codigo, 1, "CODIGO_COMPOSICAO")
    ws.cell(linha_header_horas_desc, 1, "COLABORADOR / ATIVIDADE")

    for idx, tarefa in enumerate(tarefas, start=2):
        ws.cell(linha_header_horas_codigo, idx, CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(tarefa["CODIGO_COMPOSICAO"], tarefa["CODIGO_COMPOSICAO"]))
        ws.cell(linha_header_horas_desc, idx, tarefa["DESCRICAO_COMPOSICAO"])

    estilizar_linha(linha_header_horas_codigo, fill_header, True)
    estilizar_linha(linha_header_horas_desc, fill_header, True)

    linha_inicio_horas_dados = linha_header_horas_desc + 1
    linhas_horas_por_pessoa = {}

    for i, chave_pessoa in enumerate(lista_pessoas_chaves):
        row_idx = linha_inicio_horas_dados + i
        linha_coef = linha_inicio_coef + i
        pessoa = pessoas[chave_pessoa]
        ws.cell(row_idx, 1, pessoa["DESCRICAO"])
        linhas_horas_por_pessoa[chave_pessoa] = row_idx

        for col_idx in range(2, len(tarefas) + 2):
            col_letter = get_column_letter(col_idx)
            ws.cell(row_idx, col_idx, f"=IF({col_letter}${linha_quantidade}=\"\",0,{col_letter}${linha_quantidade}*{col_letter}{linha_coef})")

        estilizar_linha(row_idx)

    linha_total_horas = linha_inicio_horas_dados + len(lista_pessoas_chaves)
    ws.cell(linha_total_horas, 1, "TOTAL")
    for col_idx in range(2, len(tarefas) + 2):
        col_letter = get_column_letter(col_idx)
        ws.cell(linha_total_horas, col_idx, f"=SUM({col_letter}{linha_inicio_horas_dados}:{col_letter}{linha_total_horas-1})")
    estilizar_linha(linha_total_horas, fill_header, True)

    # ============================================================
    # TABELA 3 - RESUMO FINAL
    # ============================================================
    linha_resumo = linha_total_horas + 3
    ws.cell(linha_resumo, 1, "RESUMO DE COLABORADORES")
    ws.cell(linha_resumo, 1).font = Font(bold=True, size=12)
    ws.cell(linha_resumo, 1).fill = fill_titulo

    linha_header_resumo = linha_resumo + 1
    headers = [
        "CODIGO",
        "DESCRICAO",
        "UNID",
        "HORAS_ATIVIDADE_PRINCIPAL",
        "HORAS_ATIVIDADE_COMPLEMENTAR",
        "HORAS_TOTAL",
        "COLABORADORES_TOTAL",
        "COLABORADORES_MES",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(linha_header_resumo, col_idx, header)
    estilizar_linha(linha_header_resumo, fill_header, True)

    linha_inicio_resumo_dados = linha_header_resumo + 1

    # ============================================================
    # TABELA INTERMEDIÁRIA - MAO_OBRA_CALCULO
    # ============================================================
    # A partir desta versão, os cálculos detalhados são escritos em uma aba própria.
    # Isso reduz fórmulas gigantes na aba PESSOAS e deixa o orçamento mais auditável.
    ws_calc = wb.create_sheet("MAO_OBRA_CALCULO")
    headers_calc = [
        "CODIGO_COMPOSICAO",
        "DESCRICAO_COMPOSICAO",
        "CODIGO_COLABORADOR",
        "DESCRICAO_COLABORADOR",
        "COEF_TOTAL",
        "COEF_PRINCIPAL",
        "COEF_COMPLEMENTAR",
        "QUANTIDADE",
        "MESES_ENTRADA",
        "MESES_EFETIVO",
        "HORAS_PRINCIPAL",
        "HORAS_COMPLEMENTAR",
        "HORAS_TOTAL",
        "COLABORADORES_TOTAL",
        "COLABORADORES_MES",
        "OBSERVACAO",
    ]
    for col_idx, header in enumerate(headers_calc, start=1):
        ws_calc.cell(1, col_idx, header)
        ws_calc.cell(1, col_idx).font = Font(bold=True)
        ws_calc.cell(1, col_idx).fill = fill_header

    linha_calc = 2
    for tarefa in tarefas:
        codigo_comp_original = CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(
            tarefa["CODIGO_COMPOSICAO"], tarefa["CODIGO_COMPOSICAO"]
        )
        for chave_pessoa in lista_pessoas_chaves:
            pessoa = pessoas[chave_pessoa]
            coef_total = converter_numero(tarefa["COEFICIENTES"].get(chave_pessoa, 0.0))
            coef_principal = converter_numero(tarefa["COEF_PRINCIPAL"].get(chave_pessoa, 0.0))
            coef_complementar = converter_numero(tarefa["COEF_COMPLEMENTAR"].get(chave_pessoa, 0.0))

            if coef_total == 0 and coef_principal == 0 and coef_complementar == 0:
                continue

            ws_calc.cell(linha_calc, 1, codigo_comp_original)
            ws_calc.cell(linha_calc, 2, tarefa["DESCRICAO_COMPOSICAO"])
            ws_calc.cell(linha_calc, 3, pessoa["CODIGO"])
            ws_calc.cell(linha_calc, 4, pessoa["DESCRICAO"])
            ws_calc.cell(linha_calc, 5, coef_total)
            ws_calc.cell(linha_calc, 6, coef_principal)
            ws_calc.cell(linha_calc, 7, coef_complementar)
            ws_calc.cell(linha_calc, 8, f'=IFERROR(INDEX(QUANTIDADES!$D:$D,MATCH(A{linha_calc},QUANTIDADES!$A:$A,0)),"")')
            ws_calc.cell(linha_calc, 9, f'=IFERROR(INDEX(QUANTIDADES!$E:$E,MATCH(A{linha_calc},QUANTIDADES!$A:$A,0)),"")')
            ws_calc.cell(linha_calc, 10, f'=IF(OR(I{linha_calc}="",I{linha_calc}=0),HISTOGRAMA_MO!$B$2,I{linha_calc})')
            ws_calc.cell(linha_calc, 11, f'=IF(H{linha_calc}="",0,H{linha_calc}*F{linha_calc})')
            ws_calc.cell(linha_calc, 12, f'=IF(H{linha_calc}="",0,H{linha_calc}*G{linha_calc})')
            ws_calc.cell(linha_calc, 13, f'=K{linha_calc}+L{linha_calc}')
            ws_calc.cell(linha_calc, 14, f'=IF(M{linha_calc}=0,0,M{linha_calc}/{HORAS_MES_PADRAO})')
            ws_calc.cell(linha_calc, 15, f'=IF(OR(J{linha_calc}="",J{linha_calc}=0),0,M{linha_calc}/(J{linha_calc}*{HORAS_MES_PADRAO}))')
            ws_calc.cell(linha_calc, 16, f'=IF(OR(H{linha_calc}="",H{linha_calc}=0),"Informar quantidade",IF(OR(J{linha_calc}="",J{linha_calc}=0),"Informar meses","OK"))')
            linha_calc += 1

    for col_idx in range(1, len(headers_calc) + 1):
        ws_calc.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_calc.column_dimensions["B"].width = 55
    ws_calc.column_dimensions["D"].width = 40
    ws_calc.column_dimensions["P"].width = 28
    ws_calc.freeze_panes = "A2"

    def formula_somases_calc(coluna_soma, row_idx):
        return (
            f'=SUMIFS(MAO_OBRA_CALCULO!${coluna_soma}:${coluna_soma},'
            f'MAO_OBRA_CALCULO!$C:$C,A{row_idx},'
            f'MAO_OBRA_CALCULO!$D:$D,B{row_idx})'
        )

    for i, chave_pessoa in enumerate(lista_pessoas_chaves):
        row_idx = linha_inicio_resumo_dados + i
        pessoa = pessoas[chave_pessoa]

        ws.cell(row_idx, 1, pessoa["CODIGO"])
        ws.cell(row_idx, 2, pessoa["DESCRICAO"])
        ws.cell(row_idx, 3, "H")
        ws.cell(row_idx, 4, formula_somases_calc("K", row_idx))
        ws.cell(row_idx, 5, formula_somases_calc("L", row_idx))
        ws.cell(row_idx, 6, formula_somases_calc("M", row_idx))
        ws.cell(row_idx, 7, f"=ROUNDUP({formula_somases_calc('N', row_idx)[1:]},0)")
        ws.cell(row_idx, 8, f"=ROUNDUP({formula_somases_calc('O', row_idx)[1:]},0)")
        estilizar_linha(row_idx)

    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = 45
    for col_idx in range(2, len(tarefas) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 18, 40)
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22

    return ws, len(lista_pessoas_chaves)





def aplicar_filtro_e_congelamento(ws, linha_header=1, freeze="A2"):
    """
    Aplica autofiltro e congelamento de painéis.
    """
    try:
        ultima_coluna = get_column_letter(ws.max_column)
        ultima_linha = max(ws.max_row, linha_header)
        ws.auto_filter.ref = f"A{linha_header}:{ultima_coluna}{ultima_linha}"
    except Exception:
        pass

    try:
        ws.freeze_panes = freeze
    except Exception:
        pass


def ocultar_linhas_grade(ws):
    """
    Remove as linhas de grade para melhorar a apresentação visual.
    """
    try:
        ws.sheet_view.showGridLines = False
    except Exception:
        pass

def recriar_aba_pessoas_resumida(wb):
    """
    Substitui a aba PESSOAS detalhada por uma visão executiva e limpa.

    A aba técnica MAO_OBRA_CALCULO continua existindo e guarda todos os detalhes.
    A nova aba PESSOAS mostra apenas o resumo consolidado por colaborador,
    reduzindo a poluição visual causada por uma coluna para cada composição.
    """
    if "MAO_OBRA_CALCULO" not in wb.sheetnames:
        return wb["PESSOAS"] if "PESSOAS" in wb.sheetnames else None

    # Remove a aba PESSOAS antiga, que continha matriz ampla por composição.
    if "PESSOAS" in wb.sheetnames:
        idx = wb.sheetnames.index("PESSOAS")
        wb.remove(wb["PESSOAS"])
    else:
        idx = len(wb.sheetnames)

    ws = wb.create_sheet("PESSOAS", idx)
    ws_calc = wb["MAO_OBRA_CALCULO"]

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_obs = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "RESUMO DE MÃO DE OBRA"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:H1")

    ws["A2"] = "Esta aba é uma visão limpa. O detalhamento por composição está na aba técnica MAO_OBRA_CALCULO."
    ws["A2"].fill = fill_obs
    ws.merge_cells("A2:H2")

    headers = [
        "CODIGO",
        "DESCRICAO",
        "UNID",
        "HORAS_ATIVIDADE_PRINCIPAL",
        "HORAS_ATIVIDADE_COMPLEMENTAR",
        "HORAS_TOTAL",
        "COLABORADORES_TOTAL",
        "COLABORADORES_MES",
    ]
    linha_header = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(linha_header, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Lista única de colaboradores a partir da aba técnica.
    colaboradores = []
    vistos = set()
    for row in range(2, ws_calc.max_row + 1):
        codigo = ws_calc.cell(row, 3).value
        descricao = ws_calc.cell(row, 4).value
        if not descricao:
            continue
        chave = (str(codigo or "").strip(), str(descricao or "").strip())
        if chave in vistos:
            continue
        vistos.add(chave)
        colaboradores.append(chave)

    linha = linha_header + 1
    for codigo, descricao in colaboradores:
        ws.cell(linha, 1, codigo)
        ws.cell(linha, 2, descricao)
        ws.cell(linha, 3, "H")
        ws.cell(linha, 4, f'=SUMIFS(MAO_OBRA_CALCULO!$K:$K,MAO_OBRA_CALCULO!$C:$C,A{linha},MAO_OBRA_CALCULO!$D:$D,B{linha})')
        ws.cell(linha, 5, f'=SUMIFS(MAO_OBRA_CALCULO!$L:$L,MAO_OBRA_CALCULO!$C:$C,A{linha},MAO_OBRA_CALCULO!$D:$D,B{linha})')
        ws.cell(linha, 6, f'=D{linha}+E{linha}')
        ws.cell(linha, 7, f'=ROUNDUP(SUMIFS(MAO_OBRA_CALCULO!$N:$N,MAO_OBRA_CALCULO!$C:$C,A{linha},MAO_OBRA_CALCULO!$D:$D,B{linha}),0)')
        ws.cell(linha, 8, f'=ROUNDUP(SUMIFS(MAO_OBRA_CALCULO!$O:$O,MAO_OBRA_CALCULO!$C:$C,A{linha},MAO_OBRA_CALCULO!$D:$D,B{linha}),0)')
        for col in range(1, 9):
            ws.cell(linha, col).border = border
            ws.cell(linha, col).alignment = Alignment(vertical="center", wrap_text=True)
        linha += 1

    linha_total = linha
    ws.cell(linha_total, 1, "TOTAL")
    ws.cell(linha_total, 1).font = Font(bold=True)
    for col_idx in range(4, 9):
        col_letter = get_column_letter(col_idx)
        ws.cell(linha_total, col_idx, f"=SUM({col_letter}{linha_header+1}:{col_letter}{linha_total-1})")
        ws.cell(linha_total, col_idx).font = Font(bold=True)
    for col in range(1, 9):
        ws.cell(linha_total, col).fill = fill_header
        ws.cell(linha_total, col).border = border

    larguras = {"A": 16, "B": 45, "C": 10, "D": 28, "E": 32, "F": 18, "G": 22, "H": 22}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A5")
    ocultar_linhas_grade(ws)
    return ws



def escrever_aba_histograma_mo(wb, ws_pessoas):
    """
    Cria a aba HISTOGRAMA_MO a partir da aba PESSOAS resumida.

    A partir da V9, a aba PESSOAS é uma visão executiva limpa, com cabeçalho
    na linha 4. Portanto, o histograma não procura mais o texto antigo
    "RESUMO DE COLABORADORES"; ele lê diretamente as colunas consolidadas.
    """
    if "HISTOGRAMA_MO" in wb.sheetnames:
        wb.remove(wb["HISTOGRAMA_MO"])
    ws = wb.create_sheet("HISTOGRAMA_MO")

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def estilizar_linha(row_idx, fill=None, bold=False):
        for cell in ws[row_idx]:
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if fill:
                    cell.fill = fill
                if bold:
                    cell.font = Font(bold=True)

    # A aba PESSOAS resumida possui cabeçalho fixo na linha 4.
    linha_header_pessoas = 4
    if ws_pessoas.max_row < linha_header_pessoas:
        ws["A1"] = "Aba PESSOAS sem dados suficientes para gerar o histograma."
        return ws

    colunas_pessoas = {}
    for col in range(1, ws_pessoas.max_column + 1):
        header = normalizar_texto(ws_pessoas.cell(row=linha_header_pessoas, column=col).value)
        if header:
            colunas_pessoas[header] = col

    col_codigo = colunas_pessoas.get("CODIGO", 1)
    col_descricao = colunas_pessoas.get("DESCRICAO", 2)
    col_unid = colunas_pessoas.get("UNID", 3)
    col_horas_total = colunas_pessoas.get("HORAS_TOTAL", 6)
    col_colab_total = colunas_pessoas.get("COLABORADORES_TOTAL", 7)
    col_colab_mes = colunas_pessoas.get("COLABORADORES_MES", 8)

    ws.cell(1, 1, "HISTOGRAMA DE MÃO DE OBRA")
    ws.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = fill_titulo
    ws.merge_cells("A1:I1")

    ws.cell(2, 1, "MESES_PADRAO_OBRA")
    ws.cell(2, 2, None)
    ws.cell(3, 1, "HORAS_MES_PADRAO")
    ws.cell(3, 2, HORAS_MES_PADRAO)
    ws.cell(4, 1, "OBSERVACAO")
    ws.cell(4, 2, "Se os meses por serviço estiverem vazios na aba QUANTIDADES, preencha um prazo padrão em B2.")
    estilizar_linha(2, fill_input, True)
    estilizar_linha(3, fill_header, True)
    estilizar_linha(4, fill_header, False)

    linha_header = 6
    headers = [
        "CODIGO",
        "DESCRICAO",
        "UNID",
        "HORAS_TOTAL",
        "HORAS_MES",
        "COLABORADORES_MES_HISTOGRAMA",
        "COLABORADORES_TOTAL",
        "COLABORADORES_MES_PESSOAS",
        "OBSERVACAO",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(linha_header, col_idx, header)
    estilizar_linha(linha_header, fill_header, True)

    linha_saida = linha_header + 1
    linha_inicio_pessoas = linha_header_pessoas + 1

    for row_pessoas in range(linha_inicio_pessoas, ws_pessoas.max_row + 1):
        marcador = normalizar_texto(ws_pessoas.cell(row=row_pessoas, column=1).value)
        if marcador == "TOTAL":
            break

        codigo = ws_pessoas.cell(row=row_pessoas, column=col_codigo).value
        descricao = ws_pessoas.cell(row=row_pessoas, column=col_descricao).value
        if (codigo is None or str(codigo).strip() == "") and (descricao is None or str(descricao).strip() == ""):
            continue

        ws.cell(linha_saida, 1, f"=PESSOAS!{get_column_letter(col_codigo)}{row_pessoas}")
        ws.cell(linha_saida, 2, f"=PESSOAS!{get_column_letter(col_descricao)}{row_pessoas}")
        ws.cell(linha_saida, 3, f"=PESSOAS!{get_column_letter(col_unid)}{row_pessoas}")
        ws.cell(linha_saida, 4, f"=PESSOAS!{get_column_letter(col_horas_total)}{row_pessoas}")
        ws.cell(linha_saida, 5, f'=IF($B$2="",0,D{linha_saida}/$B$2)')
        ws.cell(linha_saida, 6, f'=IF($B$2="",PESSOAS!{get_column_letter(col_colab_mes)}{row_pessoas},ROUNDUP(E{linha_saida}/$B$3,0))')
        ws.cell(linha_saida, 7, f"=PESSOAS!{get_column_letter(col_colab_total)}{row_pessoas}")
        ws.cell(linha_saida, 8, f"=PESSOAS!{get_column_letter(col_colab_mes)}{row_pessoas}")
        ws.cell(linha_saida, 9, f'=IF($B$2="","Usando meses por serviço da aba QUANTIDADES","Usando MESES_PADRAO_OBRA")')
        estilizar_linha(linha_saida)
        linha_saida += 1

    linha_total = linha_saida
    ws.cell(linha_total, 1, "TOTAL")
    for col_idx in [4, 5, 6, 7, 8]:
        col_letter = get_column_letter(col_idx)
        ws.cell(linha_total, col_idx, f"=SUM({col_letter}{linha_header+1}:{col_letter}{linha_total-1})")
    estilizar_linha(linha_total, fill_header, True)

    larguras = {"A": 16, "B": 45, "C": 10, "D": 18, "E": 18, "F": 28, "G": 22, "H": 26, "I": 45}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A7")
    ocultar_linhas_grade(ws)
    return ws

def criar_aba_alertas(wb, codigos_nao_encontrados=None):
    """
    Cria a aba ALERTAS contendo apenas os códigos que não foram encontrados
    durante a extração, conforme decisão de simplificação da V9.
    """
    codigos_nao_encontrados = codigos_nao_encontrados or []

    if "ALERTAS" in wb.sheetnames:
        wb.remove(wb["ALERTAS"])
    if "NAO_ENCONTRADOS" in wb.sheetnames:
        wb.remove(wb["NAO_ENCONTRADOS"])

    ws = wb.create_sheet("ALERTAS")

    fill_titulo = PatternFill("solid", fgColor="C00000")
    fill_header = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "CÓDIGOS NÃO ENCONTRADOS"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:A1")

    ws.cell(3, 1, "CODIGO")
    ws.cell(3, 1).font = Font(bold=True)
    ws.cell(3, 1).fill = fill_header
    ws.cell(3, 1).border = border
    ws.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center")

    linha = 4
    codigos_unicos = []
    vistos = set()
    for codigo in codigos_nao_encontrados:
        codigo_original = CODIGOS_ORIGINAIS_POR_NORMALIZADO.get(normalizar_codigo(codigo), codigo)
        codigo_txt = limpar_codigo_original(codigo_original)
        if not codigo_txt or codigo_txt in vistos:
            continue
        vistos.add(codigo_txt)
        codigos_unicos.append(codigo_txt)

    if codigos_unicos:
        for codigo in codigos_unicos:
            ws.cell(linha, 1, codigo)
            ws.cell(linha, 1).border = border
            ws.cell(linha, 1).alignment = Alignment(vertical="center")
            linha += 1
    else:
        ws.cell(linha, 1, "SEM_CODIGOS_NAO_ENCONTRADOS")
        ws.cell(linha, 1).border = border
        ws.cell(linha, 1).alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 28
    aplicar_filtro_e_congelamento(ws, linha_header=3, freeze="A4")
    ocultar_linhas_grade(ws)
    return ws



# ============================================================
# MÓDULO V11 - MODELO MANUAL DE ADMINISTRAÇÃO LOCAL
# ============================================================

def aplicar_estilo_tabela_adm(ws, linha_titulo, linha_header, ultima_coluna):
    """Aplica formatação padrão nas tabelas do módulo de Administração Local."""
    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(linha_titulo, 1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(linha_titulo, 1).fill = fill_titulo
    ws.merge_cells(start_row=linha_titulo, start_column=1, end_row=linha_titulo, end_column=ultima_coluna)

    for col in range(1, ultima_coluna + 1):
        cell = ws.cell(linha_header, col)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def configurar_colunas_adm(ws, larguras=None):
    """Configura larguras, filtros, congelamento e linhas de grade das abas de Administração Local."""
    if larguras is None:
        larguras = {
            "A": 18, "B": 42, "C": 12, "D": 14, "E": 14,
            "F": 14, "G": 16, "H": 16, "I": 45, "J": 18,
        }
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura
    try:
        ws.freeze_panes = "A5"
        ws.sheet_view.showGridLines = False
    except Exception:
        pass


def escrever_bloco_adm(ws, linha_inicio, titulo, linhas_base):
    """
    Escreve um bloco editável de Administração Local.

    As colunas seguem o padrão do modelo enviado:
        Código, Descrição, Unidade, Coeficiente, Quantidade, Meses,
        Preço Unitário, Preço Total e Observação.
    """
    headers = [
        "CODIGO_SICRO", "DESCRICAO", "UNID", "COEFICIENTE", "QUANTIDADE",
        "MESES", "PRECO_UNITARIO", "PRECO_TOTAL", "OBSERVACAO"
    ]
    linha_titulo = linha_inicio
    linha_header = linha_inicio + 1
    ws.cell(linha_titulo, 1, titulo)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(linha_header, col_idx, header)
    aplicar_estilo_tabela_adm(ws, linha_titulo, linha_header, len(headers))

    linha = linha_header + 1
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for item in linhas_base:
        codigo, descricao, unid, coef, qtd, meses, preco, obs = item
        ws.cell(linha, 1, codigo)
        ws.cell(linha, 2, descricao)
        ws.cell(linha, 3, unid)
        ws.cell(linha, 4, coef)
        ws.cell(linha, 5, qtd)
        ws.cell(linha, 6, meses)
        ws.cell(linha, 7, preco)
        ws.cell(linha, 8, f'=IFERROR(D{linha}*E{linha}*F{linha}*G{linha},0)')
        ws.cell(linha, 9, obs)
        for col in range(1, 10):
            cell = ws.cell(linha, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col in {1, 4, 5, 6, 7}:
                cell.fill = fill_input
        linha += 1

    ws.cell(linha, 7, "TOTAL")
    ws.cell(linha, 7).font = Font(bold=True)
    ws.cell(linha, 8, f"=SUM(H{linha_header+1}:H{linha-1})")
    ws.cell(linha, 8).font = Font(bold=True)
    for col in range(1, 10):
        ws.cell(linha, col).border = border
    return linha + 2


def criar_modelo_administracao_local(wb):
    """
    Cria o esqueleto manual da Administração Local.

    Importante:
        Este módulo NÃO automatiza a administração local. Ele apenas cria as abas e
        tabelas editáveis, inspiradas no arquivo-modelo de Administração Local e no
        Manual SICRO V07, para o orçamentista preencher conforme a situação da obra.
    """
    abas_adm = [
        "01-ADM-Resumo",
        "02-ADM-Fixa",
        "03-ADM-Vinculada",
        "04-ADM-Variavel",
        "05-ADM-Manutencao-Canteiro",
        "06-ADM-Custos-Diversos",
        "07-GUIA_DNIT_AL",
    ]
    for nome in abas_adm:
        if nome in wb.sheetnames:
            wb.remove(wb[nome])

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ------------------------------------------------------------
    # 01 - Resumo
    # ------------------------------------------------------------
    ws = wb.create_sheet("01-ADM-Resumo")
    ws["A1"] = "RESUMO DAS PARCELAS DA ADMINISTRAÇÃO LOCAL"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")

    parametros = [
        ("Data-base", ""),
        ("Estado / Referência SICRO", ""),
        ("Natureza da obra", ""),
        ("Porte da obra", ""),
        ("Prazo contratual (meses)", ""),
        ("Extensão / parâmetro de porte", ""),
        ("Observações gerais", ""),
    ]
    ws["A3"] = "PARÂMETROS GERAIS"
    ws["A3"].font = Font(bold=True)
    ws["A3"].fill = fill_header
    ws.merge_cells("A3:C3")
    for idx, (nome, valor) in enumerate(parametros, start=4):
        ws.cell(idx, 1, nome)
        ws.cell(idx, 2, valor)
        ws.cell(idx, 2).fill = fill_input
        ws.cell(idx, 1).border = ws.cell(idx, 2).border = border

    linha_resumo = 13
    resumo_headers = ["PARCELA", "VALOR_TOTAL", "OBSERVACAO"]
    for col_idx, header in enumerate(resumo_headers, start=1):
        ws.cell(linha_resumo, col_idx, header)
        ws.cell(linha_resumo, col_idx).font = Font(bold=True)
        ws.cell(linha_resumo, col_idx).fill = fill_header
        ws.cell(linha_resumo, col_idx).border = border
    resumo_linhas = [
        ("Parcela Fixa", "='02-ADM-Fixa'!H1000", "Gerência técnica e administrativa."),
        ("Parcela Vinculada", "='03-ADM-Vinculada'!H1000", "Equipes de produção, topografia, medicina e segurança."),
        ("Parcela Variável", "='04-ADM-Variavel'!H1000", "Frentes de serviço, controle tecnológico e manejo florestal."),
        ("Manutenção do Canteiro", "='05-ADM-Manutencao-Canteiro'!H1000", "Manutenção de canteiro e acampamentos."),
        ("Custos Diversos", "='06-ADM-Custos-Diversos'!H1000", "Custos recorrentes diversos."),
    ]
    for idx, (parcela, formula, obs) in enumerate(resumo_linhas, start=linha_resumo+1):
        ws.cell(idx, 1, parcela)
        ws.cell(idx, 2, formula)
        ws.cell(idx, 3, obs)
        for col in range(1, 4):
            ws.cell(idx, col).border = border
            ws.cell(idx, col).alignment = Alignment(vertical="center", wrap_text=True)
    linha_total = linha_resumo + 1 + len(resumo_linhas)
    ws.cell(linha_total, 1, "TOTAL ADMINISTRAÇÃO LOCAL")
    ws.cell(linha_total, 2, f"=SUM(B{linha_resumo+1}:B{linha_total-1})")
    ws.cell(linha_total, 1).font = ws.cell(linha_total, 2).font = Font(bold=True)
    for col in range(1, 4):
        ws.cell(linha_total, col).fill = fill_header
        ws.cell(linha_total, col).border = border
    configurar_colunas_adm(ws, {"A": 34, "B": 22, "C": 65})

    # ------------------------------------------------------------
    # 02 - Parcela fixa
    # ------------------------------------------------------------
    ws = wb.create_sheet("02-ADM-Fixa")
    ws["A1"] = "ADMINISTRAÇÃO LOCAL - PARCELA FIXA"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")
    linha = 3
    linha = escrever_bloco_adm(ws, linha, "GERÊNCIA TÉCNICA", [
        ("", "Engenheiro chefe / supervisor", "mês", 1, "", "", "", "Ver Manual V07: parcela fixa e gerência técnica."),
        ("", "Engenheiro auxiliar", "mês", 1, "", "", "", "Avaliar conforme natureza e porte da obra."),
        ("", "Encarregado geral", "mês", 1, "", "", "", "Equipe de gerenciamento da obra."),
        ("", "Técnico ambiental", "mês", 1, "", "", "", "Aplicável quando exigido pelo escopo/licenciamento."),
        ("", "Motorista", "mês", 1, "", "", "", "Associado aos veículos de apoio da gerência."),
        ("", "Veículo leve - gerência técnica", "mês", 1, "", "", "", "Selecionar composição/insumo SICRO aplicável."),
    ])
    linha = escrever_bloco_adm(ws, linha, "GERÊNCIA ADMINISTRATIVA", [
        ("", "Administrador / gerente administrativo", "mês", 1, "", "", "", "Ver Manual V07: gerência administrativa."),
        ("", "Auxiliar administrativo", "mês", 1, "", "", "", "Dimensionar conforme estrutura da obra."),
        ("", "Almoxarife", "mês", 1, "", "", "", "Avaliar necessidade de almoxarifado/canteiro."),
        ("", "Apontador", "mês", 1, "", "", "", "Controle de produção, presença e recursos."),
        ("", "Vigia / segurança patrimonial", "mês", 1, "", "", "", "Quando aplicável ao canteiro/acampamento."),
    ])
    ws.cell(1000, 7, "TOTAL GERAL")
    ws.cell(1000, 8, '=SUMIF(G:G,"TOTAL",H:H)')
    configurar_colunas_adm(ws)

    # ------------------------------------------------------------
    # 03 - Parcela vinculada
    # ------------------------------------------------------------
    ws = wb.create_sheet("03-ADM-Vinculada")
    ws["A1"] = "ADMINISTRAÇÃO LOCAL - PARCELA VINCULADA"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")
    linha = 3
    linha = escrever_bloco_adm(ws, linha, "EQUIPES DE PRODUÇÃO EM CAMPO", [
        ("", "Encarregado de produção", "mês", 1, "", "", "", "Selecionar conforme natureza/porte e frentes de serviço."),
        ("", "Encarregado de turma", "mês", 1, "", "", "", "Aplicável às frentes produtivas."),
        ("", "Apontador de campo", "mês", 1, "", "", "", "Apoio ao controle das equipes de produção."),
        ("", "Veículo de apoio de produção", "mês", 1, "", "", "", "Ver veículos disponíveis no SICRO."),
    ])
    linha = escrever_bloco_adm(ws, linha, "EQUIPE DE TOPOGRAFIA", [
        ("", "Topógrafo", "mês", 1, "", "", "", "Ver tabelas de equipe de topografia no Manual V07."),
        ("", "Auxiliar de topografia", "mês", 1, "", "", "", "Dimensionar conforme frentes e extensão."),
        ("", "Veículo de topografia", "mês", 1, "", "", "", "Quando necessário ao deslocamento da equipe."),
        ("", "Equipamentos de topografia", "mês", 1, "", "", "", "GPS, estação total, nível ou equivalente, quando aplicável."),
    ])
    linha = escrever_bloco_adm(ws, linha, "MEDICINA E SEGURANÇA DO TRABALHO", [
        ("", "Técnico de segurança do trabalho", "mês", 1, "", "", "", "Dimensionar conforme quantidade total de profissionais."),
        ("", "Engenheiro de segurança do trabalho", "mês", 1, "", "", "", "Quando exigido pela legislação/porte."),
        ("", "Auxiliar de enfermagem / técnico de saúde", "mês", 1, "", "", "", "Quando aplicável."),
        ("", "Veículo de segurança/medicina", "mês", 1, "", "", "", "Apoio aos deslocamentos e inspeções."),
    ])
    ws.cell(1000, 7, "TOTAL GERAL")
    ws.cell(1000, 8, '=SUMIF(G:G,"TOTAL",H:H)')
    configurar_colunas_adm(ws)

    # ------------------------------------------------------------
    # 04 - Parcela variável
    # ------------------------------------------------------------
    ws = wb.create_sheet("04-ADM-Variavel")
    ws["A1"] = "ADMINISTRAÇÃO LOCAL - PARCELA VARIÁVEL"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")
    linha = 3
    linha = escrever_bloco_adm(ws, linha, "ACOMPANHAMENTO DAS FRENTES DE SERVIÇO", [
        ("", "Equipe de acompanhamento - terraplenagem", "mês", 1, "", "", "", "Ver Manual V07: serviços de terraplenagem."),
        ("", "Equipe de acompanhamento - pavimentação", "mês", 1, "", "", "", "Ver Manual V07: serviços de pavimentação."),
        ("", "Equipe de acompanhamento - drenagem", "mês", 1, "", "", "", "Ver Manual V07: drenagem e OAC."),
        ("", "Equipe de acompanhamento - sinalização/obras complementares", "mês", 1, "", "", "", "Ver Manual V07: sinalização, complementares e ambiental."),
        ("", "Equipe de acompanhamento - OAE/intervenções pontuais", "mês", 1, "", "", "", "Ver Manual V07: OAE e intervenções restritas."),
    ])
    linha = escrever_bloco_adm(ws, linha, "CONTROLE TECNOLÓGICO", [
        ("", "Laboratório de solos - terraplenagem", "mês", 1, "", "", "", "Ver Manual V07: controle tecnológico de solos."),
        ("", "Laboratório de solos - pavimentação", "mês", 1, "", "", "", "Ver Manual V07: pavimentação."),
        ("", "Laboratório de asfaltos", "mês", 1, "", "", "", "Quando houver serviços asfálticos."),
        ("", "Laboratório de concreto", "mês", 1, "", "", "", "Quando houver concreto/OAE/OAC."),
        ("", "Veículo de laboratório / controle tecnológico", "mês", 1, "", "", "", "Quando aplicável."),
    ])
    linha = escrever_bloco_adm(ws, linha, "MANEJO FLORESTAL", [
        ("", "Equipe de controle e manejo florestal", "mês", 1, "", "", "", "Aplicável quando previsto no escopo ambiental."),
        ("", "Veículo de manejo florestal", "mês", 1, "", "", "", "Quando aplicável."),
    ])
    ws.cell(1000, 7, "TOTAL GERAL")
    ws.cell(1000, 8, '=SUMIF(G:G,"TOTAL",H:H)')
    configurar_colunas_adm(ws)

    # ------------------------------------------------------------
    # 05 - Manutenção do canteiro
    # ------------------------------------------------------------
    ws = wb.create_sheet("05-ADM-Manutencao-Canteiro")
    ws["A1"] = "ADMINISTRAÇÃO LOCAL - MANUTENÇÃO DO CANTEIRO"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")
    linha = 3
    linha = escrever_bloco_adm(ws, linha, "MANUTENÇÃO DO CANTEIRO DE OBRAS E ACAMPAMENTOS", [
        ("", "Equipe de manutenção do canteiro", "mês", 1, "", "", "", "Ver Manual V07: manutenção do canteiro e acampamentos."),
        ("", "Servente / auxiliar de serviços gerais", "mês", 1, "", "", "", "Dimensionar conforme área/porte do canteiro."),
        ("", "Eletricista / manutenção predial", "mês", 1, "", "", "", "Quando aplicável."),
        ("", "Encanador / manutenção hidráulica", "mês", 1, "", "", "", "Quando aplicável."),
        ("", "Veículo de apoio à manutenção", "mês", 1, "", "", "", "Quando aplicável."),
    ])
    ws.cell(1000, 7, "TOTAL GERAL")
    ws.cell(1000, 8, '=SUMIF(G:G,"TOTAL",H:H)')
    configurar_colunas_adm(ws)

    # ------------------------------------------------------------
    # 06 - Custos diversos
    # ------------------------------------------------------------
    ws = wb.create_sheet("06-ADM-Custos-Diversos")
    ws["A1"] = "ADMINISTRAÇÃO LOCAL - CUSTOS DIVERSOS"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:I1")
    linha = 3
    linha = escrever_bloco_adm(ws, linha, "CUSTOS RECORRENTES E APOIO ADMINISTRATIVO", [
        ("", "Energia elétrica", "mês", 1, "", "", "", "Preencher conforme estimativa/localidade."),
        ("", "Água e esgoto", "mês", 1, "", "", "", "Preencher conforme estimativa/localidade."),
        ("", "Internet / telefonia", "mês", 1, "", "", "", "Preencher conforme necessidade."),
        ("", "Material de escritório", "mês", 1, "", "", "", "Custos administrativos recorrentes."),
        ("", "Limpeza e conservação", "mês", 1, "", "", "", "Quando não incluído em outra parcela."),
        ("", "Segurança patrimonial / vigilância", "mês", 1, "", "", "", "Quando aplicável."),
        ("", "Aluguel de imóveis / apoio local", "mês", 1, "", "", "", "Quando aplicável."),
    ])
    ws.cell(1000, 7, "TOTAL GERAL")
    ws.cell(1000, 8, '=SUMIF(G:G,"TOTAL",H:H)')
    configurar_colunas_adm(ws)

    # ------------------------------------------------------------
    # 07 - Guia DNIT Administração Local
    # ------------------------------------------------------------
    ws = wb.create_sheet("07-GUIA_DNIT_AL")
    ws["A1"] = "GUIA DE CONSULTA - ADMINISTRAÇÃO LOCAL (SICRO V07)"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:G1")
    headers = ["ITEM", "PARCELA", "MANUAL", "PAGINAS", "TABELAS/FIGURAS", "CRITERIO DE VERIFICACAO", "AUTOMATIZADO"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(3, col_idx, header)
        ws.cell(3, col_idx).font = Font(bold=True)
        ws.cell(3, col_idx).fill = fill_header
        ws.cell(3, col_idx).border = border
        ws.cell(3, col_idx).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    guia = [
        ("Natureza e porte da obra", "Premissa", "SICRO V07 - Administração Local", "22 a 27", "Figura 2; Tabelas 1 a 4", "Classificar obra por natureza, porte, extensão e prazo.", "Não"),
        ("Parcelas da Administração Local", "Premissa", "SICRO V07 - Administração Local", "27 a 30", "Figura 3", "Separar fixa, vinculada, variável, manutenção de canteiro e custos diversos.", "Não"),
        ("Parcela fixa", "Fixa", "SICRO V07 - Administração Local", "46 a 54", "Tabelas 5 a 14", "Selecionar estrutura de gerência técnica e administrativa conforme natureza/porte.", "Não"),
        ("Equipes de produção", "Vinculada", "SICRO V07 - Administração Local", "55 a 58", "Tabelas 15 a 18", "Dimensionar equipes vinculadas à produção em campo.", "Não"),
        ("Equipe de topografia", "Vinculada", "SICRO V07 - Administração Local", "58 a 59", "Tabelas 19 e 20", "Verificar necessidade conforme frentes, extensão e tipo de obra.", "Não"),
        ("Medicina e segurança do trabalho", "Vinculada", "SICRO V07 - Administração Local", "59 a 60", "Tabelas 21 e 22", "Dimensionar conforme quantidade total de trabalhadores e exigências legais.", "Não"),
        ("Acompanhamento das frentes de serviço", "Variável", "SICRO V07 - Administração Local", "61 a 73", "Tabelas 23 a 29", "Avaliar serviços de terraplenagem, pavimentação, drenagem, OAC, sinalização, OAE e conservação.", "Não"),
        ("Controle tecnológico", "Variável", "SICRO V07 - Administração Local", "73 a 85", "Tabelas 30 a 43", "Selecionar laboratórios e equipes conforme grupos de serviços executados.", "Não"),
        ("Manejo florestal", "Variável", "SICRO V07 - Administração Local", "85", "Tabela 44", "Aplicar quando previsto no escopo ambiental.", "Não"),
        ("Quantidade total de profissionais", "Consolidação", "SICRO V07 - Administração Local", "86", "Seção 3.4", "Conferir efetivo total e impactos nas equipes vinculadas e canteiro.", "Parcial"),
        ("Veículos da Administração Local", "Veículos", "SICRO V07 - Administração Local", "87 a 88", "Tabela 45", "Escolher veículos compatíveis com as equipes dimensionadas.", "Não"),
        ("Manutenção do canteiro", "Manutenção", "SICRO V07 - Administração Local", "88 a 92", "Tabelas 46 a 49", "Dimensionar manutenção conforme área, tipo de obra e canteiro/acampamento.", "Não"),
        ("Custos diversos", "Custos diversos", "SICRO V07 - Administração Local", "92", "Seção 3.7", "Prever custos recorrentes: energia, água, escritório, vigilância, aluguéis etc.", "Não"),
        ("Critérios de medição", "Medição", "SICRO V07 - Administração Local", "93", "Seção 3.8", "Verificar critério de medição da administração local no orçamento/edital.", "Não"),
    ]
    for row_idx, linha_guia in enumerate(guia, start=4):
        for col_idx, valor in enumerate(linha_guia, start=1):
            cell = ws.cell(row_idx, col_idx, valor)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    larguras = {"A": 34, "B": 18, "C": 34, "D": 16, "E": 26, "F": 70, "G": 16}
    configurar_colunas_adm(ws, larguras)
    aplicar_filtro_e_congelamento(ws, linha_header=3, freeze="A4")

    return wb


def escrever_aba_histograma_equip(wb, ws_equipamentos=None):
    """
    Cria a aba HISTOGRAMA_EQUIP como visão gerencial dos equipamentos.

    Nesta etapa o sistema ainda não automatiza horas de equipamento, pois isso
    depende de premissas de planejamento, produtividade e cronograma. A aba traz
    o ranking por ocorrência extraída do SICRO e deixa campos editáveis para uso
    futuro no dimensionamento.
    """
    if "HISTOGRAMA_EQUIP" in wb.sheetnames:
        wb.remove(wb["HISTOGRAMA_EQUIP"])
    ws = wb.create_sheet("HISTOGRAMA_EQUIP")

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "HISTOGRAMA DE EQUIPAMENTOS"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:H1")

    ws["A2"] = "Observação"
    ws["B2"] = "Ranking baseado nas ocorrências extraídas do SICRO. Horas, meses e equipamento equivalente ficam como campos manuais para planejamento."
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = fill_header
    ws["B2"].fill = fill_header

    headers = [
        "CODIGO",
        "DESCRICAO",
        "OCORRENCIAS",
        "HORAS_PREVISTAS",
        "MESES",
        "HORAS_MES",
        "EQUIPAMENTOS_EQUIVALENTES",
        "OBSERVACAO",
    ]
    linha_header = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(linha_header, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    if ws_equipamentos is None and "EQUIPAMENTOS" in wb.sheetnames:
        ws_equipamentos = wb["EQUIPAMENTOS"]

    linha_saida = linha_header + 1
    if ws_equipamentos is not None and ws_equipamentos.max_row >= 2:
        # A aba EQUIPAMENTOS possui colunas: CODIGO, DESCRICAO e OCORRENCIAS.
        for row in range(2, ws_equipamentos.max_row + 1):
            codigo = ws_equipamentos.cell(row, 1).value
            descricao = ws_equipamentos.cell(row, 2).value
            if not codigo and not descricao:
                continue
            ws.cell(linha_saida, 1, f"=EQUIPAMENTOS!A{row}")
            ws.cell(linha_saida, 2, f"=EQUIPAMENTOS!B{row}")
            ws.cell(linha_saida, 3, f"=EQUIPAMENTOS!C{row}")
            ws.cell(linha_saida, 4, None)
            ws.cell(linha_saida, 5, None)
            ws.cell(linha_saida, 6, f'=IF(OR(D{linha_saida}="",E{linha_saida}="",E{linha_saida}=0),"",D{linha_saida}/E{linha_saida})')
            ws.cell(linha_saida, 7, f'=IF(OR(F{linha_saida}="",$B$3="",$B$3=0),"",ROUNDUP(F{linha_saida}/$B$3,0))')
            ws.cell(linha_saida, 8, "Preencher horas e meses se desejar estimar equipamento equivalente.")
            for col in [4, 5]:
                ws.cell(linha_saida, col).fill = fill_input
            linha_saida += 1
    else:
        ws.cell(linha_saida, 1, "Nenhum equipamento encontrado.")
        linha_saida += 1

    ws["A3"] = "HORAS_MES_EQUIPAMENTO"
    ws["B3"] = HORAS_MES_PADRAO
    ws["A3"].font = Font(bold=True)
    ws["A3"].fill = fill_input
    ws["B3"].fill = fill_input

    for row in range(1, max(ws.max_row, linha_saida) + 1):
        for col in range(1, 9):
            cell = ws.cell(row, col)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    larguras = {"A": 16, "B": 55, "C": 16, "D": 18, "E": 12, "F": 18, "G": 28, "H": 65}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A5")
    ocultar_linhas_grade(ws)
    return ws


def criar_aba_check_orcamento(wb):
    """
    Cria a aba CHECK_ORCAMENTO.

    Objetivo:
        Servir como painel de conferência do orçamento antes da entrega,
        apontando pendências de preenchimento em módulos manuais e automáticos.

    Diretriz V14:
        - não automatizar decisões técnicas;
        - apenas indicar pontos que precisam ser revisados pelo orçamentista;
        - manter fórmulas simples, auditáveis e editáveis no Excel.
    """
    if "CHECK_ORCAMENTO" in wb.sheetnames:
        wb.remove(wb["CHECK_ORCAMENTO"])

    ws = wb.create_sheet("CHECK_ORCAMENTO")

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_obs = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "CHECK ORÇAMENTO - PAINEL DE CONFERÊNCIA"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:G1")

    ws["A2"] = "STATUS GERAL"
    ws["B2"] = '=IF(COUNTIF(A6:A200,"PENDENTE")>0,"PENDÊNCIAS ENCONTRADAS","APTO PARA REVISÃO")'
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)
    ws["A2"].fill = fill_header
    ws["B2"].fill = fill_obs

    ws["A3"] = "Observação"
    ws["B3"] = "Esta aba não substitui a revisão técnica. Ela apenas aponta campos vazios, totais zerados e pontos que precisam de conferência manual."
    ws["A3"].font = Font(bold=True)
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells("B3:G3")

    headers = ["STATUS", "MÓDULO", "VERIFICAÇÃO", "ABA", "PENDÊNCIAS", "AÇÃO NECESSÁRIA", "CRITICIDADE"]
    linha_header = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(linha_header, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    checks = [
        (
            "Extração SICRO",
            "Códigos não encontrados",
            "ALERTAS",
            '=MAX(COUNTA(ALERTAS!A4:A10000)-COUNTIF(ALERTAS!A4:A10000,"SEM_CODIGOS_NAO_ENCONTRADOS"),0)',
            "Verificar referência SICRO, mês/ano, estado, código digitado ou item manual.",
            "ALTA",
        ),
        (
            "Quantidades",
            "Serviços sem quantidade",
            "QUANTIDADES",
            '=COUNTIFS(QUANTIDADES!A2:A10000,"<>",QUANTIDADES!D2:D10000,"")',
            "Preencher ou validar a quantidade dos serviços.",
            "ALTA",
        ),
        (
            "Quantidades",
            "Serviços sem meses",
            "QUANTIDADES",
            '=COUNTIFS(QUANTIDADES!A2:A10000,"<>",QUANTIDADES!E2:E10000,"")',
            "Preencher meses por serviço ou usar o prazo padrão no HISTOGRAMA_MO.",
            "MÉDIA",
        ),
        (
            "Mão de obra",
            "Linhas sem quantidade na base técnica",
            "MAO_OBRA_CALCULO",
            '=COUNTIF(MAO_OBRA_CALCULO!P:P,"Informar quantidade")',
            "Revisar a aba QUANTIDADES para os serviços com mão de obra calculada.",
            "ALTA",
        ),
        (
            "Mão de obra",
            "Linhas sem meses na base técnica",
            "MAO_OBRA_CALCULO",
            '=COUNTIF(MAO_OBRA_CALCULO!P:P,"Informar meses")',
            "Preencher meses na aba QUANTIDADES ou prazo padrão no HISTOGRAMA_MO.",
            "MÉDIA",
        ),
        (
            "Mão de obra",
            "Colaboradores/mês zerado",
            "PESSOAS",
            '=IFERROR(IF(INDEX(PESSOAS!H:H,MATCH("TOTAL",PESSOAS!A:A,0))=0,1,0),1)',
            "Conferir quantidades, meses e extração de mão de obra.",
            "ALTA",
        ),
        (
            "DMT",
            "Itens DMT extraídos para revisão",
            "DMT",
            '=COUNTA(DMT!A2:A10000)',
            "Conferir se há itens de transporte que exigem distância, fonte ou decisão manual.",
            "INFORMATIVO",
        ),
        (
            "Mobilização Equipamentos",
            "Equipamentos sem quantidade",
            "MOB_EQUIP",
            '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!C3:C1000,"")',
            "Preencher quantidade apenas para os equipamentos que serão mobilizados.",
            "MÉDIA",
        ),
        (
            "Mobilização Equipamentos",
            "Equipamentos sem FU",
            "MOB_EQUIP",
            '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!F3:F1000,"")',
            "Preencher FU quando o equipamento entrar no cálculo de mobilização.",
            "MÉDIA",
        ),
        (
            "Mobilização Equipamentos",
            "Equipamentos sem K",
            "MOB_EQUIP",
            '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!G3:G1000,"")',
            "Preencher fator K quando aplicável.",
            "MÉDIA",
        ),
        (
            "Mobilização Equipamentos",
            "Equipamentos sem distância",
            "MOB_EQUIP",
            '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!I3:I1000,"")',
            "Preencher distância de mobilização/desmobilização quando aplicável.",
            "MÉDIA",
        ),
        (
            "Mobilização Pessoas",
            "Pessoas sem quantidade",
            "MOB_PESSOAS",
            '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!C4:C103,"")',
            "Preencher apenas as pessoas que serão mobilizadas.",
            "MÉDIA",
        ),
        (
            "Mobilização Pessoas",
            "Pessoas sem custo de passagem",
            "MOB_PESSOAS",
            '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!D4:D103,"")',
            "Informar custo unitário de passagem quando aplicável.",
            "BAIXA",
        ),
        (
            "Mobilização Pessoas",
            "Pessoas sem custo de alimentação",
            "MOB_PESSOAS",
            '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!E4:E103,"")',
            "Informar custo unitário de alimentação quando aplicável.",
            "BAIXA",
        ),
        (
            "Administração Local",
            "Resumo de ADM Local zerado ou ausente",
            "01-ADM-Resumo",
            '=IFERROR(IF(SUM(\'01-ADM-Resumo\'!A1:Z500)=0,1,0),1)',
            "Conferir se a Administração Local foi preenchida conforme o modelo e manual.",
            "MÉDIA",
        ),
        (
            "Canteiro",
            "Resumo de canteiro zerado ou ausente",
            "CANT_Resumo",
            '=IFERROR(IF(SUM(CANT_Resumo!A1:Z500)=0,1,0),1)',
            "Conferir as abas CANT_Princ, CANT_Industrial, CANT_Complementar e CANT_Resumo.",
            "MÉDIA",
        ),
        (
            "Canteiro",
            "Guia de canteiro ausente",
            "GUIA_CANTEIRO_SICRO",
            '=IFERROR(IF(COUNTA(GUIA_CANTEIRO_SICRO!A:A)=0,1,0),1)',
            "Verificar a criação da aba de orientação do Manual SICRO V06.",
            "BAIXA",
        ),
    ]

    for row_idx, (modulo, verificacao, aba, formula, acao, criticidade) in enumerate(checks, start=6):
        ws.cell(row_idx, 1, f'=IF(OR(E{row_idx}=0,G{row_idx}="INFORMATIVO"),"OK","PENDENTE")')
        ws.cell(row_idx, 2, modulo)
        ws.cell(row_idx, 3, verificacao)
        ws.cell(row_idx, 4, aba)
        ws.cell(row_idx, 5, formula)
        ws.cell(row_idx, 6, acao)
        ws.cell(row_idx, 7, criticidade)
        for col in range(1, 8):
            cell = ws.cell(row_idx, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Bloco de orientação manual.
    linha_obs = 6 + len(checks) + 2
    ws.cell(linha_obs, 1, "ORIENTAÇÃO")
    ws.cell(linha_obs, 1).font = Font(bold=True)
    ws.cell(linha_obs, 1).fill = fill_header
    ws.cell(linha_obs, 2, "Itens informativos, como DMT extraído, não significam erro; indicam apenas necessidade de conferência técnica.")
    ws.merge_cells(start_row=linha_obs, start_column=2, end_row=linha_obs, end_column=7)
    ws.cell(linha_obs, 2).alignment = Alignment(wrap_text=True, vertical="center")

    larguras = {"A": 16, "B": 28, "C": 42, "D": 24, "E": 16, "F": 70, "G": 16}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A6")
    ocultar_linhas_grade(ws)
    return ws

def criar_aba_resumo(wb):
    """Cria uma aba inicial limpa com os principais indicadores e instruções de uso."""
    if "RESUMO" in wb.sheetnames:
        wb.remove(wb["RESUMO"])
    ws = wb.create_sheet("RESUMO", 0)

    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_secao = PatternFill("solid", fgColor="D9EAF7")
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "PAINEL RESUMO - EXTRATOR SICRO"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:F1")

    indicadores = [
        ("Serviços informados", '=COUNTA(QUANTIDADES!A2:A10000)'),
        ("Total de funções de mão de obra", '=MAX(COUNTA(PESSOAS!A5:A10000)-1,0)'),
        ("Horas totais de mão de obra", '=IFERROR(INDEX(PESSOAS!F:F,MATCH("TOTAL",PESSOAS!A:A,0)),0)'),
        ("Colaboradores totais", '=IFERROR(INDEX(PESSOAS!G:G,MATCH("TOTAL",PESSOAS!A:A,0)),0)'),
        ("Colaboradores/mês", '=IFERROR(INDEX(PESSOAS!H:H,MATCH("TOTAL",PESSOAS!A:A,0)),0)'),
        ("Equipamentos consolidados", '=COUNTA(EQUIPAMENTOS!A2:A10000)'),
        ("Itens DMT", '=COUNTA(DMT!A2:A10000)'),
        ("Códigos não encontrados", '=MAX(COUNTA(ALERTAS!A4:A10000)-COUNTIF(ALERTAS!A4:A10000,"SEM_CODIGOS_NAO_ENCONTRADOS"),0)'),
    ]

    ws["A3"] = "INDICADOR"
    ws["B3"] = "VALOR"
    ws["A3"].fill = fill_secao
    ws["B3"].fill = fill_secao
    ws["A3"].font = ws["B3"].font = Font(bold=True)

    for i, (nome, formula) in enumerate(indicadores, start=4):
        ws.cell(i, 1, nome)
        ws.cell(i, 2, formula)
        ws.cell(i, 1).border = ws.cell(i, 2).border = border

    ws["D3"] = "COMO USAR"
    ws["D3"].fill = fill_secao
    ws["D3"].font = Font(bold=True)
    instrucoes = [
        "1. Valide as quantidades e preencha os meses na aba QUANTIDADES.",
        "2. Consulte a mão de obra consolidada na aba PESSOAS.",
        "3. Use HISTOGRAMA_MO para verificar efetivo mensal.",
        "4. Consulte EQUIPAMENTOS e DMT para logística e orçamento.",
        "5. Verifique pendências na aba ALERTAS.",
    ]
    for i, txt in enumerate(instrucoes, start=4):
        ws.cell(i, 4, txt)
        ws.cell(i, 4).alignment = Alignment(wrap_text=True, vertical="center")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 90
    ws.sheet_view.showGridLines = False
    return ws


def aplicar_limpeza_visual_final(wb):
    """
    Reordena abas, oculta bases técnicas e aplica filtros/congelamentos.
    Mantém a planilha funcional, mas com visual mais limpo para uso diário.
    """
    # Abas técnicas que permanecem disponíveis, mas ocultas.
    abas_ocultas = {
        "MAO_OBRA_CALCULO",
        "ATIVIDADES_AUXILIARES",
        "EQUIPAMENTOS_POR_COMPOSICAO",
    }

    # Oculta abas técnicas.
    for nome in abas_ocultas:
        if nome in wb.sheetnames:
            wb[nome].sheet_state = "hidden"

    # Aplica navegação limpa nas abas visíveis.
    for nome in ["CHECK_ORCAMENTO", "SINTETICO", "ANALITICO", "QUANTIDADES", "PESSOAS", "HISTOGRAMA_MO", "HISTOGRAMA_EQUIP", "EQUIPAMENTOS", "DMT", "ALERTAS", "01-ADM-Resumo", "02-ADM-Fixa", "03-ADM-Vinculada", "04-ADM-Variavel", "05-ADM-Variavel", "06-Manutenção Canteiro", "07-Acordão", "05-ADM-Manutencao-Canteiro", "06-ADM-Custos-Diversos", "07-GUIA_DNIT_AL", "01-MO Ordinária", "05-Relação Equip", "06-Mob e Desm", "08-MOB-Resumo", "09-MOB-Pessoas", "10-MOB-Equipamentos", "11-MOB-Transportadores", "12-MOB-Parametros", "13-MOB-Velocidades", "14-MOB-Checklist", "15-GUIA_MOB_DNIT", "MOB_EQUIP", "MOB_PESSOAS", "MOB_RESUMO", "CANT_Resumo", "CANT_Princ", "CANT_Princ Container", "CANT_Industrial", "CANT_Complementar", "GUIA_CANTEIRO_SICRO"]:
        if nome in wb.sheetnames:
            ocultar_linhas_grade(wb[nome])
            if nome == "PESSOAS":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=4, freeze="A5")
            elif nome == "HISTOGRAMA_MO":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=6, freeze="A7")
            elif nome == "HISTOGRAMA_EQUIP":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=4, freeze="A5")
            elif nome == "ALERTAS":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=3, freeze="A4")
            else:
                aplicar_filtro_e_congelamento(wb[nome], linha_header=1, freeze="A2")

    # Reordena colocando as abas de uso no começo.
    ordem = ["RESUMO", "CHECK_ORCAMENTO", "SINTETICO", "ANALITICO", "QUANTIDADES", "PESSOAS", "HISTOGRAMA_MO", "HISTOGRAMA_EQUIP", "EQUIPAMENTOS", "DMT", "ALERTAS", "01-ADM-Resumo", "02-ADM-Fixa", "03-ADM-Vinculada", "04-ADM-Variavel", "05-ADM-Variavel", "06-Manutenção Canteiro", "07-Acordão", "05-ADM-Manutencao-Canteiro", "06-ADM-Custos-Diversos", "07-GUIA_DNIT_AL", "MOB_RESUMO", "MOB_PESSOAS", "MOB_EQUIP", "CANT_Resumo", "CANT_Princ", "CANT_Princ Container", "CANT_Industrial", "CANT_Complementar", "GUIA_CANTEIRO_SICRO"]
    for pos, nome in enumerate(ordem):
        if nome in wb.sheetnames:
            ws = wb[nome]
            wb._sheets.remove(ws)
            wb._sheets.insert(pos, ws)



# Guarda a versão manual simplificada da V11 como fallback.
_criar_modelo_administracao_local_manual = criar_modelo_administracao_local


def localizar_arquivo_modelo_adm_local():
    """
    Localiza automaticamente o arquivo-modelo de Administração Local.

    A V11.2 tenta usar o arquivo Excel enviado como referência para copiar as
    tabelas e a formatação das abas de Administração Local. Para não quebrar o
    programa quando o arquivo não estiver disponível, a busca é opcional e, se
    não localizar o modelo, o sistema usa o esqueleto manual da V11.1.
    """
    candidatos = []
    try:
        base_script = Path(__file__).resolve().parent
        candidatos.extend(base_script.glob("Exemplo*Administra*.xlsx"))
        candidatos.extend(base_script.glob("*Administracao*Local*.xlsx"))
        candidatos.extend(base_script.glob("*Administração*Local*.xlsx"))
    except Exception:
        pass

    try:
        candidatos.extend(Path.cwd().glob("Exemplo*Administra*.xlsx"))
        candidatos.extend(Path.cwd().glob("*Administracao*Local*.xlsx"))
        candidatos.extend(Path.cwd().glob("*Administração*Local*.xlsx"))
    except Exception:
        pass

    # Remove duplicidades preservando a ordem.
    vistos = set()
    unicos = []
    for caminho in candidatos:
        try:
            chave = caminho.resolve()
        except Exception:
            chave = caminho
        if chave in vistos:
            continue
        vistos.add(chave)
        if caminho.exists() and caminho.is_file():
            unicos.append(caminho)

    return unicos[0] if unicos else None


def formula_referencia_aba_excluida(formula):
    """Retorna True quando uma fórmula depende de abas de referência que não serão copiadas."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return False
    texto = formula.upper()
    referencias_excluidas = [
        "TABELA REFERENCIA!", "'TABELA REFERENCIA'!",
        "MO!", "'MO'!",
        "EQUI!", "'EQUI'!", "EQUI!", "EQUÍ!",
        "JAN2026!", "'JAN2026'!",
        "MANUAL!", "'MANUAL'!",
        "SICRO!", "'SICRO'!", "SICRO (2)!", "'SICRO (2)'!",
        "COPIE AQUI OS EQUIPAMENTOS!", "'COPIE AQUI OS EQUIPAMENTOS'!",
        "CPU_ADM!", "'CPU_ADM'!",
        "ITENS_SINTETICO!", "'ITENS_SINTETICO'!",
        "BASE - MATERIAL!", "'BASE - MATERIAL'!",
        "[1]", "[2]", "[3]",
    ]
    return any(ref in texto for ref in referencias_excluidas)


def copiar_aba_modelo_adm(ws_origem, ws_origem_valores, wb_destino, nome_destino=None):
    """
    Copia valores, fórmulas independentes, estilos, larguras, alturas e mesclagens.

    Fórmulas que dependem de abas de referência removidas são substituídas pelo
    valor calculado salvo no arquivo-modelo. Isso evita erros #REF! quando não
    copiamos as abas TABELA REFERENCIA, MO, Equi e Orçamento Sintético.
    """
    nome_destino = nome_destino or ws_origem.title
    if nome_destino in wb_destino.sheetnames:
        wb_destino.remove(wb_destino[nome_destino])
    ws_destino = wb_destino.create_sheet(nome_destino)

    for row in ws_origem.iter_rows():
        for cell in row:
            novo = ws_destino[cell.coordinate]
            valor = cell.value
            if formula_referencia_aba_excluida(valor):
                valor = ws_origem_valores[cell.coordinate].value
            novo.value = valor
            if cell.has_style:
                novo.font = copy(cell.font)
                novo.fill = copy(cell.fill)
                novo.border = copy(cell.border)
                novo.alignment = copy(cell.alignment)
                novo.number_format = cell.number_format
                novo.protection = copy(cell.protection)
            if cell.hyperlink:
                novo._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                novo.comment = copy(cell.comment)

    for col_key, dim in ws_origem.column_dimensions.items():
        ws_destino.column_dimensions[col_key].width = dim.width
        ws_destino.column_dimensions[col_key].hidden = dim.hidden
        ws_destino.column_dimensions[col_key].outlineLevel = dim.outlineLevel
    for row_key, dim in ws_origem.row_dimensions.items():
        ws_destino.row_dimensions[row_key].height = dim.height
        ws_destino.row_dimensions[row_key].hidden = dim.hidden
        ws_destino.row_dimensions[row_key].outlineLevel = dim.outlineLevel

    for intervalo in ws_origem.merged_cells.ranges:
        ws_destino.merge_cells(str(intervalo))

    ws_destino.sheet_view.showGridLines = ws_origem.sheet_view.showGridLines
    ws_destino.freeze_panes = ws_origem.freeze_panes
    try:
        ws_destino.sheet_properties.pageSetUpPr = copy(ws_origem.sheet_properties.pageSetUpPr)
        ws_destino.page_margins = copy(ws_origem.page_margins)
        ws_destino.page_setup = copy(ws_origem.page_setup)
    except Exception:
        pass

    return ws_destino


def copiar_modelo_adm_local_de_arquivo(wb_destino, caminho_modelo):
    """
    Copia do Excel-modelo apenas as abas úteis de Administração Local.

    Não copia as abas de referência auxiliares solicitadas para exclusão:
        - TABELA REFERENCIA
        - MO
        - Equi

    Não copia a aba Orçamento Sintético do modelo, pois o sistema já gera
    a aba SINTETICO como saída principal da extração SICRO.
    """
    abas_copiar = [
        "01-ADM-Resumo",
        "02-ADM-Fixa",
        "03-ADM-Vinculada",
        "04-ADM-Variavel",
        "05-ADM-Variavel",
        "06-Manutenção Canteiro",
        "07-Acordão",
    ]
    abas_remover = set(abas_copiar) | {
        "05-ADM-Manutencao-Canteiro",
        "06-ADM-Custos-Diversos",
        "07-GUIA_DNIT_AL",
        "TABELA REFERENCIA",
        "MO",
        "Equi",
        "Orçamento Sintético",
    }
    for nome in list(abas_remover):
        if nome in wb_destino.sheetnames:
            wb_destino.remove(wb_destino[nome])

    wb_modelo = load_workbook(caminho_modelo, data_only=False)
    wb_modelo_valores = load_workbook(caminho_modelo, data_only=True)

    for nome in abas_copiar:
        if nome not in wb_modelo.sheetnames:
            continue
        copiar_aba_modelo_adm(
            ws_origem=wb_modelo[nome],
            ws_origem_valores=wb_modelo_valores[nome],
            wb_destino=wb_destino,
            nome_destino=nome,
        )

    return True


def criar_modelo_administracao_local(wb):
    """
    Cria o módulo de Administração Local da V11.2.

    Regra principal:
        1. se o arquivo 'Exemplo Administração Local.xlsx' estiver na mesma pasta
           do script ou no diretório de execução, copia as tabelas e formatações
           oficiais do modelo;
        2. não copia a aba Orçamento Sintético do modelo, pois já existe SINTETICO;
        3. não copia as abas auxiliares de referência TABELA REFERENCIA, MO e Equi;
        4. se o modelo não for encontrado, usa o esqueleto manual da V11.1.
    """
    caminho_modelo = localizar_arquivo_modelo_adm_local()
    if caminho_modelo:
        try:
            copiar_modelo_adm_local_de_arquivo(wb, caminho_modelo)
            logger.info(f"Modelo de Administração Local copiado de: {caminho_modelo}")
            return
        except Exception as exc:
            logger.warning(f"Não foi possível copiar o modelo de Administração Local ({exc}). Usando esqueleto manual da V11.1.")

    _criar_modelo_administracao_local_manual(wb)


# ============================================================
# MÓDULO V12 - MODELO MANUAL DE MOBILIZAÇÃO E DESMOBILIZAÇÃO
# ============================================================


def localizar_arquivo_modelo_mobilizacao():
    """
    Localiza a planilha-base simplificada de mobilização.

    A partir da V12.4, a prioridade é o arquivo mob_e.xlsx, pois ele possui
    apenas três abas objetivas: MOB_EQUIP, MOB_PESSOAS e MOB_RESUMO.
    """
    candidatos = []
    padroes = [
        "mob_e*.xlsx",
        "MOB_E*.xlsx",
        "*mob_e*.xlsx",
        "*Mobiliza*.xlsx",
        "*Mobilizacao*.xlsx",
        "*Mobilização*.xlsx",
    ]

    locais = []
    try:
        locais.append(Path(__file__).resolve().parent)
    except Exception:
        pass
    try:
        locais.append(Path.cwd())
    except Exception:
        pass

    vistos = set()
    for base in locais:
        for padrao in padroes:
            try:
                for caminho in base.glob(padrao):
                    chave = str(caminho.resolve())
                    if chave in vistos:
                        continue
                    vistos.add(chave)
                    if caminho.exists() and caminho.is_file():
                        candidatos.append(caminho)
            except Exception:
                continue

    # Prioriza explicitamente mob_e.
    for caminho in candidatos:
        if "MOB_E" in normalizar_texto(caminho.name):
            return caminho
    return candidatos[0] if candidatos else None


def _coletar_equipamentos_extraidos_para_mob(wb):
    """
    Coleta somente os equipamentos extraídos dos códigos SICRO informados.

    Regra V12.4:
        - a mobilização usa os equipamentos da aba EQUIPAMENTOS;
        - não copia equipamentos da planilha-base;
        - quantidade, transportador, FU, K, distância, custos e tempos ficam manuais.
    """
    equipamentos = []
    vistos = set()

    if "EQUIPAMENTOS" in wb.sheetnames:
        ws = wb["EQUIPAMENTOS"]
        for row in range(2, ws.max_row + 1):
            codigo = limpar_codigo_original(ws.cell(row, 1).value)
            descricao = str(ws.cell(row, 2).value or "").strip()
            if not codigo and not descricao:
                continue
            chave = (normalizar_texto(codigo), normalizar_texto(descricao))
            if chave in vistos:
                continue
            vistos.add(chave)
            equipamentos.append({"CODIGO": codigo, "DESCRICAO": descricao})

    if not equipamentos and "EQUIPAMENTOS_POR_COMPOSICAO" in wb.sheetnames:
        ws = wb["EQUIPAMENTOS_POR_COMPOSICAO"]
        for row in range(2, ws.max_row + 1):
            codigo = limpar_codigo_original(ws.cell(row, 3).value)
            descricao = str(ws.cell(row, 4).value or "").strip()
            if not codigo and not descricao:
                continue
            chave = (normalizar_texto(codigo), normalizar_texto(descricao))
            if chave in vistos:
                continue
            vistos.add(chave)
            equipamentos.append({"CODIGO": codigo, "DESCRICAO": descricao})

    return equipamentos


def _copiar_estilo_linha(ws, linha_origem, linha_destino, max_col):
    """Copia estilo e altura de uma linha-modelo para outra linha."""
    try:
        ws.row_dimensions[linha_destino].height = ws.row_dimensions[linha_origem].height
    except Exception:
        pass
    for col in range(1, max_col + 1):
        origem = ws.cell(linha_origem, col)
        destino = ws.cell(linha_destino, col)
        if origem.has_style:
            destino.font = copy(origem.font)
            destino.fill = copy(origem.fill)
            destino.border = copy(origem.border)
            destino.alignment = copy(origem.alignment)
            destino.number_format = origem.number_format
            destino.protection = copy(origem.protection)


def _limpar_intervalo_valores(ws, linha_ini, linha_fim, col_ini, col_fim):
    """Limpa valores, preservando formatação."""
    for row in range(linha_ini, linha_fim + 1):
        for col in range(col_ini, col_fim + 1):
            ws.cell(row, col).value = None


def _localizar_linha_por_texto(ws, texto, colunas=(1, 2), inicio=1):
    alvo = normalizar_texto(texto)
    for row in range(inicio, ws.max_row + 1):
        for col in colunas:
            if alvo in normalizar_texto(ws.cell(row, col).value):
                return row
    return None


def _localizar_linha_total_mob_equip(ws):
    """Localiza a linha de total da aba MOB_EQUIP."""
    for row in range(3, ws.max_row + 1):
        valor_b = normalizar_texto(ws.cell(row, 2).value)
        valor_n = str(ws.cell(row, 14).value or "").upper()
        if valor_b == "TOTAL" or "SUBTOTAL" in valor_n:
            return row
    return ws.max_row


def _localizar_linha_total_mob_pessoas(ws):
    """Localiza a linha de total da aba MOB_PESSOAS."""
    for row in range(3, ws.max_row + 1):
        if normalizar_texto(ws.cell(row, 2).value) == "TOTAL":
            return row
        valor_f = str(ws.cell(row, 6).value or "").upper()
        if "SUBTOTAL" in valor_f:
            return row
    return ws.max_row


def _garantir_linhas_tabela(ws, linha_inicio, linha_total, qtd_linhas, max_col):
    """Garante linhas suficientes antes da linha de total."""
    capacidade = max(0, linha_total - linha_inicio)
    if qtd_linhas > capacidade:
        inserir = qtd_linhas - capacidade
        ws.insert_rows(linha_total, inserir)
        linha_modelo = max(linha_inicio, linha_total - 1)
        for i in range(inserir):
            _copiar_estilo_linha(ws, linha_modelo, linha_total + i, max_col)
        linha_total += inserir
    return linha_total


def preencher_mob_equipamentos_extraidos_v123(wb):
    """
    Preenche MOB_EQUIP com equipamentos extraídos, sem copiar itens do modelo.

    Somente código e descrição são automáticos. Todo o restante permanece manual
    para respeitar a análise técnica do orçamentista.
    """
    if "MOB_EQUIP" not in wb.sheetnames:
        return

    ws = wb["MOB_EQUIP"]
    equipamentos = _coletar_equipamentos_extraidos_para_mob(wb)
    linha_inicio = 3
    linha_total = _localizar_linha_total_mob_equip(ws)
    qtd_linhas = max(1, len(equipamentos))
    linha_total = _garantir_linhas_tabela(ws, linha_inicio, linha_total, qtd_linhas, 14)

    _limpar_intervalo_valores(ws, linha_inicio, linha_total - 1, 1, 14)

    if equipamentos:
        for idx, equip in enumerate(equipamentos):
            row = linha_inicio + idx
            ws.cell(row, 1).value = equip.get("CODIGO")
            ws.cell(row, 2).value = equip.get("DESCRICAO")
            ws.cell(row, 3).value = None  # Quantidade manual
            ws.cell(row, 4).value = None  # Transportador - código manual
            ws.cell(row, 5).value = None  # Transportador - descrição manual
            ws.cell(row, 6).value = None  # FU manual
            ws.cell(row, 7).value = None  # K manual
            ws.cell(row, 8).value = None  # Velocidade manual
            ws.cell(row, 9).value = None  # Distância manual
            ws.cell(row, 10).value = f"=IFERROR(I{row}/H{row},0)"
            ws.cell(row, 11).value = None  # Custo produtivo manual
            ws.cell(row, 12).value = None  # Tempo descanso manual
            ws.cell(row, 13).value = None  # Custo improdutivo manual
            ws.cell(row, 14).value = f"=IFERROR(C{row}*F{row}*G{row}*(J{row}*K{row}+L{row}*M{row}),0)"
    else:
        ws.cell(linha_inicio, 1).value = "SEM_EQUIPAMENTOS_EXTRAIDOS"
        ws.cell(linha_inicio, 2).value = "Nenhum equipamento foi extraído dos códigos SICRO informados."
        ws.cell(linha_inicio, 14).value = 0

    # Limpa sobras da planilha-base caso existam linhas além da lista real.
    for row in range(linha_inicio + qtd_linhas, linha_total):
        for col in range(1, 14):
            ws.cell(row, col).value = None
        ws.cell(row, 14).value = None

    ws.cell(linha_total, 2).value = "TOTAL"
    ws.cell(linha_total, 14).value = f"=SUBTOTAL(9,N{linha_inicio}:N{linha_total-1})"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def limpar_mob_pessoas_manual_v123(wb):
    """
    Mantém MOB_PESSOAS como tabela manual limpa.

    A planilha-base possuía fórmulas e profissionais de exemplo. Nesta versão,
    essas linhas são removidas para evitar trazer mão de obra que não pertence ao
    orçamento. O usuário preenche os profissionais efetivamente mobilizados.
    """
    if "MOB_PESSOAS" not in wb.sheetnames:
        return

    ws = wb["MOB_PESSOAS"]
    linha_inicio = 4
    linha_total = _localizar_linha_total_mob_pessoas(ws)
    linhas_manuais = 12
    linha_total = _garantir_linhas_tabela(ws, linha_inicio, linha_total, linhas_manuais, 6)
    _limpar_intervalo_valores(ws, linha_inicio, linha_total - 1, 1, 6)

    for row in range(linha_inicio, linha_inicio + linhas_manuais):
        ws.cell(row, 6).value = f"=IFERROR(ROUND(C{row}*(D{row}+E{row}),2),0)"

    ws.cell(linha_total, 2).value = "TOTAL"
    ws.cell(linha_total, 6).value = f"=SUBTOTAL(9,F{linha_inicio}:F{linha_total-1})"
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def ajustar_mob_resumo_v123(wb):
    """Ajusta MOB_RESUMO para apontar para as tabelas simplificadas."""
    if "MOB_RESUMO" not in wb.sheetnames:
        return

    ws = wb["MOB_RESUMO"]
    # Mantém visual da planilha-base, mas atualiza os vínculos principais.
    ws.cell(6, 1).value = "MOB_PESSOAS"
    ws.cell(6, 2).value = "Mobilização Pessoas"
    ws.cell(6, 3).value = '=IFERROR(INDEX(MOB_PESSOAS!F:F,MATCH("TOTAL",MOB_PESSOAS!B:B,0)),0)'
    ws.cell(7, 1).value = "MOB_EQUIP"
    ws.cell(7, 2).value = "Mobilização Equipamentos"
    ws.cell(7, 3).value = '=IFERROR(INDEX(MOB_EQUIP!N:N,MATCH("TOTAL",MOB_EQUIP!B:B,0)),0)'
    ws.cell(8, 1).value = "TOTAL"
    ws.cell(8, 2).value = "Total"
    ws.cell(8, 3).value = "=SUM(C6:C7)"
    ws.sheet_view.showGridLines = False


def copiar_modelo_mobilizacao_simplificado(wb_destino, caminho_modelo):
    """
    Copia somente as três abas da nova base simplificada: MOB_EQUIP,
    MOB_PESSOAS e MOB_RESUMO.

    Correção V12.4:
        A cópia do modelo é aceita apenas se as três abas forem efetivamente
        criadas no arquivo final. Se alguma aba não for encontrada/copiada,
        a rotina lança erro e o fallback manual cria a estrutura completa.
    """
    abas_copiar = ["MOB_EQUIP", "MOB_PESSOAS", "MOB_RESUMO"]
    for nome in abas_copiar:
        if nome in wb_destino.sheetnames:
            wb_destino.remove(wb_destino[nome])

    wb_modelo = load_workbook(caminho_modelo, data_only=False)
    wb_modelo_valores = load_workbook(caminho_modelo, data_only=True)

    abas_copiadas = []
    for nome in abas_copiar:
        if nome not in wb_modelo.sheetnames:
            continue
        copiar_aba_modelo_adm(wb_modelo[nome], wb_modelo_valores[nome], wb_destino, nome)
        abas_copiadas.append(nome)

    faltantes = [nome for nome in abas_copiar if nome not in wb_destino.sheetnames]
    if faltantes:
        raise RuntimeError(f"Abas de mobilização não copiadas do modelo: {faltantes}. Abas copiadas: {abas_copiadas}")

    preencher_mob_equipamentos_extraidos_v123(wb_destino)
    limpar_mob_pessoas_manual_v123(wb_destino)
    ajustar_mob_resumo_v123(wb_destino)


def _estilizar_cabecalho_simples(ws, titulo, max_col):
    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    for col in range(1, max_col + 1):
        cell = ws.cell(2, col)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def criar_modelo_mobilizacao_manual(wb):
    """
    Fallback da V12.4: cria apenas as três abas limpas de mobilização,
    seguindo a estrutura da planilha mob_e.
    """
    for nome in ["MOB_EQUIP", "MOB_PESSOAS", "MOB_RESUMO"]:
        if nome in wb.sheetnames:
            wb.remove(wb[nome])

    equipamentos = _coletar_equipamentos_extraidos_para_mob(wb)

    ws = wb.create_sheet("MOB_EQUIP")
    headers = [
        "Equipamento - Cdg", "Equipamento - Desc", "Quantidade",
        "Transportador - Cdg", "Transportador - Desc", "FU", "K",
        "Velocidade (km/h)", "Distância (km)", "Tempo (h)",
        "Custo Prod", "Tempo descanso (h)", "Custo Improd.", "Total (R$)"
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    _estilizar_cabecalho_simples(ws, "MOBILIZAÇÃO DE EQUIPAMENTOS", len(headers))
    linha = 3
    for equip in equipamentos or [{"CODIGO": "SEM_EQUIPAMENTOS_EXTRAIDOS", "DESCRICAO": "Nenhum equipamento extraído."}]:
        ws.cell(linha, 1, equip.get("CODIGO"))
        ws.cell(linha, 2, equip.get("DESCRICAO"))
        ws.cell(linha, 10, f"=IFERROR(I{linha}/H{linha},0)")
        ws.cell(linha, 14, f"=IFERROR(C{linha}*F{linha}*G{linha}*(J{linha}*K{linha}+L{linha}*M{linha}),0)")
        linha += 1
    ws.cell(linha, 2, "TOTAL")
    ws.cell(linha, 14, f"=SUBTOTAL(9,N3:N{linha-1})")
    for col, largura in zip(range(1, 15), [18, 55, 14, 20, 45, 10, 10, 18, 18, 14, 16, 20, 16, 18]):
        ws.column_dimensions[get_column_letter(col)].width = largura
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws = wb.create_sheet("MOB_PESSOAS")
    headers = [
        "CÓDIGO",
        "DESCRIÇÃO",
        "QUANTIDADE",
        "CUSTO UNT PASSAGEM",
        "CUSTO UNT ALIMENTAÇÃO",
        "CUSTO TOTAL",
        "OBSERVAÇÃO",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    _estilizar_cabecalho_simples(ws, "MOBILIZAÇÃO DE PESSOAL", len(headers))

    # V13.5 - A aba MOB_PESSOAS deixa de ser uma tabela curta e vazia.
    # Ela passa a trazer a lista de colaboradores da aba PESSOAS como apoio,
    # mantendo QUANTIDADE, PASSAGEM, ALIMENTAÇÃO e OBSERVAÇÃO totalmente manuais.
    # A decisão de quem realmente será mobilizado continua sendo do orçamentista.
    linha_inicio_pessoas = 4
    linha_total_pessoas = 104
    for row in range(linha_inicio_pessoas, linha_total_pessoas):
        # MOB linha 4 corresponde à primeira linha de dados da aba PESSOAS, linha 5.
        linha_pessoas = row + 1
        ws.cell(row, 1, f'=IFERROR(IF(PESSOAS!A{linha_pessoas}="TOTAL","",PESSOAS!A{linha_pessoas}),"")')
        ws.cell(row, 2, f'=IFERROR(IF(PESSOAS!A{linha_pessoas}="TOTAL","",PESSOAS!B{linha_pessoas}),"")')
        ws.cell(row, 6, f"=IFERROR(ROUND(C{row}*(D{row}+E{row}),2),0)")

    ws.cell(linha_total_pessoas, 2, "TOTAL")
    ws.cell(linha_total_pessoas, 6, f"=SUBTOTAL(9,F{linha_inicio_pessoas}:F{linha_total_pessoas-1})")

    for col, largura in zip(range(1, 8), [18, 55, 14, 20, 22, 18, 45]):
        ws.column_dimensions[get_column_letter(col)].width = largura
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws = wb.create_sheet("MOB_RESUMO")
    ws["A1"] = "MOBILIZAÇÃO"
    ws["A2"] = "MOBILIZAÇÃO - Resumo"
    ws["A5"] = "Código"
    ws["B5"] = "Instalações"
    ws["C5"] = "Custo"
    ws["A6"] = "MOB_PESSOAS"
    ws["B6"] = "Mobilização Pessoas"
    ws["C6"] = '=IFERROR(INDEX(MOB_PESSOAS!F:F,MATCH("TOTAL",MOB_PESSOAS!B:B,0)),0)'
    ws["A7"] = "MOB_EQUIP"
    ws["B7"] = "Mobilização Equipamentos"
    ws["C7"] = '=IFERROR(INDEX(MOB_EQUIP!N:N,MATCH("TOTAL",MOB_EQUIP!B:B,0)),0)'
    ws["A8"] = "TOTAL"
    ws["B8"] = "Total"
    ws["C8"] = "=SUM(C6:C7)"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.sheet_view.showGridLines = False


def remover_abas_mobilizacao_existentes(wb):
    """Remove abas antigas de MOB/DESMOB antes de recriar a V12.4."""
    abas_mob = [
        "MOB_EQUIP", "MOB_PESSOAS", "MOB_RESUMO",
        "01-MO Ordinária", "05-Relação Equip", "06-Mob e Desm",
        "08-MOB-Resumo", "09-MOB-Pessoas", "10-MOB-Equipamentos",
        "11-MOB-Transportadores", "12-MOB-Parametros", "13-MOB-Velocidades",
        "14-MOB-Checklist", "15-GUIA_MOB_DNIT", "Jan2026", "MANUAL",
        "copie aqui os equipamentos", "SICRO", "SICRO (2)",
    ]
    for nome in abas_mob:
        if nome in wb.sheetnames:
            wb.remove(wb[nome])


def criar_modelo_mobilizacao(wb):
    """
    Cria o módulo V12.5 de Mobilização e Desmobilização.

    Correção V12.5:
        A rotina NÃO copia mais a aba MOB_EQUIP da planilha-base, nem qualquer
        tabela operacional de equipamentos. Isso elimina o risco de trazer
        equipamentos que não pertencem aos códigos SICRO informados.

    Diretriz:
        - máximo de simplicidade;
        - não validar nem remover abas de canteiro dentro da mobilização;
        - apenas MOB_EQUIP, MOB_PESSOAS e MOB_RESUMO;
        - MOB_EQUIP é recriada do zero;
        - equipamentos vêm somente das abas EQUIPAMENTOS / EQUIPAMENTOS_POR_COMPOSICAO;
        - decisões logísticas e custos permanecem manuais.
    """
    remover_abas_mobilizacao_existentes(wb)

    # Regra definitiva: a mobilização é sempre gerada do zero.
    # A planilha mob_e deixa de ser fonte de dados; ela serviu apenas como
    # referência visual para simplificar o modelo.
    criar_modelo_mobilizacao_manual(wb)
    logger.info("Módulo V12.5/V13.5 de Mobilização criado do zero, sem copiar equipamentos da planilha-base.")

    faltantes = [nome for nome in ["MOB_EQUIP", "MOB_PESSOAS", "MOB_RESUMO"] if nome not in wb.sheetnames]
    if faltantes:
        raise RuntimeError(f"Falha ao criar o módulo de mobilização. Abas ausentes: {faltantes}")


# ============================================================
# MÓDULO V13 - CANTEIRO DE OBRAS MANUAL BASEADO NO MODELO CANT
# ============================================================

ABAS_CANTEIRO_MODELO = [
    "CANT_Princ",
    "CANT_Princ Container",
    "CANT_Industrial",
    "CANT_Complementar",
    "CANT_Resumo",
]


def localizar_arquivo_modelo_canteiro():
    """
    Localiza automaticamente o arquivo-modelo de Canteiro de Obras.

    Diretriz V13.2:
        - o modelo oficial do usuário é a planilha cant/canteiro;
        - as abas de canteiro são copiadas preservando fórmulas e formatação;
        - referências externas do modelo são substituídas pelo valor salvo no próprio arquivo,
          evitando vínculos quebrados quando o usuário não tiver as bases externas.
    """
    candidatos = []

    try:
        base_script = Path(__file__).resolve().parent
        candidatos.extend(base_script.glob("cant*.xlsx"))
        candidatos.extend(base_script.glob("CANT*.xlsx"))
        candidatos.extend(base_script.glob("*canteiro*.xlsx"))
        candidatos.extend(base_script.glob("*Canteiro*.xlsx"))
    except Exception:
        pass

    try:
        candidatos.extend(Path.cwd().glob("cant*.xlsx"))
        candidatos.extend(Path.cwd().glob("CANT*.xlsx"))
        candidatos.extend(Path.cwd().glob("*canteiro*.xlsx"))
        candidatos.extend(Path.cwd().glob("*Canteiro*.xlsx"))
    except Exception:
        pass

    unicos = []
    vistos = set()
    for caminho in candidatos:
        try:
            chave = caminho.resolve()
        except Exception:
            chave = caminho
        if chave in vistos:
            continue
        vistos.add(chave)
        if caminho.exists() and caminho.is_file():
            unicos.append(caminho)

    return unicos[0] if unicos else None


def remover_abas_canteiro_existentes(wb):
    """Remove abas antigas de canteiro antes de copiar o modelo-base."""
    abas = set(ABAS_CANTEIRO_MODELO) | {"GUIA_CANTEIRO_SICRO"}
    for nome in list(abas):
        if nome in wb.sheetnames:
            wb.remove(wb[nome])


def formula_canteiro_deve_virar_valor(formula):
    """
    Identifica fórmulas do modelo de canteiro que dependem de arquivos/abas externas.

    Mantemos as fórmulas internas entre as abas CANT_*.
    Fórmulas que dependem de CPU_ADM, ITENS_SINTETICO ou Base - Material são trocadas
    pelo valor calculado salvo no arquivo-modelo, pois essas bases não fazem parte do
    módulo manual de canteiro.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return False

    texto = formula.upper()
    referencias_externas = [
        "[",  # referências externas como [1]CPU_ADM e [2]Base - Material
        "CPU_ADM!", "'CPU_ADM'!",
        "ITENS_SINTETICO!", "'ITENS_SINTETICO'!",
        "BASE - MATERIAL!", "'BASE - MATERIAL'!",
    ]
    return any(ref in texto for ref in referencias_externas)


def copiar_aba_modelo_canteiro(ws_origem, ws_origem_valores, wb_destino, nome_destino=None):
    """
    Copia uma aba de canteiro preservando layout, fórmulas internas e formatação.

    Para manter o modelo manual e evitar vínculos externos quebrados:
        - fórmulas internas entre abas CANT_* são preservadas;
        - fórmulas que dependem de arquivos externos são substituídas pelo valor salvo.
    """
    nome_destino = nome_destino or ws_origem.title
    if nome_destino in wb_destino.sheetnames:
        wb_destino.remove(wb_destino[nome_destino])
    ws_destino = wb_destino.create_sheet(nome_destino)

    for row in ws_origem.iter_rows():
        for cell in row:
            novo = ws_destino[cell.coordinate]
            valor = cell.value
            if formula_canteiro_deve_virar_valor(valor):
                valor = ws_origem_valores[cell.coordinate].value
            novo.value = valor

            if cell.has_style:
                novo.font = copy(cell.font)
                novo.fill = copy(cell.fill)
                novo.border = copy(cell.border)
                novo.alignment = copy(cell.alignment)
                novo.number_format = cell.number_format
                novo.protection = copy(cell.protection)
            if cell.hyperlink:
                novo._hyperlink = copy(cell.hyperlink)
            if cell.comment:
                novo.comment = copy(cell.comment)

    for col_key, dim in ws_origem.column_dimensions.items():
        ws_destino.column_dimensions[col_key].width = dim.width
        ws_destino.column_dimensions[col_key].hidden = dim.hidden
        ws_destino.column_dimensions[col_key].outlineLevel = dim.outlineLevel
    for row_key, dim in ws_origem.row_dimensions.items():
        ws_destino.row_dimensions[row_key].height = dim.height
        ws_destino.row_dimensions[row_key].hidden = dim.hidden
        ws_destino.row_dimensions[row_key].outlineLevel = dim.outlineLevel

    for intervalo in ws_origem.merged_cells.ranges:
        ws_destino.merge_cells(str(intervalo))

    ws_destino.sheet_view.showGridLines = ws_origem.sheet_view.showGridLines
    ws_destino.freeze_panes = ws_origem.freeze_panes
    try:
        ws_destino.sheet_properties.pageSetUpPr = copy(ws_origem.sheet_properties.pageSetUpPr)
        ws_destino.page_margins = copy(ws_origem.page_margins)
        ws_destino.page_setup = copy(ws_origem.page_setup)
    except Exception:
        pass

    return ws_destino


def criar_guia_canteiro_sicro(wb):
    """
    Cria a aba informativa com os locais do Manual SICRO V06 a consultar.

    A aba é orientativa e não automatiza o dimensionamento. Ela apenas direciona
    o orçamentista para as seções/tabelas corretas do Manual de Canteiro de Obras.
    """
    if "GUIA_CANTEIRO_SICRO" in wb.sheetnames:
        wb.remove(wb["GUIA_CANTEIRO_SICRO"])

    ws = wb.create_sheet("GUIA_CANTEIRO_SICRO")
    fill_titulo = PatternFill("solid", fgColor="1F4E78")
    fill_header = PatternFill("solid", fgColor="D9EAF7")
    fill_input = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "GUIA DE CONSULTA - CANTEIRO DE OBRAS SICRO V06"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = fill_titulo
    ws.merge_cells("A1:G1")

    ws["A2"] = "Objetivo"
    ws["B2"] = "Indicar onde consultar no Manual SICRO V06 os critérios para preenchimento manual das abas de canteiro copiadas do modelo cant."
    ws["A2"].font = Font(bold=True)
    ws["A2"].fill = fill_input
    ws["B2"].fill = fill_input
    ws.merge_cells("B2:G2")

    headers = ["ABA", "ITEM / DECISÃO", "MANUAL", "SEÇÃO", "PÁGINAS", "TABELAS / FIGURAS", "ORIENTAÇÃO"]
    linha_header = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(linha_header, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    registros = [
        ("Todas", "Requisitos legais e técnicos", "SICRO V06 - Canteiro de Obras", "1.1", "19 a 24", "NR-18, NR-21, NR-24, CTB/CONTRAN, CONAMA", "Verificar exigências mínimas legais, conforto, segurança, áreas de vivência e limites de circulação."),
        ("Todas", "Planejamento e organização do canteiro", "SICRO V06 - Canteiro de Obras", "2.1", "26 a 35", "Figuras 2 a 9", "Avaliar layout, interligações, fluxos internos e relação entre sistemas administrativo, técnico, vivência e industrial."),
        ("Todas", "Natureza e porte da obra", "SICRO V06 - Canteiro de Obras", "2.2", "38 a 43", "Tabelas 4 a 7", "Classificar a obra antes de escolher o tipo de canteiro e as instalações necessárias."),
        ("CANT_Princ", "Canteiro principal fixo", "SICRO V06 - Canteiro de Obras", "2.3 e 2.4", "43 a 64", "Figuras 11 a 26; Tabelas 8 a 18", "Usar quando houver canteiro principal montado in loco ou adaptado; preencher os parâmetros conforme projeto e efetivo."),
        ("CANT_Princ Container", "Canteiro principal em contêiner", "SICRO V06 - Canteiro de Obras", "2.3.2.2 e 2.4.6", "46 e 64 a 65", "Tabela 18; Figuras 27 a 32", "Usar quando o padrão móvel/contêiner for mais adequado ao caso da obra."),
        ("CANT_Princ / Container", "Áreas mínimas referenciais", "SICRO V06 - Canteiro de Obras", "2.4.4 e 3.1.1", "54 a 62; 71 a 74", "Tabelas 9 a 16", "Consultar áreas referenciais para escritório, seção técnica, sanitários, vestiários, refeições, cozinha, alojamento, residências e lavanderia."),
        ("CANT_Princ / Container", "Áreas de referência por tipo de obra", "SICRO V06 - Canteiro de Obras", "3.1.2", "75 a 89", "Tabelas 19 a 31", "Consultar instalações de referência para obras rodoviárias, OAE, ferroviárias, hidroviárias e intervenções pontuais."),
        ("CANT_Complementar", "Canteiro complementar / instalações adicionais", "SICRO V06 - Canteiro de Obras", "3.1.1.5 e 3.1.2", "74 a 89", "Tabelas 19 a 31", "Preencher manualmente quando houver apoio complementar ao canteiro principal."),
        ("CANT_Industrial", "Instalações industriais", "SICRO V06 - Canteiro de Obras", "3.4", "96 a 110", "Tabelas 38 a 45; Figuras 33 a 35", "Usar para instalações industriais específicas, como fábrica de dormentes, centrais de pré-moldagem, estaleiro e pátios."),
        ("CANT_Industrial", "Composições industriais", "SICRO V06 - Canteiro de Obras", "3.4", "98 a 107", "Tabelas 40 a 44", "Conferir composições e premissas antes de preencher custos industriais."),
        ("CANT_Resumo", "Resumo do canteiro", "SICRO V06 - Canteiro de Obras", "3.6", "113 a 116", "Equações 3.6.1 a 3.6.4", "Consolidar os resultados das abas CANT_ conforme o tipo de canteiro adotado."),
        ("Todas", "Fatores de equivalência e ajustes", "SICRO V06 - Canteiro de Obras", "3.2", "89 a 95", "Tabelas 32 a 36", "Consultar FEAC, FEAD, FEAT e fatores k1, k2, k3 e kCI quando aplicável."),
        ("Todas", "Guias de dimensionamento", "SICRO V06 - Canteiro de Obras", "Apêndice A", "123 a 128", "Guias de dimensionamento", "Usar como roteiro final para conferência e preenchimento manual."),
    ]

    for row_idx, registro in enumerate(registros, start=linha_header + 1):
        for col_idx, valor in enumerate(registro, start=1):
            cell = ws.cell(row_idx, col_idx, valor)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    larguras = {"A": 24, "B": 36, "C": 32, "D": 16, "E": 16, "F": 38, "G": 75}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    ws.freeze_panes = "A5"
    ocultar_linhas_grade(ws)
    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A5")
    return ws


def copiar_modelo_canteiro_de_arquivo(wb_destino, caminho_modelo):
    """
    Copia as abas CANT_* do arquivo-modelo mantendo o máximo de fidelidade.

    Essa rotina substitui o esqueleto simplificado anterior. O arquivo cant passa a ser
    a base visual e lógica do módulo, preservando fórmulas internas e formatação.
    """
    remover_abas_canteiro_existentes(wb_destino)

    wb_modelo = load_workbook(caminho_modelo, data_only=False)
    wb_modelo_valores = load_workbook(caminho_modelo, data_only=True)

    for nome in ABAS_CANTEIRO_MODELO:
        if nome not in wb_modelo.sheetnames:
            continue
        copiar_aba_modelo_canteiro(
            ws_origem=wb_modelo[nome],
            ws_origem_valores=wb_modelo_valores[nome],
            wb_destino=wb_destino,
            nome_destino=nome,
        )

    criar_guia_canteiro_sicro(wb_destino)
    return True


def criar_modelo_canteiro_manual_fallback(wb):
    """
    Fallback simples caso o arquivo cant.xlsx não seja encontrado.

    Mantém o programa executável, mas registra que o usuário deve colocar o modelo
    cant.xlsx na mesma pasta do script para obter as abas completas e formatadas.
    """
    remover_abas_canteiro_existentes(wb)
    for nome in ABAS_CANTEIRO_MODELO:
        ws = wb.create_sheet(nome)
        ws["A1"] = f"{nome} - MODELO NÃO LOCALIZADO"
        ws["A2"] = "Coloque o arquivo cant.xlsx ou cant(1).xlsx na mesma pasta do script para copiar o modelo completo com fórmulas e formatação."
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
        ws.column_dimensions["A"].width = 120
    criar_guia_canteiro_sicro(wb)
    return True


def criar_modelo_canteiro(wb):
    """
    Cria o módulo de Canteiro de Obras V13.2.

    Regra:
        - copiar as abas do arquivo cant.xlsx/cant(1).xlsx;
        - preservar fórmulas internas e formatação;
        - manter tabelas manuais para preenchimento pelo orçamentista;
        - acrescentar GUIA_CANTEIRO_SICRO com locais de consulta no Manual V06.
    """
    caminho_modelo = localizar_arquivo_modelo_canteiro()
    if caminho_modelo:
        logger.info("Copiando módulo de Canteiro a partir do modelo: %s", caminho_modelo)
        return copiar_modelo_canteiro_de_arquivo(wb, caminho_modelo)

    logger.warning("Modelo de canteiro não encontrado. Criando fallback manual simples.")
    return criar_modelo_canteiro_manual_fallback(wb)


def separar_material_e_transporte(descricao):
    """
    Separa:
        descrição do material
        x
        equipamento de transporte

    Exemplo:
        'Brita - Caminhão basculante'

    Resultado:
        descrição = Brita
        transporte = Caminhão basculante
    """
    """
    Mantém a descrição original completa e extrai o equipamento de transporte
    usando o último separador " - ".

    Exemplo:
    "Cimento Portland CP II - 32 - a granel - Caminhão silo 30 m³"
    DESCRICAO = texto completo
    EQUIPAMENTO_TRANSPORTE = "Caminhão silo 30 m³"
    """
    descricao_completa = str(descricao or "").strip()

    if " - " in descricao_completa:
        partes = descricao_completa.rsplit(" - ", 1)
        equipamento = partes[1].strip() if len(partes) > 1 else ""
        return descricao_completa, equipamento

    return descricao_completa, ""


def limpar_codigo_original(codigo):
    codigo_txt = str(codigo or "").strip()
    if codigo_txt.endswith(".0"):
        codigo_txt = codigo_txt[:-2]
    return codigo_txt


def codigo_dmt_valido(codigo):
    """
    Valida códigos reais de itens dos blocos E/F.
    Remove linhas estruturais do relatório, como nan, Obs., linhas vazias etc.
    Mantém códigos originais como M1954, 6416224, 0000028 etc.
    """
    codigo_txt = limpar_codigo_original(codigo)
    codigo_norm = normalizar_texto(codigo_txt)

    if not codigo_txt:
        return False

    if codigo_norm in {"NAN", "NONE", "OBS", "OBS."}:
        return False

    if "OBS" in codigo_norm:
        return False

    # Código válido precisa ter ao menos um número.
    # Isso mantém M1954 e 6416224, mas remove textos como Obs.
    if not re.search(r"\d", codigo_txt):
        return False

    return True


def descricao_dmt_valida(descricao):
    descricao_txt = str(descricao or "").strip()
    descricao_norm = normalizar_texto(descricao_txt)

    if not descricao_txt:
        return False

    if descricao_norm in {"NAN", "NONE", "OBS", "OBS."}:
        return False

    return True


def criar_dataframe_dmt(registros_tempo_fixo, registros_momento_transporte):
    """
    Consolida os itens de DMT oriundos dos blocos:
        E - TEMPO FIXO
        F - MOMENTO DE TRANSPORTE

    Remove:
        - linhas vazias;
        - OBS;
        - NaN;
        - duplicidades.
    """
    registros_dmt = []
    vistos = set()

    def adicionar_item(codigo, descricao):
        codigo_original = limpar_codigo_original(codigo)
        descricao_completa, equipamento_transporte = separar_material_e_transporte(descricao)

        # Filtro para impedir que linhas como nan, Obs. e linhas vazias
        # entrem na aba DMT.
        if not codigo_dmt_valido(codigo_original):
            return

        if not descricao_dmt_valida(descricao_completa):
            return

        chave = (
            normalizar_texto(codigo_original),
            normalizar_texto(descricao_completa),
            normalizar_texto(equipamento_transporte),
        )

        if chave in vistos:
            return

        vistos.add(chave)
        registros_dmt.append({
            "CODIGO": codigo_original,
            "DESCRICAO": descricao_completa,
            "EQUIPAMENTO_TRANSPORTE": equipamento_transporte,
        })

    for item in registros_tempo_fixo:
        adicionar_item(item.get("CODIGO_ITEM"), item.get("DESCRICAO_ITEM_TEMPO_FIXO"))

    for item in registros_momento_transporte:
        adicionar_item(item.get("COLUNA_1"), item.get("COLUNA_2"))

    return pd.DataFrame(registros_dmt)

def gerar_arquivo_final(arquivo_sintetico, arquivo_analitico, arquivo_saida, codigos_desejados):
    """
    Orquestra todo o processamento.

    Etapas:
        1. valida arquivos;
        2. carrega sintético;
        3. carrega analítico em memória;
        4. localiza composições;
        5. executa o fluxo E -> F -> D -> E -> F;
        6. gera as abas finais;
        7. salva o Excel final.
    """
    logger.info("Iniciando geração do arquivo SICRO consolidado.")
    arquivo_sintetico = Path(arquivo_sintetico)
    arquivo_analitico = Path(arquivo_analitico)
    arquivo_saida = Path(arquivo_saida)
    if not arquivo_sintetico.exists():
        raise FileNotFoundError(f"Arquivo sintético não encontrado: {arquivo_sintetico}")
    if not arquivo_analitico.exists():
        raise FileNotFoundError(f"Arquivo analítico não encontrado: {arquivo_analitico}")
    df_sintetico = carregar_itens_sintetico(arquivo_sintetico, codigos_desejados)
    codigos_encontrados_sintetico = set()
    if not df_sintetico.empty:
        coluna_codigo_sint = identificar_coluna_codigo(df_sintetico)
        codigos_encontrados_sintetico = set(df_sintetico[coluna_codigo_sint].apply(normalizar_codigo))
    logger.info("Carregando relatório analítico em memória.")
    dados_analitico, max_col_analitico = carregar_analitico_em_memoria(arquivo_analitico)

    logger.info("Localizando blocos de composições no analítico.")
    blocos = localizar_blocos_composicoes(dados_analitico)
    producao_por_codigo = mapear_producao_equipe_por_codigo(
        dados=dados_analitico,
        max_col=max_col_analitico,
        blocos=blocos
    )
    df_sintetico = adicionar_producao_equipe_ao_sintetico(
        df_sintetico=df_sintetico,
        producao_por_codigo=producao_por_codigo
    )
    wb_destino = Workbook()
    ws_padrao = wb_destino.active
    wb_destino.remove(ws_padrao)
    escrever_dataframe_em_aba(wb_destino, "SINTETICO", df_sintetico)
    escrever_aba_entradas_orcamento(wb_destino, df_sintetico, codigos_desejados)
    ws_comp = None
    if GERAR_COMPOSICOES:
        ws_comp = wb_destino.create_sheet("ANALITICO")
        ajustar_larguras_composicoes(ws_comp, max_col_analitico)
    codigos_nao_encontrados = []
    codigos_composicoes_copiadas = []
    codigos_auxiliares_adicionados = []
    # Listas acumuladoras preenchidas durante o fluxo recursivo.
    registros_equipamentos = []              # Itens do bloco A - EQUIPAMENTOS.
    registros_mao_obra = []                  # Itens do bloco B - MÃO DE OBRA.
    registros_tempo_fixo = []                # Itens do bloco E - TEMPO FIXO.
    registros_momento_transporte = []        # Itens do bloco F - MOMENTO DE TRANSPORTE.
    registros_atividades_auxiliares = []     # Relações do bloco D - ATIVIDADES AUXILIARES.
    relacoes_atividades_auxiliares_vistas = set()
    def adicionar_equipamentos_do_codigo(codigo):
        df_eq = extrair_itens_equipamentos(dados_analitico, max_col_analitico, blocos, [codigo])
        if not df_eq.empty:
            registros_equipamentos.extend(df_eq.to_dict("records"))
    def adicionar_mao_obra_do_codigo(codigo):
        df_mo = extrair_itens_mao_obra(dados_analitico, max_col_analitico, blocos, [codigo])
        if not df_mo.empty:
            registros_mao_obra.extend(df_mo.to_dict("records"))
    def adicionar_tempo_fixo_do_codigo(codigo):
        df_tf = extrair_itens_tempo_fixo(dados_analitico, max_col_analitico, blocos, [codigo])
        if not df_tf.empty:
            registros_tempo_fixo.extend(df_tf.to_dict("records"))
    def adicionar_momento_transporte_do_codigo(codigo):
        df_mt = extrair_itens_momento_transporte(dados_analitico, max_col_analitico, blocos, [codigo])
        if not df_mt.empty:
            registros_momento_transporte.extend(df_mt.to_dict("records"))
    def adicionar_atividades_auxiliares_do_codigo(codigo):
        registros_aux = extrair_registros_atividades_auxiliares(
            dados=dados_analitico,
            max_col=max_col_analitico,
            blocos=blocos,
            codigo_composicao=codigo,
        )
        for registro_aux in registros_aux:
            chave = (
                registro_aux.get("CODIGO_COMPOSICAO"),
                registro_aux.get("CODIGO_ATIVIDADE_AUXILIAR_NORMALIZADO"),
            )
            if chave in relacoes_atividades_auxiliares_vistas:
                continue
            relacoes_atividades_auxiliares_vistas.add(chave)
            registros_atividades_auxiliares.append(registro_aux)
    def copiar_composicao_do_codigo(codigo):
        if codigo not in blocos:
            return False
        if codigo in codigos_composicoes_copiadas:
            return True
        if GERAR_COMPOSICOES and ws_comp is not None:
            linha_inicio, linha_fim = blocos[codigo]
            copiar_bloco_dados(dados_analitico, max_col_analitico, ws_comp, linha_inicio, linha_fim, 2)
        codigos_composicoes_copiadas.append(codigo)
        return True
    # Conjuntos de controle para evitar retrabalho e recursão infinita.
    codigos_processados_fluxo = set()
    codigos_equipamentos_processados = set()
    codigos_mao_obra_processados = set()
    codigos_tempo_fixo_processados = set()
    codigos_momento_transporte_processados = set()
    def processar_codigo_completo(codigo, nivel=0):
        """
        Processa uma composição e seus vínculos técnicos.

        Fluxo preservado:
            1. copia composição;
            2. extrai equipamentos;
            3. extrai mão de obra;
            4. extrai E - TEMPO FIXO;
            5. processa códigos encontrados em E;
            6. extrai F - MOMENTO DE TRANSPORTE;
            7. processa códigos encontrados em F;
            8. extrai D - ATIVIDADES AUXILIARES;
            9. processa auxiliares recursivamente.
        """
        if codigo in codigos_processados_fluxo:
            return
        if codigo not in blocos:
            if codigo not in codigos_nao_encontrados:
                codigos_nao_encontrados.append(codigo)
            return
        codigos_processados_fluxo.add(codigo)
        copiar_composicao_do_codigo(codigo)
        if codigo not in codigos_equipamentos_processados:
            adicionar_equipamentos_do_codigo(codigo)
            codigos_equipamentos_processados.add(codigo)
        if codigo not in codigos_mao_obra_processados:
            adicionar_mao_obra_do_codigo(codigo)
            codigos_mao_obra_processados.add(codigo)
        if codigo not in codigos_tempo_fixo_processados:
            adicionar_tempo_fixo_do_codigo(codigo)
            codigos_tempo_fixo_processados.add(codigo)
        codigos_encontrados_no_tempo_fixo = extrair_codigos_tempo_fixo(dados_analitico, max_col_analitico, blocos, codigo)
        for codigo_tf in codigos_encontrados_no_tempo_fixo:
            processar_codigo_completo(codigo_tf, nivel=nivel + 1)
        if codigo not in codigos_momento_transporte_processados:
            adicionar_momento_transporte_do_codigo(codigo)
            codigos_momento_transporte_processados.add(codigo)
        codigos_encontrados_no_momento_transporte = extrair_codigos_momento_transporte(dados_analitico, max_col_analitico, blocos, codigo)
        for codigo_mt in codigos_encontrados_no_momento_transporte:
            processar_codigo_completo(codigo_mt, nivel=nivel + 1)
        adicionar_atividades_auxiliares_do_codigo(codigo)
        codigos_auxiliares = extrair_codigos_atividades_auxiliares(dados_analitico, max_col_analitico, blocos, codigo)
        for codigo_auxiliar in codigos_auxiliares:
            if codigo_auxiliar not in codigos_auxiliares_adicionados:
                codigos_auxiliares_adicionados.append(codigo_auxiliar)
            processar_codigo_completo(codigo_auxiliar, nivel=nivel + 1)
    for codigo in codigos_desejados:
        encontrou_em_algo = False
        if codigo in codigos_encontrados_sintetico:
            encontrou_em_algo = True
        if codigo in blocos:
            encontrou_em_algo = True
            processar_codigo_completo(codigo)
        if not encontrou_em_algo:
            codigos_nao_encontrados.append(codigo)
    if GERAR_COMPOSICOES and ws_comp is not None:
        if len(codigos_composicoes_copiadas) == 0:
            ws_comp["A1"] = "Nenhuma composição encontrada para os códigos informados."
        else:
            aplicar_formatacao_basica(ws_comp)
    df_equipamentos = criar_dataframe_equipamentos(registros_equipamentos)
    ws_equipamentos = escrever_dataframe_em_aba(wb_destino, "EQUIPAMENTOS", df_equipamentos)
    escrever_aba_histograma_equip(wb_destino, ws_equipamentos)

    # Aba técnica de rastreabilidade: mostra de qual composição saiu cada equipamento.
    # Fica oculta na limpeza visual final, mas mantém a auditoria do orçamento.
    df_equipamentos_por_composicao = criar_dataframe_equipamentos_por_composicao(registros_equipamentos)
    escrever_dataframe_em_aba(wb_destino, "EQUIPAMENTOS_POR_COMPOSICAO", df_equipamentos_por_composicao)
    ws_pessoas, qtd_pessoas = escrever_aba_pessoas(
        wb_destino,
        registros_mao_obra,
        registros_atividades_auxiliares=registros_atividades_auxiliares,
        codigos_tarefas_principais=codigos_desejados,
        producao_por_codigo=producao_por_codigo,
    )
    # V7: troca a matriz ampla da aba PESSOAS por uma visão executiva limpa.
    ws_pessoas = recriar_aba_pessoas_resumida(wb_destino) or ws_pessoas
    escrever_aba_histograma_mo(wb_destino, ws_pessoas)
    df_dmt = criar_dataframe_dmt(registros_tempo_fixo, registros_momento_transporte)
    escrever_dataframe_em_aba(wb_destino, "DMT", df_dmt)
    df_atividades_auxiliares = pd.DataFrame(registros_atividades_auxiliares)
    escrever_dataframe_em_aba(wb_destino, "ATIVIDADES_AUXILIARES", df_atividades_auxiliares)
    codigos_nao_encontrados = list(dict.fromkeys(codigos_nao_encontrados))
    criar_aba_alertas(wb_destino, codigos_nao_encontrados)
    criar_modelo_administracao_local(wb_destino)
    criar_modelo_mobilizacao(wb_destino)
    criar_modelo_canteiro(wb_destino)
    criar_aba_check_orcamento(wb_destino)
    criar_aba_resumo(wb_destino)
    aplicar_limpeza_visual_final(wb_destino)
    arquivo_saida.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Salvando arquivo final em: %s", arquivo_saida)
    wb_destino.save(arquivo_saida)
    logger.info("Arquivo final salvo com sucesso.")
    print("=" * 60)
    print("PROCESSO CONCLUÍDO COM SUCESSO")
    print("=" * 60)
    print(f"Arquivo sintético: {arquivo_sintetico}")
    print(f"Arquivo analítico: {arquivo_analitico}")
    print(f"Arquivo gerado: {arquivo_saida}")
    print(f"Códigos solicitados: {len(codigos_desejados)}")
    print(f"Atividades auxiliares encontradas no bloco D: {len(codigos_auxiliares_adicionados)}")
    print(f"Códigos encontrados no sintético: {len(codigos_encontrados_sintetico)}")
    print(f"Códigos com composição copiada: {len(codigos_composicoes_copiadas)}")
    print(f"Itens de equipamentos encontrados: {len(registros_equipamentos)}")
    print(f"Itens consolidados na aba EQUIPAMENTOS: {len(df_equipamentos)}")
    print(f"Relações na aba EQUIPAMENTOS_POR_COMPOSICAO: {len(df_equipamentos_por_composicao)}")
    print(f"Itens de mão de obra encontrados: {len(registros_mao_obra)}")
    print(f"Itens consolidados na aba PESSOAS: {qtd_pessoas}")
    print(f"Itens de tempo fixo encontrados: {len(registros_tempo_fixo)}")
    print(f"Itens de momento de transporte encontrados: {len(registros_momento_transporte)}")
    print(f"Itens consolidados na aba DMT: {len(df_dmt)}")
    print(f"Relações na aba ATIVIDADES_AUXILIARES: {len(df_atividades_auxiliares)}")
    print(f"Não encontrados: {len(codigos_nao_encontrados)}")


data_hoje = datetime.now().strftime("%d-%m-%Y")


# ============================================================
# MÓDULO V13.5 - CANTEIRO MANUAL COM ESTRUTURA EMBUTIDA
# ============================================================
# Esta versão NÃO depende mais da leitura do arquivo cant(1).xlsx em tempo de execução.
# A estrutura, fórmulas internas, formatação e tabelas do modelo foram convertidas
# em um template interno do próprio código.

import json as _json_canteiro
import base64 as _base64_canteiro
import zlib as _zlib_canteiro

_CANTEIRO_TEMPLATE_B64 = """
eNrtfUtzHDl27l9Jc8YOvlRKIAEkUDcYdrGyKMs9LWkoteNGdE90lMiSumZIFl2kNO2ZmMVd+meMZ+FoLz1xN3dp7f2bLs7BG4kqkmpKoh4dM5KAzMIBTuJxPuDDOX/cuLj815PZxcbw2z9uvFicXW4M/7hxNj2dbQw3ns1PZxfVo9nvq8PF6fRsY3fj4g8bQ0IH9e7G88XJ8cbwcvlq9qfdjRfzkxP4Hfz9/eW/nsOPLxYn8+ONP8Gby+PZEh6fzF5g8ShRv3L5w/wM3ljOX/5QfHC5OC9lP19cXi5O+0/0o+nJ/OXZ6cw044fFcv4H3aSprtzGkc7U1djdeD1bXs6P4jz9w17bx7qk58u5azMZ1HE7X44XJwvdpo3a/qffe97PTFp/08a8l0rdmuoLtV3dfX5urd95R/q0WnONYaHfOXt1+ny2/P7FYnk6vYTKblxfC/Vd0cLatqKIuziu3r6/feAJ92pdvuX4+KQH4PVWoE97wimuw5/IJPSWmlk5Pd1ue2/U1T5eQ+9DtvPWe8Xuz54z3q6BH3pG2N34/XJ6/v3l7MdLW/WPpvfeJhh5/0vBHezBdwbqfNQf9osW7qqlM3gvts6tGqJvN79+HGvH+97u+TKD3oG2fhQw7S03Pj58//loK37be1CfysL4HoyA21nobv0TvINqfWCj6Y4i0LtqwV3Lurj9feUbq+PLEcenNeW+Fyzyi91f/ELPcKYJH3DmuZtntp9Rt/ntq4vL+Yt//dxHz8oZ5ANjzI9re+O9TCYfrXL+9u5OtR+wwu98p/pOYuKfU/3+lFT4fO/52909Y/1Od+k7OEN9f++7jcNffrfx3XfVduXMw+/v/a/vvlvxIPnBdxs69fd/j9n/8P29jVvBjZ+Sxm/RjujB6d1IVQcHinY1v9W14EN1jfo2usYdOA75sFtx76/nwX+fmNI/sw2J2zRKPo2WftqA6N137vdIvP08N0Jvdyi+K9bBB5kePtxA/ji5eH/6NOH+HV2aPqedus94u/Ldds73AZ0/8P7bx7w7cE14naPqDEx/8O2Vz7FNnw9J++OmUX6qlMlPZJh9anz1z0IjXw53b7g/HkZdfxx+PCPxY1htblMPH3jq/ojP2z6q5n6xea/eEDND4XY3ha59WvRZMSg/wnvIH+SO7kdl/d3xfd13uaX9UWOmd2WtfRT7pavn/PfApnqLxeGTRtqf4d2Djc95G/jLpsP77igfxMD6KO/fv18O2V3aDL5b9t0XYP3ZNPFOTnN3eDK761PWxzgxfR5j8654xuqr8h82PiNfL5/ZvcrPdSy9A6b5x3GWcqfN3I+fbf8JNvfu3MT/FLalb/ly0CfiXOZ2rki8y4PIt2defNmV/ePdp6vc8WH5p8/plOcOURjuyN7sBzno/uiOtX6jW/zDbHZpIgq5tWr06Nn3T5bzsyOtkNPpj98vF7/fGPJGmNTR4uTV6dnGsN3VlTk5gZ9+uzEi+l344eTh4eOqm1SP9w9HG7v1b3a/3djXz85enZzsEkiNk1TnUhRSI6pLmbyYXc5fL6rjWXX65i/49+L5cuoKo0lhSaqjSWGNLuzR+YuN3QZ/qJNE7DL8GT568/9OZ0ss/8Wrs6P54uzNn5fzxUV1PK3Op0vdtmn1Yv6jFsyx8MYWLrBwhoXfe+1K12llC2fXLPy11vCrk6nOcSJYIoKb+gcZkLG3z3agKUYUv0E7qllBIk8kCpTo5elkawWJawr65+ly/ubPr2cn1ebpm/+8gHfP50eLLSdPJPJaKPV0ocWgPJ1UtRXYXiHw66hrVI+Xx3N8Nl0hto3FTqDsr9/8W/dwBD314JtH44ePH40OHz5+qiEk2dW9SGL1JNThxdSpA5J7m/t8ZxO+Qbu1dZ+67yDX1XZ0svjt9Fj/42xRHU318J0vF65iMtGHQn1Mf9R6QIkKPzjf2Rc7oByCI2CsImGnb/784/y0L1S/bIaESoYEgbVsW7ep2qs2sW9tVTvwr9c7+jtgg4h9Ewb09jbUJntXv2d/sxXepvj29j/D3DCr2t3q/M2fX87PppUUut8RKM29CsPjYHq5WM7wO40ALM30Y6NkwtLJgqWzBUubw11ZUNIUS9Jd8XiJnWNaHS3OLvQM/uY/dFpLYEaC6/OEGxFQyJPl4vX84s1foQvBPGGk4ZOZnl/PYNZ1+SMCo+EraJLpt8T1aiJNiTpdD+QuUaYYndS9qja/hc5X0pNuGSXmFRW36uvF8/nJHL+q1uRID7TZyQ/Tl7PTKrTIf2Rq5KukhR0W+BX19afQC8aRbnTB+mtcTl8tpya9XBxrfZgRBWNp9i+vZrr3Ll7pDvfTsa7J+WKJCjEKgAJNBRozK/s0M/OyTmvIw61GJlgBWCGq8eNHTx92k8NRh2PxyeTX30wePa6ePD58NtEK4aa+5Ib1fbmcnkG100qSrJIkqyTBSjJbyRGlVups+bovJhRLs2JpVizofUAaX2yTN0ZrdTk7enU+s63Z1Um9qr/5D2jn9GS6fDkFQwAH+ePRJP4gWQubrCpNVpUGWyh8VdgtVMV2CP3LVOtUmDr5EduaOvm0NHViVuvU1onndXoxWy77WudZU3nWVJ70uBEV5WFHaj0DN9aycMPYpMdZugtpM06pW1Uaan6fpsdZekRlPK7NvAezVDe/uHzz73ry1n13UY3tClFN9Yox1spe6r8PFsuz2dHsGCZNXV8zVWJ5WlXHc6Mo/eMn09dz/EDupTGVqYUkU3vLzRSNmUkgvfGr2Vx/4UfTSz28TkJB8Ohw9lqPOyOi0ibiKc5L7p3OvIODZOrros0N98KogYH/VTOsDvRUe6pNEWgNN+YZPCJ6Zal365qw6seqe7ZhP984eVir5GGXPpTJw0njpyIUc5ClRw3MLd2zavOr061QF51Z7zbCyI4SHSQIzO0mPWlI8o0PQro1xcMk8kQbR/Pp/KLqZkfzU/2PIAge74GFTtg2inUN9g9qtT0OD7rogdzuwoNJQ9Pe1jRX9vhJ02S/YbiMm15XfbN8PtVzDFRYf+s3f0VDY66725G2Zio9Lb75c/VkdrzUvXV6/+uRX78abowBrdU3P1XHSe8+inv3ke3dL5Le7Uppsb9dvALrBl4yJoNbxPYbP7rMopuluyax+UaNdGu2Ma0wvYdLtVO5TO0ArSWr8nTMNH41tSWhlWaWD1uSyucwvXAMCvM2MxNLkxppzAwSLwDTe0SbnQ3d0T1jR3cC3VeZsZDwaTSFRNOGF8DqVACubHr6WSQTDkK++flU2+7j8ZMtB7hYirjSZMdSzMUa/GondsGEeswqPUnMprpWC41XL/U/fM7l4hI6F3xeDWqXWjNhdmNNKrZJxTapWOi4z+bnEVp0xTC04VDttkVjzPsa1y2b1WHWA1zAbNaI8b7VcZhYHVqWney8qvfhV/Wg0f/tMrMMmBxmk51NcpseMdEzMw5jM8OVizYl94UmyS5NjlhbWNgP04X9MCzso2xh1/ozprnupNXk4nyGs1eoSotNDHVJ012WnrA2nXKZLM5MF5VU1ebl9DngSEa3ditFQppt6YrGGWLLL9zMDVBmlqos3WXpiU/bCvE6Xpgn//Jq/np68uY/zZylO+v/wc56PLs4cj34XnUwGXVuhPA66appsuPp4ONXGrO2C1/E5iYMGmNz6c9g1mrulhpmcEyW7iBdgwnE7HYCjPpvLjWY+IMbnT8enby60I2FojVWunzzn/Oz2fIiiKCZCJqJoCii8SJWrDmK+2md46j6ejz2yuOp8niqPJ4qT/hfuyriBM7UgOwyY2Gad3B2s7ZphgSro/nr+Yn+hv80enQfjDpmVgwuYlkPfJIZE5IjctNr+SUoCzsCTJ8aG/6Iy6KWqD/UiUbL327+jmor5HeN/uN//jIa6790dxlXW7/R/4Law/aLmWS4wtlbj70Ljd3dGNSf286Qr+1+Cgw/TkyLVbrwZenOp42FghIe7z8dVo+huBODu/WMoAe/lQcLgDbjlpXuhqCf6YlOw47FbnWG/cRkA8Q+WixhFV+cVlPYz9DQ+CJZ1nUtzZcSMKYeaqVPcS34v7Mw/+t37P4SvGNH13dnQQvHU78MjPEdPSRdIboSubq+O0OFYTneFBXJqDvAYqCLT6GHO2T5ALP/5fK46j0aCRink4uj5fwSzAEYphcz038u3/x0dKatIN0Q0wnx3T3eDhTf2WQDrk052HoygwYfjib3nj7TXYO3u4prU5Xtct0XcHeSG8tSuOHLW1Nj+NnXdS1EqC3BjQRbP4or7YtZqN/R4g/zsx+gWtJUC21FMuB8e7OGWqlgNuCz0eG9sa4U0VXVteH131Zm96mxm3U01SE1NeLtDvwpQ71oXC+YA3DLC7/ORagNPNhrBoRsQ2U299udfRFVCH83+pWuj8bKxNbnR9xw0goz202uYk1ascZWTIUq6azGVwkW9/3p2Q/QRWEmRRBjoYurHTN2ddv2awaPRvv3/llXrd5t21WVYmmlmKkUI6FSLNYTTISj0+caBJkPGKrCTVUofLFQC3x/9PU+1oJi90k+Fk/F83734bF44cYeDCkN+vW/7D6Zq4awvcd2HhlpRGBdDk3fcV0Hd0ptZURaGXj/XqhIkhwJa+nPj82aG3UZeLInB0xsw+a3k95i15081eKlXnrMQMJtciu9TaW3/e7RxqqQK7bkZLA2ahjPxhwXfs/WWJM+bUBUl6VHbTTLwT48WA9r1//E/LNmbJuaGGmya1MToyVr5l6z6uIrTwwK4bZMkq4jPm336mm6VdDS2LAWtgzas6wxK7WsWzNFnC5+1KvbC73Q6N+bLRR4Qmo2kHJXmPkfczgdCGE/fwc5tNELfmtzDlo3/lv/gfFn7gO3MBS72fmbv17MjYl75HZJnFw4vmnAKndiYahS4mVCUolBS7xM1pcZD+8Wxt/jF/Mj3ZOCGLSD+IAwL4fDJ24HMjSPgzapnrS9KDuUGYV5V9T4Jw9i42HdwtB6tjhfvFxOX8wjyWAmMT27ecFg99QD0Xi5OsOnDtp0xKbJUQsj6sEr/fUuIxE6U2gzzAmwSVc8HO7oud8LsIOyDdNjGw/KVvZHzUpg4gaJTAdJuvXVpjC+VVcPEtUfJJmx1SbG1kjW6SCBdI4+Mc+NETNuOsxzg8TkjSRZNUrgiVQDqdyMKPF4hQ9asSvM/ifmmGFjckaSXjUI4A3S0kEjfbk6hzI+aEK5OqdhLEgayabQ0SGTSP/1x5BWcqCkLwdeaNWAEV8OK/dczN/WIwa6hyuOuc7rimPYeV1RvNA9IVNENbJJVwC33dMVIa6BUUmOSfmWwzn7UqSrRJbusvRI+u0qszMY0vb3bfb7Nvu9GS+Hk9HTavx4f3L4bPT0u7PR+Ym2U5dm8wy++awHbg3O8Mj2u7NNu9NWTc/1n7iBC2eXU9TBoCKEgsWvVTNoqy0/9mQ69tJkJ9OxJ3tjzw05fOLNEm3eDqrN0//+ry1drdGBlzbG1wBTuVUm/O7CAXg4srsIv574X49UfR3T3pgg+O4eWvi250BOW7t5zTx/eDA5PHx8uKlf397Ub9wndb21W3uLaaTIldY6vrKHRruTRBJJJJVEQBLpS6Ir7G98sIdmuCufJuXTtHwK5dN++c3VxjS+s4c2tRPVJKKaVFQDopq+KLbKRMYne2gpOwEwAwQBLBXAQADrC+BXGcHKcC7AYnWCOIB3L4ingjgI4n1BYpWJi0/20NJ1AkSiKpEKECBA9AW0vcXCCUAbGi0rJ6BNWtCmAloQ0PYFyJXLhxMkjaDoo8tEkEwFSRAk+4JUvJ64sg0Too2+t0rKVmnZCspWvbIJcvniRcaWbx7soZlkBWBW6FH2DS9Cp7UM/WdBCImXHy/BDO9oFcOsNpJAMgkEJZCCBIrNAFbgrhC2fBy+T7/5GiahIfxa/0C0VpIz3oW0svzbegYbdvHbunjcxnuynL2e6ym6ehXvGdqdwqndKLSrHf4kEjDOM7o8Q0sxh024M3bPHe9cZ43arRbPL+eWUFPYbGRRtVheLZZXi+XV8pQoZcvgeRk8L4PnZWTg+sVs6VqzcCcdus2jZ2Z58oso/jBmv6TpLqQdnaf1guAs5WS1OOhpVoazHFonJM3oTKl70KXu43lG66QhUDbKr6aXy+nrNz8hGWtZPvCxxzuzYF0cOytCkV00Iao20F5qhR1iBPtko4PqfnVwXasllIG0Jad5t18Pqu5yVZOMlUgyXiJJmYkEKU55yUGrhGZazTI6U8AeaPeeGa1OrQSH2/Fcf7bZ2Zv/mMZKdJ81dBv3wNfAN92MJ93SPehZus0H0cOoR0bbtaYQ7IvjnoKyvkiyvkiyvkh8X4y0kvc1kvc14vtapBDsZ6Mx9oPQBkNbwo192IAO0xBsqIxDxWm6V5Klu5B23DU8kXnz1+P5S9BENwPL0JoBzM7e+M6vX2m9zY+nANLaxhaOGynLGZykfXOmrTtjCJnHnXn89NXzSzNZt8yKxL4w0X9Xf1d9twH/+7vqALPtIRv+u/ZLBT6RjVAD2hBpQXJnyzl8/M2jbhNe2obf7bKt+zBq8Z2JeQe3DxzSNoLgEMqe+KA+fzs9AxBzr4ItWF0Iq051Z4OPIHZrpRObpHo2+WbL43UrnQ4Y2xaD2u+u6eY1YX1q7frkyUiuK2QZnfmVWZN00VpOa9cjw1PCD2+Gg/v8M/P5N8fjrUgQywWxXJCxHhu1vc/qbSPYy0Lygj0mgzUGeAqhp13MXr460yMIJ1mYBvWIHOEJUC3u0Rp6s2ULEU9Jau2qYDhI22NQaTgg2iT3ebVd/c9ffn20PT7yALIjnm7kf98aXoWVR608/36bvy9T/lLIcG8gc2Y8fgIDaDUdIIysJhtZTTaymmxkIblGlx8mBM+PcR8ny+jMb/Y2N8FG2d5v5Lb9Tls7m4GY/jcwnW7HaU62tuKMfS52zJc2s8qEGOoOE6ohA6ka1kjJpdxtlRkSRi7Iv2feNQ90I5Dbqlux9+3m7+Bcwn+75GxvZ1MnO5PswknfDk5NbgozpJzyF9wnno5jTI9xnqF/77qydPpk+U/SjBFvRDjs/s3uxuls+XJ2vDE0i8LQTOOwGzbEzauNsc7DHG1rDI3BAYdFQzwM2hjzYYfnEKwZIgcDTkeGnTkh0Uvm0KybG2OdiYcuVA6Rcwab0kPcQwb6ybAzFBShSxGmFF2wOeHQnWhoehKcjg7xOBNYLkNksWjDf9hhDtfV46amww63ylr9Dm62j6SWihsiY6Hz8JBvX7cQd+tga3GIm4FwNDfEczeg7Aw7Q9uBn8JbTTtEDhEuPMPOrj6NfoxEobF+jCwLof9hTiS4fo0bWbpc3GEf6fYjyQ4OvoedOfwWoGKsuVYE6kGrAbfEiVabXsZ/A0eVJ9//fn58+cMF3P8YwR+Y3BgyYJfQutHzy5+0qOgJJfGTcfykGQjBaoEPuviB0JM3cU8m0RPcLG71KGnhyUH0BHcV3U8epD/RNpz5wT9G+cplw/Wvxe+//2EG12SwYaTGS04zc2+G0AH+WHehUi4t5rZJbjOgOlNPnYVX9Wgq5bJiLi/mqlIuS6qgZxeoApPFV9MC2IDpTF6sF29LuaIuyBJFWbKoBKkKBQA4K+m2ZuVsWc5W5Y9ZLpuUyyblsiktVZvy8suinN2Ws0vfBGb90svlPkRKnQiujS1nsz/Mvj/XlpXu7jALwyW03jUrOA507IPowlXdpBeu5Md34QpYcO/uwhWhX25c/ewbV5K++xtXhZtUVH25OvXl6tSXq1M3vDr15S7Ul7tQX+5CfbkL9b7vQkkS3YUyCbx8RAeqtekPcxeq/nL/6Y7cf6q/3H+y95+uPqO4+eWm9ecT+ErheMLcflp1OoFPe4cTeO/pKnaMpP62k9mPbPxdpyjdhTQzN294MiUcZOkHKBvPt/xhAt6P2pswEZ+UMJFcjPKsLkwwKRQdNFwjm5b761H+mATIxEzssi1/RIKPE670gbuVFR+QLOMJX6tieaoBz647N5lVzy0ppnSEoqEqwTMUmp6hmJrhEQrRc2lodbuaL+R032a6bzPdt5nu23Q6xmtYWrEyUawMipWxYiHBa41G9ITPlFesjBUrtWJloliZk2RRwNqTJ7YrZfnkqTp+dX6yiHSHwoEsnB4/MWUappKGqdAwFTcMEpyTptXrS9QwFTcMDiBU0jBlGyZDw9TKhs0q6jvHxc2bqcrNxK3mhOhlO4a/cGY7RpYOF9Jsx/Bp2zHwhtrehJNYf5Bp9Qf/dERhvHUmlCRiQLm2XNtw98yrj5Nt/ZtYfZzkjPwDdy3utgbcijNLU7H+kSVejlvFanN6pZleaaZXmumVZnptjF6bRK9N0GsT9UtM1I1s20Gk1iZWa6PV2iRqbfIrJ1h8rtZ+Z7yW5poVmuuT9Jy+WKYvlumLZfpimb640RdP9MWDvnisL16c+U0ZTmNca4wnGuvdkjng/P3M/KZm/Zmf96mCTp8i06fI9CkyfYpMn63RZ5vosw36bGN9tsVx3cbqbLU620SdbX9ct+9nXLcreqfs8SKdNmWmTZlpU2balJk2zSrDk1WGh1WGx6sMJvrLJ49XGa5XGZ6sMlz1lk+ubm355CvWFTwu9QRMqy2RrSlZOly3tNoS2ZpibkhORLKmiLCmmKuN7oIXScguVlkiXlOEXlNEsqa4+5IRzUWQW2C5iBUrhrlcOREJg0cEAo+5FBndtiy0KGbvCKpbRJMW9Yk74jZ4OyLQdtggWDDm5qZuUbI+ibA+iXh9gkR/fhDxAiX0AiWSBap/UdPc3Xz384NoVrSZxVxd19mzhStLh5uerrNnC5e5wDkRycIlwsIl4oULEryhte4atG29IuN1S+h1SyTrlrvdGW5PmQufmSIbq7MLOC2aVi9NO0tqa3Zrjmq7ny1TIixTzaDmG8n90ZTi5S99tuGWaEztMRdKkd/FgAIBt2VkdCMUH2lENHwQPfpHfHT63/+VXhe9LhtMZATELG2YG8A2EmJ780C0W1ueb+RqMMF3uhnchrmYT/3+Yu/KuLlouncg1P0DEXUxeUuMMpHStyZY8GsN+JGhHWoBopmXrtaSzjYN6yzQzoTKhMDvT2cX5sTJnAs7QQpwAI9uua4kp/kLqpZ55u+Wuucke06v45HjKDiUmXqurL+8l27bpMmuTbdt8Abpi83Rs/ujbsv7lmib1P1FlsabqOhhQwp/z9Q8Z/5WafJ7lv3ep0W4M3ojkm6bOs9Ik12bOs9oAwfYlmJZeW02cLM0Xg3d2MNBaim6eHvTMHSfLk6t4R/7rTi2nRzHZLSr2aqcsF5uaM6M9/css2uYJL2FSf0tzERI2lx/SdM2N0uby5h7utX38Qtb/mB4S4XrmZas3lX3X1S+99imSpr0cXNR8so+bTnz6GDmmb8Vl24XpslOptuFeGHyZn5eZNZRszRer0Q/L7aj4wXL9X5eZJv5efEXFkl6n5Gk1xlpuM2Y+HmRZr86IXhuOJcv5vW3cPmS3kt84JPW5Yt0/cp6a8HLfuNzvMA/AwsZzrZxX3u5OF8sgUMwPbE7sLbjqpQymyY7lW5M4xW/yY+Xs7MLPO2bAokErnQsTuxZ47m7lDM7BaG/neFw+d3pOWT87tSfCaqMZ5ul7WVC2m6Ds9s2XBxcK9xdLplhnwWxy1xsdgsiS+MdQ1rXbhrBe4DX1KcdvCqjkWdpe7WQ7GzqJt7T8rbuo1AnEA94tESNy3eqTa23e7oRW/f1H36tUiwd6Iqv5uzi/b/kZez24/F6RrUfkdkpQXSgqdLhkiY7lQ4XvFo3htHitCQzLclMS8aXsTZ/dvR0t60H+DaMtq1t1J+d8cKPDEU6S5ubetqyGO99W4E9oc2KiBr96/EcOND6z60dPTUaYvQzS4vegr/Po6s/9TpedF1nJOcsw1y5S3jRPsP/hGS8aH8fLedFewa0aofmdqRnPbd02FkOs+c/B65xn/Psmc568h/iPF1gMgf+smcoSy0ZJ9ARXM4zs06fodzqaqHJELjKfX5yxEC+w2xjGZjDMdVYJj94kMr4PLnGiTQmBrLHNXYE5BLRl5PCz1P+sXuTl95sS5klQYKUMkskXtEUyhQl6WWasyjqqS1+7Lb4sdvkqzQYDKQtcoXbVNGWUV0sVNJ+oZJfzdNmgz5LW5gXlSgQlVWZdF2XO3tN3pKn/ND7t4spypTL90NR9qX4ZVWbyqnzFl/D+cXNmct4wj559Oxw9Cuo6Pjxo/Hh5NnjarOpq9P//uv9HwISSUmWKccyoVhOwGj2MLzvncOXmN6zTG9Zpncs21VuM9prec1oM6cZbdlnRuYtQ6Z0AH+jXSIVSDoSb+uYOF16zx+u+Rdu+fc4BiU3EpEEFUvIbvvDZf/CJXmgb02fL5a58wh4okePFK5sYm76O85pftEfb+H3yyerPC4QEtUbEjyUnV3xxxv+hQv+fd9+BfXAW2k7aKwkkvrzIODPg/T9eeDt4563ArAC9drrS26SklP3HQTcd5C++w6kK+eeHAiLlcMS5aROOwg47SB9px1ILV7riQLekNoSCIrhiZzUZwcBnx2k77MjvjKtR8g8mW3CzXPlecxE2zcN9TJFIKchjzkWCV48SN+LB2n7uyP+/r5rHGwIAoZyctpYTurMg4AzD9J35oEXrO3mrbQNMN47CMZgMWVF3A7ctNWmJWm3dvFmLgt06+2bOxnC3ehUofMwfVdXeCBSwjsgCrsthhft14Y1hefbTboMRVJWtKIpK1o1nhWd7EH8cv+XjdpOM1i9na2Yf6NVmb0EexcKdfggCOGBan1rbmYpSYnYBv088GxoI/QfQ1pEbGjUJn6rSz2J9p3TzqpTQK5AqF+vc71qdw91i9uUBq1kSoNWytOgN37x6P4IGBC1qS/LlARDuFZ1I8F/hzZrPBE6Wh/tRQbgAhrPGBpnHy1njpo3PZ8e4d4CPGpqXOahLJqSqInnUMc2iUi9dWV3tQNd2qJN2l6LFkhT0J8mO5qCfkN5XksLxFcKtEAq19EC8WmPFkjNeTdNzrtpOO+m8Xk3JiI2BqkjUrQ71qJqex+uAcXeCqjqETKoei+EDMPH3ptoiy9qIGTaBsI/fQMxER3ouwaaQmwDm3pb/yptIL6Qnuk39W2d6RsGt25DcszehGP2Jj5mx0REfSM187zu0AYCXOysDaTHfmvIu2G/GdJ4es7YZJuMWdqGyMAlC7ZiGhJcSRhS+C2cxXnmuD2wcszx6/l28Jf03a/5usOzJnMk0WSTT5NNPk02+ZidocI0GKMbuwR61rhdAn2a1NzTymHnEGjYlOn/k61tQDpmNTM08+3xw4d7kRZgcXoYHCJ4pYxP5vof4/EcbsX9z19wjXDsd7V6Q3DEyCq4xvpwLQWaabJjJEVsJIJsuL23CrPdnMBdxG345Grghq/FyC38bi10y6nc3ojEIBTOgkQXT96EZKl1zMA6Zn3rGInLVyI4jEvh5YhETmoSMzCJWd8kRiJ0EcbBkwT+QEaAcSw1hRmYwqxvCjO5CsaxGOAyGSMIlkJcBhiX9UEuU9eBcchATtqRYF2Wgl0GaJf14S6ygXMYB5kxjON1XDJPgS4HoMv7QBcpszmM4zHG5QnG5SnG5YBxeR/jcnoVjMPoEl4GTWSk2JYDtuV9bMubG0A4eFkvhQE1QkaQmGJeDpiX9zEvZ1cjOHinqQOC4yyWk4JgDiCY90EwskdTBGcIpVi6RXAx+dXQbuAggyUQDummdwPCcfnzIZxnadr1K6QDq/OtIBznqyFcEMIDNfS2IVwI66F86IoYwoW0CIzLdwXhfLAKC+FCWgWuZgrhRIZzkYVoIBwPEA5z3w7CsRzCiXTHN41mMfIxJKwVJVKLbiwy30uCXQvCiXR/OE1mMSIMUXA9hMNXChBO8HUQDp/2IJxh2U1EcpVKhKtUIr5KJUQZwon4LpUQ2/vgmT9GB6J/m0q8n9tUhhqoG5gw3EVguIuY4Y6JAoQTMcddtLqBbdbAtgfhRHtrEE6Ya1kiuZYlwrUsEV/LwkQC4RpPFQxtkLoNMmuD7EE4Id8RhEMKYUYVVRktU2W0TOUhHHoGkwHCGaLhLUC4Eifx+hCuyFhczX+k2dvZ5NNmk0+bTT4YbaII4VgPwnlWoV0CfdpBOCwMyB9qR0/UO3pNBAhHPIZDTsEtYLhWrMFwxgFbCcPpMZmDuJsEdpi0MgJxlgpYBnEy5WalyU6m3CxJVvqsJ9fzWU9yn/XkeudvdNUBHLUxC9whXHL8I1MTWYKJLPsmMpJSrsRxsokP+5LDIJkaxhIMY9k3jJGVWMRxEncPIMaEK5/FOE6mBrEEg1j2DWKkJBZxnESn4QM9A7ryk9MgmeJdCXhX9vGuFNfBcvCWUBBBw8lKMK9MMa8EzCv7mFe2BSyHmXolYtsbvvA2KTwFvBIAr+wDXuQq5XAOMjFehys5gbsyO9HFI90+3MUYDGsRnYxPdKVKZKQwVwLMlX2Yq+obIDp4mbJ6ULf+FLmOoNZNQjdcAengHcGVHHiorUgs6QahG1JQZ4M2QAEW1CmaH8vRoRaQgDokQd4NUIfsyZ8J6jzJ0q5oId1EQRjeAtQpuhrUBSE8hHC4bVCn0q3aB57XaUFdSIuI6PmOQJ3nd1pQF9IqkEZTUKcy5IuUUAPqRAB1Sr01qNPGQIbqVEr9V7mXtjq1rEJGK9eQR6/jV7zOOEt1xlqqM4/xGGPhCr/i+E7Jrzg+WO1XHB/3/YrX5gqf/jvxK16HS3z4bxJCUDTJ7TOHH2xBzrF43WzDDzPH4nXTu4Rm8m79FpoNFpF5FK9zR9917ugbf2U9itcU3Bk3kVvx+tZcfde5q+/6Zq6+69zVd73e1Xedu/quZd7rZd7rZd7r1SpIERvgylGwVToFhwyHKkyBe6BptqPnCv1/rnEFWOMWVxDjQvHnAwtiPSyWkYWJpxBBi/3Dh89GDyZfV5uyhyzw5STKQZNFOUiOciaY9uiCEL4GXhDCs7J5VjbPIiiIVRAjid+wBmOY92KQEf1yLcowgRhKMIMYhpHnsyVGJ8kZR4ZyVOIcEXkdpIGvpaw2IlOJMpMoUaIsSFQrKYBEJW1SCQGQqEyCQgkFjiGtV5IAaR1LgFTEQKMZx5AiyZAWWIaUXI8KmHAOKUmJgBnrkCLtkJZ4h7REBqQ0KTxjGeY0Q8MzLBANaVPiA9ImKbxJ9ZQxDSlSDWmBa0jZlaxAmlAPaco9pBn5kCL7kBboh5TfhBeIzhdTkiaPCXs0IyNSZCPSAh2RimuQA/FuGxVRMxMaIs14iBSJiLTARETGUkYRpDZWC5TiSIK0zeCIHvvgLV+kREEq7wwisVEvfh4kCaEu3IIYMpoQDOPt2IK0XQ1LIjk8CrJx28AkxOWwyIR4lo2FJlGGiIJbvCtwEmJhKJkFx1COldz08En0Eg8RNCxCaQNCMdlrIMpzmBGhzn2I4oyKAFFCEA4X0yTzvewZPc5cyyg+4+gNF/lDXA+kNFnUpiaL2tRkUZsMOWg9SMF3SiAFH6wGKfi4D1KM68mJ/jsBKU0448B/B5CCqcJRlC3JoRSIFUOa7KDDvJMeR5m893AeRYxvTF0JlTZVRU1VSVNV7PYrNFQlDVXQUJU3VPW8f5FG3ZL7rxFhdR+BsewmfZ7RmV9ZBAa3FHWNIgSGRK5bQWAsO44hjN4IgTGa/75Zi8BYk7/PsiHNWDakWRaZhyDTalyYY2Kc4tYblu2BhQwPwAxzaxOUvgOzIfxRawiGmMVhMKRd3QYGY+06DIaspW+ePnw0qg4e/u9R9fXDp8++ORx1jw9HAMmePv7V46fAP6yrywSOMZX5sFeZE3uVwjGmYjjGyTo4xrP9G57t3/Bs/4bTlXCM0+vBMXwvgWPhl+vhGG9WwTGeGMw8vZmT0ZQI8pRIgahEkGh0NRzjidEMqVhaZjQjW4nw0p0dvhKK8Z6FDDkRHOP5fR1zYadgIpt4UEU4xkXSDpFeCMrMYo5mMS+Yxby9Fhzjba9NKWzmGWzmCJt5ATYjY6kHySA3uaDFU5DMM5DMESTzAkjmpXi7mBupS6XqyvAxR3zMC/gYfcmtR2Uiwckixckiw8kCcbIo4GRBbnRbi2RcP8yJ5GaIWSBiFgXELOg1UBm8xAWJmkkTaRmEFgihRQFCi6aPyoxrMyPEobLYD5tZfHkzBEEpKkNO0x1BZULcAirzLsfcKhkyrCVl6Etvg8pEswaVBTl2KUJG0K2jMu++zKEyITNUFjJcwBX1LlGZJ+44VBYyXOdWfVQWXrLaQlKPQWUyQmWYnaOyVxfgauh0fgExGY71ggI99GJxsrgwF7nQpIjAWJsFqWyzIJWeyeMst4zaM47esJZbS68HxjIfZFm6I5kXMoJuyK4AY/hOCYzhg9VgDB/3wZghA0303wlCwSju1tiBfweEgqkSGDMlOYzSsm34ZYZR8J0MjLXsPYGxltum8rSpPGoqT5rKi2Cs5UlDOTSU5w3lfTDW8lsDY8YjmxYj0qaIqCkiaYookx1tSb4xAhoj8saIHuHR5N0O45G0rW1Nm7amjVrTJq1py6eWbZs0poXGtHlj2v6pZdu+m1PLVvYxc5u5ZsozOvMri5lbWLbbNsLM6CHoVjBzqzIMK+sbYWZZ578nazGzzDG695DnZl5Js5lX0mzmRbbZN6tXgAhSOpvAu8dzNkFwn+eQs+GwgbrlDqxY8IcA5Azw0iFnyW4JOUu+DjkjjcsgZw2Unxy++bd7HjxXGk0fPnxcbYocN2fe9bJ0R1L/ehNME99PpVqHm2WGyWWGyWVGwVD1Stwc07fW4WZ8L8HN4ZfrcbNxoVfCzSo5IlPpEVnG0CJI0SIFjhZR9Fq4WSVnZio9M1OZwa/Q4FcFgx+ZXGXcDI9SjAk5EW5W2V6Awr0AVdgLQLd4ZdysEvyv0kMzleF/hfhfFfA/UrSuxs2qtxegeKq5bC9A4V6AKuwFIF+rh5sxN2VRYl4sItsMULgZoAqbAegcrgedVXJcDqlYYxnuV4j7VQH3I9tqPXRWMpEkU0nZDoDCHQBV2AFQ6ibQGd5OobNSMZhV2eaAws0B1d8coHV9NXTGl6hivpmY4aWZMoI0ndbS9J8FaaQHnU3enhFiobPNi6GzIkMQlEBnigSsO+L6pGa34Pukzq4NRBnWd4NhU70FdNY6XeP/pM59exhe1m17QKl55gKlFpkPlJDhYn2279ILiidxOTcoIcN1btxFrLlS4HvX+0Kp21xh0qNnFXlDwewyekYXq29+uudtqGpavYBpuX+8KTI8TessimKdhVH0zDDnG6VWmXOUWmXeUUh9PfcoJL2wkaW7kLZVQZ7YFS5S8J2SjxR8sNpJCj7ue0kxXrcm+u/ETwoJgSHw38FTivHiVXCVQuLoEDq1Db/MvKUQ2neXQuj7wdPUuAHTlWjSpjZRU5ukqU0JT9tyfEMbaGiTN7Qf2sjk3Qqepsb1mBbD0qawqCksaQor42lbkm8Mg8awvDGsh6dN3u3gaWocnGk5PG0Nj1rDk9bwIp62BfnGcGgMzxvTD0Vh8m4fT1PSDzpBSea8Ps/ozK+sPxgCaznhAU9TE1X6FvA0lhTjW4pMxWvjaXw9/b1ah6cpyfA7pRn3PWS4mZdm3HeK1ECDp8sLQoCazlDoeUnzGQ5Pm1L3QN1iB5Yx+IMBngbYafE0NY7Ofj6eprRZg6cpEuw8nh49PRj96hkA6V9/M3n0bKK/Bs2wNE1df41p7gos9QU2wbTH0pS2a7A0zT185S6+ch9fVK7C0ubR1VjavBdj6eiXa7E0pWoFlsYn8dVDzAgYyvwyMsspgAD9Z98sR5bclXAaXwsYoEn8mJgyImkNgoCmAAKQK1eE0/gouYaIOSKSQjIpBKWQghS6Ck7jo+QyIubwSArNpFCUQgtSmutAanwtuZKIObH2mkxigxKbgkRWgNQmN4XUmBeLYJkIhiJYQQQvQGrMja8nYkasNJ6Vz7F8XihfXIWq8ZWoq4lUksgkCZQkCpLaG6BqfDu9qohZkeA2E9yi4LYgWF4DVsNLgisapMlEmsykSZQmC9JUH1YbtpoR4mB1zGAzSzFVQxCUwmrki90RWB2HiX5rWO2JZW61DBnWyEIxbwWre2/FsJrRDCUaRtttw2rWZLCasQxWhwznW5S/U+eiPPcuynP3ohClRq/adTPgMnIxynOFCQer0WO0g9VMrIPV04sX05NLMJ8g0jmqunenkeaQmqUk3ZC2VoAP3ewMO9Zmhh1rM8OOyetBapa6VsjSXUi7qqirITW+U4LU+GA1pMbHfUht/JRN9N8JnOHBNSf+O8AZSJXgDI+dc+rUNvwwgzO87sMZXr8bOMNJH87wLHBPntGZX1k4w2AO5XUEZ5BfeCtwhmeUWMqbG8EZ3uS/Z2vhDGf5+xlLPmS4Xs8zljxFIt83K8diZOy7+ZlnjKGQ4dGMYQdugubB5yWHP6hGM2j4OzSDhL7bQDNcrkAzSZAa4MYMDZdlY2QDz4SgNHAxcWjuF8Lt6yHeksapa2gmF3BnNURvVIiQhgbUhFA0LgROK4foXgXPJYfmGBExztBAFeTCDA2RBfxyDtGjJrjxGpropvtYOJYtdYXxhBnOG4fm2BBvXgzNBQnwkjzsbKmNGJq7CriPOTS7jMgrHhoSMN7bHppL17iDODTbfzj1Dc28hYzgoSH24r3QobnPaUPhvG3gm2YgpWi1LvIoNoRjmI+DJB5NCIXzYGX0nDSCDedMmMIfxoUnv/in68W8adIwIlA73o820hbCspBiXBFajItD07gm3OaqQgyXpi7IakrBYm4hgo4shOvJwuqYTFb39cRoX1GMlSLlFAMA8aL+0gg4TlOiFNRHlL6KKEX6uVGsm+IHbHlBKa0oZbZ9TbWqEMCm9ElV8ZOqov6yuDbcBbApRQuC+/fFaDeinN2WA0HVhfaSLD6Uy6V9NRDC+nqA69eFwUWLoYxIUw5GlXV6p4u0fztdNKVwTHA/plQyK8e5Yk05u9Qj4CJIKVcWNMTrgoZ4KSoUUJNLVRDlb5cNAaehNLKT01Da253Atvih2rZYQjnokiwHXSoHiAI+UUFvkhdzRUGbUha0qUoTLJxWl+bt8riBtbY0n6ed3uoCTmEKAgkr5vJiCcVvCtvhxewVCxMtxWajTTGX9bVJaSEMF6Wl+Zay4nem5VFDs1HjtJmNBasLvmKBLRfNWTm7GKeO95fDa4X+Gi9Oz09mgM+nyzj4F6/rNPiXeEfBvw7s3o226Ee/1YhWA0rS3DDAVxJ7d4qlaOB2vHQ3faNYsXAca3zhu/JMUFxDr168nl/Aduu00u8J4xIfHsyWp1qFZ1g54WOKfUV8aFR/JVi6wGH1QHon33iCYILMroitC/RD5zsxbszXi+ca6yGS1oBzpNHo7OQH3DbxDfFnCtSFAYvahVHAvqK+2vZqbAidC0FfLy6nr5Z2t2O5ONZKAIGGRzADcLWoFq9c7N3zxRLV0LowYUZ44yKFmSRzscIIRNAmyocFu4nwl8vpGdQhkahSiSqVqFAicxKJxfYXs+XrvgxfZmAWND5kWFwq7hANSOOLJXlLtH6Ws6NX5zPblF3A4IslbIXoHnkyXb4MUXYfjyaxatPmEZJVhWRVIdhC4atiPZMWupTQotGZjiuaZkXTrGiaKa/JW/litlz2lddkxTZZsU3aC/AsvFBfdGLU1D6OGJZh0uMs3YV0iBpm0tTHCIvS4yxton/5QWbmHZgpuvnF5Zt/N9vQUWjf6eKiwqu8+u+DxfJsdjQ7XsTxfInlyx/P7Z7honoyfe3OHhzvNbusn93VdwO3YT4S2MavZnCO8WgKp7YnoSBzTTEifTxZzk/tdlvThFhhh9jXp74u0+Np4NLCVPBVM6wO9IR3+uoEjisaHqKFkWqnqnfrmrDqx6p7tmE/3zh5WKvkYZc+lMnDiT8Bd8T8NG3ijXXPqs2vTrdCXZB+2AhLRgyJDhKE29TEH5ebL3wQ0nbrDGaCJ9PlEe4qd7Oj+SnGkmwcjQl2geoBtHcbhdrmhge12h6HB130QG534cHEn8I3NMQuW9/fJ/6Q3f3mVsOE2fjm9iR9W6v4zU/VcdLVj+KufmS7+oukq/stLOx8F69OoZtfVPbsBfZkfZgubEqdRtqq00hbtY+0ZZbRWvhQWxt7Y+ZzMNhWvDLDk9aU5QYQb0MYLlzrlI+3tbGHi5DyEbfSCU0vBoPCXMzqlSG5vmpC+dIGI9e9Z0d3FOBB+LMzE4Urmk+iOSQISLffY9c10ewTm2jV5ng83nKWVpMy5tJk16R8ucb4mnLHGlAVf9I2didtPgd26fVAgS9sjxfDbNekFl6a7JrUxjMhnubnuPQtni+noRh0gmg0b1uE0Z42vkZLwxEaMOsBmgI2ywR6yiyJw8SSAM6EmfwuQngsuFo/aPR/u8zGP8IcZpOdTXKbNmGfUuvhMLYeXLlo7XFfaJLs0qQJBpWbDoep6XAYTIdRZjpo/RlTGWaGycX57MiExHVV4djEUJc03WXpiXdE4+JxieJcdVFJVW1eTp8DYY3Rrd1KkZBmW7qicYbY8gu5d0jDmjR6FbMR1dL0JDiw4SHilF+oJ72DYkNzOZ5d+LPiezD/dX6EpAdgabJrsvGXG6i7mX1qe/BFbETCmDFGnf4KFlq4tYdZbJGmMdB7DRYRs1gCxr091LHd+cejk1cXuq3Wd6M9m7rwIrwrFisiS2MM+hp4xU7EikUIbly4uFtN4A29mP+Iqlw7xqKxUH13tlk9s5+f6B8+wZN33WvsyXvVbt1WdC0/WSVhtp6YmZsTH1LLrKfSR9QyR0k+clZsa+GZtZ+JBPGBs7KpCLPSqYj1nVII6qNjkZoNpNwVzIfIIpwOIKYG8WGyaKMGpE0iYpUpLK5cYNk1ME25YoFITokvE5JKDFqSRMHytBtXDMSGIHxAmC8H7JGmhfs5rijgkRA64DwJevVscb7Q2OzFPCrNhLlqW18YOkEaiMaXBV4Ym3JAK1sIZIoBcUW4pC0AkshKiyNXrYWTVw1X2x8zzzGZ45jMb0zfbUyvP+IraX/0p622P/q08JGs4v7Im/7KiHmuOwoSwla5/iiiGFXFDglPpBpI5UwUyCCUD1phVY7xqWwPFSyEolrbH9HDS0uBg+bKhUmFcWDbuXJ1TsNYkGQiU+V9EjKRCeYKAsaLHCjpy4EXWjVgxJfTljukObHVnRvvGdviWtcnXXEt9klXlCz1SRl3wrFLugKk7ZOuCIQwoKCL+eKsT3t5rQt/8+fXs7m3p/dD5Kc6iwRlw3yl6ZGPBGXhd0jXNrRRltH5DFtJUfCj9N3ZzVlZes5fT7wihHrmVbUFs6FK40sRUqcBpmxG5zOsaS9WOmsS1/PVJHJXTWKFp6bw64griw5XipRP44pF23fbON+7wD3Gz6mduERKwRTAwBR9Aqa40scpvmGFhV4tWCIsJWMK4GKKPhVTlJiYmGnLl6F8npSfkjEFcDFFn4opRDYsnQgRRKiNOBqVCCJSFqYAEqboczBFKY6GCe+E5SODKASDakP5KdlSANdS9KmWIrreL4SPzmTIOforDvUvPbdx7D2zEEKiIE0myFEz7KJ3TcikbYTp9xx+vs6I29XL2eUchxiOqtyW83vIIQKTkGkEJiF9BKYobcIurbf9lqtsvw3Lz9n3rles0CzdZWkTkskPXeQsuaZ7Tquux+iZGZMhXlC6XqfJELqJhshNCV92lSjoIqb8LAJxlsZwTxt7ujPcR0jZUh/0aWPbfKBqermcvn7z0wXsKS7LkNsC7FmYT489YZXsWqM5HAoYdpXWxV41OqjuVwfXnabTgE1O3Q4vgX67XL+pk9E02bXptmXbFkoNqmwzVbaZKs0WTUvv4dhyqpT93rgS+brW2ZDCHSjoGTTrwD+LQzhFW12WzgvvjjMVyLSLpclOpl1M+i7m2y2zLpSlbfimuM3SBHUZ4/cNNefX4nnK9E5Lmuxk6ogewxSt53jiKwWKpxTrGJ74tEfwNIGLJjLx/iKD8xeJRHg7UWO+bIQa0IZIPZfSKPiRpXbKdlv/KCV2Suv2BWjRdndZXhW07zpXB2XfxYvMPLxkaRtOycTuiTy7yOCoy8xD7nvOzPfcxG09JyMLoJelbTilfao1QeU2irRi1G3Fz1OZ+xd1o/h5KnMGo9bGz1MZF1U16WmPyrim6JhiPH4Cg2DtbqkfICqN25kmO5U6cDZBjWB42dHsgwXZz5GlbYSiTb3QS/1BxLb9MFs7m70j9r/RU952IbdRW1sx3V9DoR38zHZ6MPGJdKX2vt38HQH1u28Qb/rfr2ilpeqszmR14RhgBycMF7dpjXvXfZWyzcdZeuTDAsk6jSzk3u8FI5E/Y/GyQXd9XB+dkXBlR40c4kZewpAVQzxQg7l/aELx7TMg0mLPoUCUxX2bRj/FffFRUw9xwxz2FYYmyrXuZ8POOE/xzFn9DxN617NZ9wmQWbE0LcEEZR9xOjThqccU6KpQu32dZyJba8NniGaKZapCLfnQhKrTq/QQlti3Za8mNNJxymsNXNUufiAGjBH3ZFJinl6f3LqSqlrkPZX5LeQG1NQycaa9mkXqaKhFDmGTFsAGDNih5G2JoI7dWXy1LTahLaqmLXKmynQ1WSxXlminqkh4UsX2KnoDkmfxO6hCI67FVDJnjjFHSaiUosR/JkUpkJL87+5VTizJWUl274KmOxWseISXH32ly1G6/ERRo0Pg8/72h7IbUk7ehmX8jIStP/F7WJjjqnIO68v8HC0ai46zG2odhocwC88EiRKLeQW2/ckPYDJEFvQ5HssugJwExKuT17OzKSJCO2u3tip0ECrT9vZPs+tewWX62AetaoLLdIck8EIjYYHVhHKolyPfRg7ryZGxHOXlNF6Oehs5cRAuZ1TEggxvCSUxL4nU1xdV9DbvI6kloogXxYMocoUocwsHwNmVnvqcVJJIpV6qCFLptaSudG8TvBk4meHyZXK1EugWviaNr0kbatJctybrbiG5WjTlWuBmtqsFs7UIHctklc1aH5HBl5yYk7/sfom+1lzh3BYOWdwTpBywUYEjZYDLWAx10bBha1mDpL2aZBaFA48OcPfjdGq3keGYmPtNY3M/aDi+4taOthtW2D0ti6yb2OwhaoXZExk3ByuNmwfXMm7Kq9dv/vT/Adp7w6Q=
"""


def _carregar_template_canteiro_embutido():
    """Carrega o template interno de canteiro, sem acessar planilha externa."""
    dados = _zlib_canteiro.decompress(_base64_canteiro.b64decode(_CANTEIRO_TEMPLATE_B64.strip()))
    return _json_canteiro.loads(dados.decode("utf-8"))


def _criar_side_canteiro(dados_side):
    if not dados_side:
        return Side()
    return Side(style=dados_side.get("style"), color=dados_side.get("color"))


def _aplicar_estilo_canteiro(cell, estilo):
    """Aplica estilo serializado do modelo cant ao cell de destino."""
    if not estilo:
        return
    font = estilo.get("font")
    if font:
        kwargs = {k: v for k, v in font.items() if k != "sz"}
        if "sz" in font:
            kwargs["size"] = font.get("sz")
        cell.font = Font(**kwargs)

    fill = estilo.get("fill")
    if fill:
        # Correção V16:
        # O modelo embutido pode armazenar preenchimentos "automáticos" do Excel como
        # fgColor/bgColor = 00000000 ou apenas fill_type="solid" sem cor explícita.
        # Ao recriar isso com openpyxl, o Excel interpreta como preenchimento preto.
        # Portanto, só aplicamos preenchimento quando há cor real e visível.
        fg = fill.get("fgColor")
        bg = fill.get("bgColor")
        fill_type = fill.get("fill_type")

        def _cor_visivel(cor):
            if not cor:
                return False
            cor_txt = str(cor).upper()
            return cor_txt not in {"00000000", "000000", "NONE", "NULL"}

        if fill_type and _cor_visivel(fg):
            cell.fill = PatternFill(fill_type=fill_type, fgColor=fg)
        elif fill_type and _cor_visivel(bg):
            cell.fill = PatternFill(fill_type=fill_type, bgColor=bg)
        # Caso contrário, mantém o preenchimento padrão da célula.

    border = estilo.get("border")
    if border:
        cell.border = Border(
            left=_criar_side_canteiro(border.get("left")),
            right=_criar_side_canteiro(border.get("right")),
            top=_criar_side_canteiro(border.get("top")),
            bottom=_criar_side_canteiro(border.get("bottom")),
            diagonal=_criar_side_canteiro(border.get("diagonal")),
            vertical=_criar_side_canteiro(border.get("vertical")),
            horizontal=_criar_side_canteiro(border.get("horizontal")),
        )

    alignment = estilo.get("alignment")
    if alignment:
        # openpyxl usa textRotation/shrinkToFit; os nomes abaixo também aceitam snake_case em versões recentes.
        kwargs = {}
        mapa = {
            "horizontal": "horizontal",
            "vertical": "vertical",
            "text_rotation": "textRotation",
            "wrap_text": "wrap_text",
            "shrink_to_fit": "shrink_to_fit",
            "indent": "indent",
        }
        for k, v in alignment.items():
            kwargs[mapa.get(k, k)] = v
        cell.alignment = Alignment(**kwargs)

    if estilo.get("number_format"):
        cell.number_format = estilo.get("number_format")


def remover_abas_canteiro_existentes(wb):
    """Remove abas antigas de canteiro antes de recriar o modelo embutido."""
    abas = set(ABAS_CANTEIRO_MODELO) | {"GUIA_CANTEIRO_SICRO"}
    for nome in list(wb.sheetnames):
        if nome in abas:
            wb.remove(wb[nome])


def aplicar_template_canteiro_embutido(wb_destino):
    """
    Recria as abas de canteiro a partir da estrutura embutida no código.

    Mantém:
        - fórmulas internas entre abas CANT_*;
        - formatação da planilha modelo;
        - larguras, alturas e mesclagens;
        - valores fixos e campos manuais.

    Observação:
        Fórmulas que dependiam de arquivos externos do modelo original foram substituídas
        pelo valor armazenado no próprio modelo para evitar vínculos quebrados.
    """
    remover_abas_canteiro_existentes(wb_destino)
    template = _carregar_template_canteiro_embutido()
    estilos = template.get("styles", [])

    for sheet_data in template.get("sheets", []):
        nome = sheet_data["name"]
        ws = wb_destino.create_sheet(nome)

        # Valores, fórmulas e estilos.
        for coord, valor, style_id in sheet_data.get("cells", []):
            cell = ws[coord]
            cell.value = valor
            if style_id is not None and style_id < len(estilos):
                _aplicar_estilo_canteiro(cell, estilos[style_id])

        # Mesclagens.
        for faixa in sheet_data.get("merged", []):
            try:
                ws.merge_cells(faixa)
            except Exception:
                pass

        # Largura das colunas.
        for col, cfg in sheet_data.get("col_widths", {}).items():
            if cfg.get("width") is not None:
                ws.column_dimensions[col].width = cfg.get("width")
            if cfg.get("hidden"):
                ws.column_dimensions[col].hidden = True

        # Altura das linhas.
        for row_txt, cfg in sheet_data.get("row_heights", {}).items():
            row_idx = int(row_txt)
            if cfg.get("height") is not None:
                ws.row_dimensions[row_idx].height = cfg.get("height")
            if cfg.get("hidden"):
                ws.row_dimensions[row_idx].hidden = True

        # Congelamento e apresentação.
        if sheet_data.get("freeze_panes"):
            ws.freeze_panes = sheet_data.get("freeze_panes")
        ocultar_linhas_grade(ws)

    criar_guia_canteiro_sicro(wb_destino)
    logger.info("Módulo V13.5 de Canteiro criado com estrutura embutida do modelo cant, sem ler planilha externa.")
    return wb_destino


def criar_modelo_canteiro(wb):
    """
    Cria o módulo de canteiro de obras manual/orientado.

    Regra V13.5:
        - não extrair nem copiar o arquivo cant em tempo de execução;
        - recriar diretamente no nosso workbook a estrutura de cálculo e tabelas do modelo;
        - manter a aba GUIA_CANTEIRO_SICRO com os locais de consulta no Manual SICRO V06.
    """
    return aplicar_template_canteiro_embutido(wb)



# ============================================================
# V16 - INTERFACE COM CANTEIRO PRESERVADO
# ============================================================
# Esta seção sobrescreve funções visuais da V14 sem alterar o motor de cálculo.
# Objetivo:
#   - melhorar a navegação do arquivo Excel;
#   - padronizar cores;
#   - destacar campos manuais, resultados e alertas;
#   - proteger visualmente o usuário contra edição acidental de fórmulas;
#   - manter abas técnicas ocultas.

COR_TITULO = "1F4E78"      # azul escuro
COR_HEADER = "D9EAF7"      # azul claro/cinza
COR_INPUT = "FFF2CC"       # amarelo: preencher manualmente
COR_RESULTADO = "E2F0D9"   # verde: resultado/fórmula
COR_ALERTA = "F4CCCC"      # vermelho claro: alerta
COR_NEUTRO = "F2F2F2"      # cinza claro
COR_BRANCO = "FFFFFF"

# Abas que vêm do modelo de canteiro e devem preservar a estrutura visual original.
# A interface global não deve inserir legenda, filtros ou proteção nelas.
ABAS_CANTEIRO_MODELO_PRESERVAR = {
    "CANT_Resumo",
    "CANT_Princ",
    "CANT_Princ Container",
    "CANT_Industrial",
    "CANT_Complementar",
}


def _borda_fina_v15():
    thin = Side(style="thin", color="D9D9D9")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill_hex(cell):
    """Retorna a cor RGB do preenchimento da célula, quando existir."""
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == "rgb" and fg.rgb:
            return str(fg.rgb).replace("FF", "", 1).upper()
    except Exception:
        pass
    return ""


def _aplicar_larguras(ws, larguras):
    for col, largura in larguras.items():
        try:
            ws.column_dimensions[col].width = largura
        except Exception:
            pass


def _estilizar_cabecalho(ws, linha_header=1, max_col=None):
    border = _borda_fina_v15()
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(linha_header, col)
        if cell.value is not None and str(cell.value).strip() != "":
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=COR_HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border


def adicionar_legenda_visual(ws, linha=1, coluna=1):
    """Insere legenda visual padronizada sem interferir em fórmulas existentes."""
    labels = [
        ("Campo manual", COR_INPUT),
        ("Resultado/Fórmula", COR_RESULTADO),
        ("Alerta/Pendência", COR_ALERTA),
        ("Cabeçalho", COR_HEADER),
    ]
    try:
        ws.cell(linha, coluna, "LEGENDA")
        ws.cell(linha, coluna).font = Font(bold=True, color=COR_BRANCO)
        ws.cell(linha, coluna).fill = PatternFill("solid", fgColor=COR_TITULO)
        for i, (texto, cor) in enumerate(labels, start=1):
            c = ws.cell(linha, coluna + i, texto)
            c.fill = PatternFill("solid", fgColor=cor)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _borda_fina_v15()
    except Exception:
        pass


def proteger_formulas_e_liberar_inputs(ws):
    """
    Protege fórmulas contra edição acidental e libera campos amarelos.
    A proteção não usa senha e serve apenas como barreira visual/operacional leve.
    """
    try:
        from openpyxl.styles import Protection
        for row in ws.iter_rows():
            for cell in row:
                cor = _fill_hex(cell)
                valor = cell.value
                if cor == COR_INPUT:
                    cell.protection = Protection(locked=False)
                elif isinstance(valor, str) and valor.startswith("="):
                    cell.protection = Protection(locked=True)
                    if not cell.fill or cell.fill.fill_type is None:
                        cell.fill = PatternFill("solid", fgColor=COR_RESULTADO)
                else:
                    cell.protection = Protection(locked=False)
        ws.protection.sheet = True
        ws.protection.enable()
    except Exception:
        pass


def criar_aba_resumo(wb):
    """Cria a aba inicial do sistema com navegação e indicadores principais."""
    if "RESUMO" in wb.sheetnames:
        wb.remove(wb["RESUMO"])
    ws = wb.create_sheet("RESUMO", 0)

    border = _borda_fina_v15()

    ws["A1"] = "EXTRATOR SICRO - PAINEL INICIAL"
    ws["A1"].font = Font(bold=True, size=18, color=COR_BRANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.merge_cells("A1:H1")

    ws["A2"] = "Versão"
    ws["B2"] = "V16 - Interface com Canteiro preservado"
    ws["D2"] = "Status do orçamento"
    ws["E2"] = '=IFERROR(CHECK_ORCAMENTO!B2,"GERAR CHECK_ORCAMENTO")'
    for cell in ["A2", "D2"]:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws["E2"].fill = PatternFill("solid", fgColor=COR_RESULTADO)
    ws["E2"].font = Font(bold=True)

    indicadores = [
        ("Serviços informados", '=COUNTA(QUANTIDADES!A2:A10000)'),
        ("Códigos não encontrados", '=MAX(COUNTA(ALERTAS!A4:A10000)-COUNTIF(ALERTAS!A4:A10000,"SEM_CODIGOS_NAO_ENCONTRADOS"),0)'),
        ("Funções de mão de obra", '=MAX(COUNTA(PESSOAS!A5:A10000)-1,0)'),
        ("Horas totais de mão de obra", '=IFERROR(INDEX(PESSOAS!F:F,MATCH("TOTAL",PESSOAS!A:A,0)),0)'),
        ("Colaboradores/mês", '=IFERROR(INDEX(PESSOAS!H:H,MATCH("TOTAL",PESSOAS!A:A,0)),0)'),
        ("Equipamentos consolidados", '=COUNTA(EQUIPAMENTOS!A2:A10000)'),
        ("Itens DMT", '=COUNTA(DMT!A2:A10000)'),
        ("Pendências no check", '=IFERROR(COUNTIF(CHECK_ORCAMENTO!A6:A200,"PENDENTE"),0)'),
    ]

    ws["A4"] = "INDICADORES"
    ws["A4"].font = Font(bold=True, color=COR_BRANCO)
    ws["A4"].fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.merge_cells("A4:C4")

    ws["A5"] = "Indicador"
    ws["B5"] = "Valor"
    ws["C5"] = "Observação"
    _estilizar_cabecalho(ws, linha_header=5, max_col=3)

    for i, (nome, formula) in enumerate(indicadores, start=6):
        ws.cell(i, 1, nome)
        ws.cell(i, 2, formula)
        ws.cell(i, 3, "Atualizado por fórmula")
        ws.cell(i, 2).fill = PatternFill("solid", fgColor=COR_RESULTADO)
        for col in range(1, 4):
            ws.cell(i, col).border = border
            ws.cell(i, col).alignment = Alignment(vertical="center", wrap_text=True)

    ws["E4"] = "NAVEGAÇÃO RÁPIDA"
    ws["E4"].font = Font(bold=True, color=COR_BRANCO)
    ws["E4"].fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.merge_cells("E4:H4")

    navegacao = [
        ("1. Conferir pendências", "CHECK_ORCAMENTO", "Abrir painel de conferência"),
        ("2. Validar sintético", "SINTETICO", "Itens extraídos do relatório sintético"),
        ("3. Validar analítico", "ANALITICO", "Composições analíticas copiadas"),
        ("4. Preencher quantidades/meses", "QUANTIDADES", "Campos manuais principais"),
        ("5. Conferir mão de obra", "PESSOAS", "Resumo de colaboradores"),
        ("6. Histograma de mão de obra", "HISTOGRAMA_MO", "Efetivo mensal"),
        ("7. Equipamentos", "EQUIPAMENTOS", "Equipamentos consolidados"),
        ("8. DMT", "DMT", "Transporte e distâncias"),
        ("9. ADM Local", "01-ADM-Resumo", "Modelo manual orientado"),
        ("10. Mobilização", "MOB_RESUMO", "Resumo MOB/DESMOB"),
        ("11. Canteiro", "CANT_Resumo", "Resumo de canteiro"),
        ("12. Alertas", "ALERTAS", "Códigos não encontrados"),
    ]

    ws["E5"] = "Ação"
    ws["F5"] = "Aba"
    ws["G5"] = "Descrição"
    _estilizar_cabecalho(ws, linha_header=5, max_col=7)

    for i, (acao, aba, desc) in enumerate(navegacao, start=6):
        ws.cell(i, 5, acao)
        ws.cell(i, 6, aba)
        ws.cell(i, 7, desc)
        if aba in wb.sheetnames:
            ws.cell(i, 6).hyperlink = f"#'{aba}'!A1"
            ws.cell(i, 6).style = "Hyperlink"
        for col in range(5, 8):
            ws.cell(i, col).border = border
            ws.cell(i, col).alignment = Alignment(vertical="center", wrap_text=True)

    linha_legenda = 20
    adicionar_legenda_visual(ws, linha=linha_legenda, coluna=1)

    ws["A22"] = "FLUXO DE USO RECOMENDADO"
    ws["A22"].font = Font(bold=True, color=COR_BRANCO)
    ws["A22"].fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.merge_cells("A22:H22")
    instrucoes = [
        "1. Validar SINTETICO e ANALITICO.",
        "2. Preencher QUANTIDADES e MESES.",
        "3. Conferir PESSOAS, HISTOGRAMA_MO, EQUIPAMENTOS e DMT.",
        "4. Preencher manualmente ADM Local, MOB/DESMOB e CANTEIRO.",
        "5. Voltar ao CHECK_ORCAMENTO antes de entregar o orçamento.",
    ]
    for i, txt in enumerate(instrucoes, start=23):
        ws.cell(i, 1, txt)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="center")

    _aplicar_larguras(ws, {"A": 34, "B": 18, "C": 30, "D": 5, "E": 30, "F": 26, "G": 55, "H": 15})
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = COR_TITULO
    return ws


def criar_aba_check_orcamento(wb):
    """Cria um painel de conferência visual, com contadores e checklist do orçamento."""
    if "CHECK_ORCAMENTO" in wb.sheetnames:
        wb.remove(wb["CHECK_ORCAMENTO"])

    ws = wb.create_sheet("CHECK_ORCAMENTO")
    border = _borda_fina_v15()

    ws["A1"] = "CHECK ORÇAMENTO - PAINEL DE CONFERÊNCIA"
    ws["A1"].font = Font(bold=True, size=16, color=COR_BRANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COR_TITULO)
    ws.merge_cells("A1:H1")

    ws["A2"] = "STATUS GERAL"
    ws["B2"] = '=IF(COUNTIF(A9:A250,"PENDENTE")>0,"PENDÊNCIAS ENCONTRADAS","APTO PARA REVISÃO")'
    ws["D2"] = "Pendências críticas"
    ws["E2"] = '=COUNTIFS(A9:A250,"PENDENTE",H9:H250,"ALTA")'
    ws["F2"] = "Pendências médias"
    ws["G2"] = '=COUNTIFS(A9:A250,"PENDENTE",H9:H250,"MÉDIA")'
    ws["H2"] = '=COUNTIFS(A9:A250,"PENDENTE",H9:H250,"BAIXA")'

    resumo_cells = ["A2", "B2", "D2", "E2", "F2", "G2", "H2"]
    for c in resumo_cells:
        ws[c].font = Font(bold=True)
        ws[c].border = border
        ws[c].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A2"].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws["B2"].fill = PatternFill("solid", fgColor=COR_RESULTADO)
    ws["D2"].fill = PatternFill("solid", fgColor=COR_ALERTA)
    ws["E2"].fill = PatternFill("solid", fgColor=COR_ALERTA)
    ws["F2"].fill = PatternFill("solid", fgColor=COR_INPUT)
    ws["G2"].fill = PatternFill("solid", fgColor=COR_INPUT)
    ws["H2"].fill = PatternFill("solid", fgColor=COR_HEADER)

    ws["A4"] = "Observação"
    ws["B4"] = "Esta aba não substitui a revisão técnica. Ela apenas aponta campos vazios, totais zerados e pontos que precisam de conferência manual."
    ws["A4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws.merge_cells("B4:H4")
    ws["B4"].alignment = Alignment(wrap_text=True, vertical="center")

    adicionar_legenda_visual(ws, linha=6, coluna=1)

    headers = ["STATUS", "MÓDULO", "VERIFICAÇÃO", "ABA", "PENDÊNCIAS", "AÇÃO NECESSÁRIA", "NAVEGAR", "CRITICIDADE"]
    linha_header = 8
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(linha_header, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    checks = [
        ("Extração SICRO", "Códigos não encontrados", "ALERTAS", '=MAX(COUNTA(ALERTAS!A4:A10000)-COUNTIF(ALERTAS!A4:A10000,"SEM_CODIGOS_NAO_ENCONTRADOS"),0)', "Verificar referência SICRO, mês/ano, estado, código digitado ou item manual.", "ALTA"),
        ("Quantidades", "Serviços sem quantidade", "QUANTIDADES", '=COUNTIFS(QUANTIDADES!A2:A10000,"<>",QUANTIDADES!D2:D10000,"")', "Preencher ou validar a quantidade dos serviços.", "ALTA"),
        ("Quantidades", "Serviços sem meses", "QUANTIDADES", '=COUNTIFS(QUANTIDADES!A2:A10000,"<>",QUANTIDADES!E2:E10000,"")', "Preencher meses por serviço ou usar o prazo padrão no HISTOGRAMA_MO.", "MÉDIA"),
        ("Mão de obra", "Linhas sem quantidade na base técnica", "MAO_OBRA_CALCULO", '=COUNTIF(MAO_OBRA_CALCULO!P:P,"Informar quantidade")', "Revisar a aba QUANTIDADES para os serviços com mão de obra calculada.", "ALTA"),
        ("Mão de obra", "Linhas sem meses na base técnica", "MAO_OBRA_CALCULO", '=COUNTIF(MAO_OBRA_CALCULO!P:P,"Informar meses")', "Preencher meses na aba QUANTIDADES ou prazo padrão no HISTOGRAMA_MO.", "MÉDIA"),
        ("Mão de obra", "Colaboradores/mês zerado", "PESSOAS", '=IFERROR(IF(INDEX(PESSOAS!H:H,MATCH("TOTAL",PESSOAS!A:A,0))=0,1,0),1)', "Conferir quantidades, meses e extração de mão de obra.", "ALTA"),
        ("DMT", "Itens DMT extraídos para revisão", "DMT", '=COUNTA(DMT!A2:A10000)', "Conferir se há itens de transporte que exigem distância, fonte ou decisão manual.", "INFORMATIVO"),
        ("Mobilização Equipamentos", "Equipamentos sem quantidade", "MOB_EQUIP", '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!C3:C1000,"")', "Preencher quantidade apenas para os equipamentos que serão mobilizados.", "MÉDIA"),
        ("Mobilização Equipamentos", "Equipamentos sem FU", "MOB_EQUIP", '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!F3:F1000,"")', "Preencher FU quando o equipamento entrar no cálculo de mobilização.", "MÉDIA"),
        ("Mobilização Equipamentos", "Equipamentos sem K", "MOB_EQUIP", '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!G3:G1000,"")', "Preencher fator K quando aplicável.", "MÉDIA"),
        ("Mobilização Equipamentos", "Equipamentos sem distância", "MOB_EQUIP", '=COUNTIFS(MOB_EQUIP!B3:B1000,"<>",MOB_EQUIP!B3:B1000,"<>TOTAL",MOB_EQUIP!I3:I1000,"")', "Preencher distância de mobilização/desmobilização quando aplicável.", "MÉDIA"),
        ("Mobilização Pessoas", "Pessoas sem quantidade", "MOB_PESSOAS", '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!C4:C103,"")', "Preencher apenas as pessoas que serão mobilizadas.", "MÉDIA"),
        ("Mobilização Pessoas", "Pessoas sem custo de passagem", "MOB_PESSOAS", '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!D4:D103,"")', "Informar custo unitário de passagem quando aplicável.", "BAIXA"),
        ("Mobilização Pessoas", "Pessoas sem custo de alimentação", "MOB_PESSOAS", '=COUNTIFS(MOB_PESSOAS!B4:B103,"<>",MOB_PESSOAS!B4:B103,"<>TOTAL",MOB_PESSOAS!E4:E103,"")', "Informar custo unitário de alimentação quando aplicável.", "BAIXA"),
        ("Administração Local", "Resumo de ADM Local zerado ou ausente", "01-ADM-Resumo", '=IFERROR(IF(SUM(\'01-ADM-Resumo\'!A1:Z500)=0,1,0),1)', "Conferir se a Administração Local foi preenchida conforme o modelo e manual.", "MÉDIA"),
        ("Canteiro", "Resumo de canteiro zerado ou ausente", "CANT_Resumo", '=IFERROR(IF(SUM(CANT_Resumo!A1:Z500)=0,1,0),1)', "Conferir as abas CANT_Princ, CANT_Industrial, CANT_Complementar e CANT_Resumo.", "MÉDIA"),
        ("Canteiro", "Guia de canteiro ausente", "GUIA_CANTEIRO_SICRO", '=IFERROR(IF(COUNTA(GUIA_CANTEIRO_SICRO!A:A)=0,1,0),1)', "Verificar a criação da aba de orientação do Manual SICRO V06.", "BAIXA"),
    ]

    for row_idx, (modulo, verificacao, aba, formula, acao, criticidade) in enumerate(checks, start=9):
        ws.cell(row_idx, 1, f'=IF(OR(E{row_idx}=0,H{row_idx}="INFORMATIVO"),"OK","PENDENTE")')
        ws.cell(row_idx, 2, modulo)
        ws.cell(row_idx, 3, verificacao)
        ws.cell(row_idx, 4, aba)
        ws.cell(row_idx, 5, formula)
        ws.cell(row_idx, 6, acao)
        ws.cell(row_idx, 7, "Abrir aba")
        ws.cell(row_idx, 8, criticidade)
        if aba in wb.sheetnames:
            ws.cell(row_idx, 7).hyperlink = f"#'{aba}'!A1"
            ws.cell(row_idx, 7).style = "Hyperlink"
        for col in range(1, 9):
            cell = ws.cell(row_idx, col)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row_idx, 5).fill = PatternFill("solid", fgColor=COR_RESULTADO)

    linha_obs = 9 + len(checks) + 2
    ws.cell(linha_obs, 1, "ORIENTAÇÃO")
    ws.cell(linha_obs, 1).font = Font(bold=True)
    ws.cell(linha_obs, 1).fill = PatternFill("solid", fgColor=COR_HEADER)
    ws.cell(linha_obs, 2, "Itens informativos, como DMT extraído, não significam erro; indicam apenas necessidade de conferência técnica.")
    ws.merge_cells(start_row=linha_obs, start_column=2, end_row=linha_obs, end_column=8)
    ws.cell(linha_obs, 2).alignment = Alignment(wrap_text=True, vertical="center")

    _aplicar_larguras(ws, {"A": 16, "B": 30, "C": 42, "D": 24, "E": 16, "F": 70, "G": 18, "H": 16})
    aplicar_filtro_e_congelamento(ws, linha_header=linha_header, freeze="A9")
    ocultar_linhas_grade(ws)
    ws.sheet_properties.tabColor = "C00000"
    return ws


def aplicar_padrao_visual_frontend(wb):
    """Aplica padronização visual geral sem alterar cálculos."""
    tab_colors = {
        "RESUMO": COR_TITULO,
        "CHECK_ORCAMENTO": "C00000",
        "SINTETICO": "70AD47",
        "ANALITICO": "70AD47",
        "QUANTIDADES": "FFC000",
        "PESSOAS": "5B9BD5",
        "HISTOGRAMA_MO": "5B9BD5",
        "HISTOGRAMA_EQUIP": "5B9BD5",
        "EQUIPAMENTOS": "A9D18E",
        "DMT": "A9D18E",
        "ALERTAS": "C00000",
        "MOB_RESUMO": "7030A0",
        "MOB_PESSOAS": "7030A0",
        "MOB_EQUIP": "7030A0",
        "CANT_Resumo": "9E480E",
        "CANT_Princ": "9E480E",
        "CANT_Princ Container": "9E480E",
        "CANT_Industrial": "9E480E",
        "CANT_Complementar": "9E480E",
        "GUIA_CANTEIRO_SICRO": "9E480E",
    }

    for nome, cor in tab_colors.items():
        if nome in wb.sheetnames:
            try:
                wb[nome].sheet_properties.tabColor = cor
            except Exception:
                pass

    # Legendas nas abas manuais/orientadas. Usa linhas distantes quando necessário para não sobrescrever estrutura.
    for nome in ["MOB_RESUMO", "MOB_PESSOAS", "MOB_EQUIP", "GUIA_CANTEIRO_SICRO"]:
        if nome in wb.sheetnames:
            adicionar_legenda_visual(wb[nome], linha=1, coluna=1)

    # As abas CANT_* preservam integralmente a aparência do modelo original.
    # Não adicionamos legenda nem proteção nelas para não sobrescrever a formatação do canteiro.
    # A orientação fica concentrada na aba GUIA_CANTEIRO_SICRO.

    # Aplica proteção leve em abas manuais principais.
    for nome in ["QUANTIDADES", "PESSOAS", "HISTOGRAMA_MO", "MOB_RESUMO", "MOB_PESSOAS", "MOB_EQUIP", "CHECK_ORCAMENTO"]:
        if nome in wb.sheetnames:
            proteger_formulas_e_liberar_inputs(wb[nome])


def aplicar_limpeza_visual_final(wb):
    """
    Reordena abas, oculta bases técnicas, aplica filtros/congelamentos e melhora a usabilidade visual.
    """
    abas_ocultas = {
        "MAO_OBRA_CALCULO",
        "ATIVIDADES_AUXILIARES",
        "EQUIPAMENTOS_POR_COMPOSICAO",
    }

    for nome in abas_ocultas:
        if nome in wb.sheetnames:
            wb[nome].sheet_state = "hidden"

    abas_visiveis = [
        "RESUMO", "CHECK_ORCAMENTO", "SINTETICO", "ANALITICO", "QUANTIDADES", "PESSOAS",
        "HISTOGRAMA_MO", "HISTOGRAMA_EQUIP", "EQUIPAMENTOS", "DMT", "ALERTAS",
        "01-ADM-Resumo", "02-ADM-Fixa", "03-ADM-Vinculada", "04-ADM-Variavel", "05-ADM-Variavel",
        "06-Manutenção Canteiro", "07-Acordão", "05-ADM-Manutencao-Canteiro", "06-ADM-Custos-Diversos", "07-GUIA_DNIT_AL",
        "MOB_RESUMO", "MOB_PESSOAS", "MOB_EQUIP",
        "CANT_Resumo", "CANT_Princ", "CANT_Princ Container", "CANT_Industrial", "CANT_Complementar", "GUIA_CANTEIRO_SICRO",
    ]

    for nome in abas_visiveis:
        if nome in wb.sheetnames:
            # As abas de canteiro devem permanecer iguais ao modelo base.
            # Não aplicar filtros/congelamentos genéricos nem limpeza visual nelas.
            if nome in ABAS_CANTEIRO_MODELO_PRESERVAR:
                continue

            ocultar_linhas_grade(wb[nome])
            if nome == "PESSOAS":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=4, freeze="A5")
            elif nome == "HISTOGRAMA_MO":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=6, freeze="A7")
            elif nome == "HISTOGRAMA_EQUIP":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=4, freeze="A5")
            elif nome == "ALERTAS":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=3, freeze="A4")
            elif nome == "CHECK_ORCAMENTO":
                aplicar_filtro_e_congelamento(wb[nome], linha_header=8, freeze="A9")
            else:
                aplicar_filtro_e_congelamento(wb[nome], linha_header=1, freeze="A2")

    ordem = [
        "RESUMO", "CHECK_ORCAMENTO",
        "SINTETICO", "ANALITICO", "QUANTIDADES",
        "PESSOAS", "HISTOGRAMA_MO", "HISTOGRAMA_EQUIP",
        "EQUIPAMENTOS", "DMT", "ALERTAS",
        "01-ADM-Resumo", "02-ADM-Fixa", "03-ADM-Vinculada", "04-ADM-Variavel", "05-ADM-Variavel", "06-Manutenção Canteiro", "07-Acordão",
        "MOB_RESUMO", "MOB_PESSOAS", "MOB_EQUIP",
        "CANT_Resumo", "CANT_Princ", "CANT_Princ Container", "CANT_Industrial", "CANT_Complementar", "GUIA_CANTEIRO_SICRO",
        "MAO_OBRA_CALCULO", "ATIVIDADES_AUXILIARES", "EQUIPAMENTOS_POR_COMPOSICAO",
    ]
    for pos, nome in enumerate(ordem):
        if nome in wb.sheetnames:
            ws = wb[nome]
            wb._sheets.remove(ws)
            wb._sheets.insert(min(pos, len(wb._sheets)), ws)

    aplicar_padrao_visual_frontend(wb)


if __name__ == "__main__":
    arquivo_sintetico = r"C:\Users\luc.silva\Documents\SEAPROJ\Valores SICRO - Jan 26\MA 01-2026 Relatório Sintético de Composições de Custos.xlsx"
    arquivo_analitico = r"C:\Users\luc.silva\Documents\SEAPROJ\Valores SICRO - Jan 26\MA 01-2026 Relatório Analítico de Composições de Custos.xlsx"
    arquivo_saida = rf"C:\Users\luc.silva\Documents\SEAPROJ\Relatorio Extração Sicro/extracao_sicro_completa_{data_hoje}.xlsx"
    codigos_desejados = ler_codigos_digitados()
    gerar_arquivo_final(
        arquivo_sintetico=arquivo_sintetico,
        arquivo_analitico=arquivo_analitico,
        arquivo_saida=arquivo_saida,
        codigos_desejados=codigos_desejados
    )
